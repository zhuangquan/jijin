import argparse
import datetime as dt
import json
import os
import re
import subprocess
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import requests
import yaml


EASTMONEY_LSJZ_URL = "https://api.fund.eastmoney.com/f10/lsjz"
TX_HEADERS = ["date", "fund_code", "action", "shares", "nav", "note"]
SYSTEM_SHEETS = {"summary", "positions", "汇总", "持仓明细"}
TEMPLATE_SHEET_NAME = "模板"
META_NAME_LABEL = "基金名称"
META_CODE_LABEL = "基金代码"
META_NAV_LABEL = "最新净值"
BASE_DATA_HEADERS = ["日期", "份额", "购买净值", "备注"]
DATA_HEADERS = BASE_DATA_HEADERS + ["最新净值", "盈利利率", "盈利额"]
DATA_HEADER_ROW = 5
DATA_START_ROW = DATA_HEADER_ROW + 1

DEFAULT_CONFIG = {
    "app": {
        "xlsx_path": "data/jijin.xlsx",
        "threshold": 0.40,
        "ntfy_url": "",
        "github_repo": "",
        "github_token": "",
        "repo_keepalive": True,
        "repo_keepalive_file": "data/.repo_activity_keepalive.txt",
        "repo_keepalive_branch": "",
        "repo_keepalive_interval_days": 30,
        "repo_inactivity_notify": True,
        "repo_inactive_warn_days": 60,
    },
    "funds": [],
}


def deep_merge(base: Dict, patch: Dict) -> Dict:
    result = dict(base)
    for key, value in (patch or {}).items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = deep_merge(result[key], value)
        else:
            result[key] = value
    return result


def normalize_fund_code(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if s.isdigit():
        s = s.zfill(6)
    if re.fullmatch(r"\d{6}", s):
        return s
    return ""


def to_float(value, default: Optional[float] = None) -> Optional[float]:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def to_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "no", "n", "off"}:
        return False
    return default


def parse_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def row_headers(ws, row_idx: int, width: int) -> List[str]:
    return [str(ws.cell(row=row_idx, column=i + 1).value or "").strip() for i in range(width)]


def parse_sheet_meta(ws) -> Tuple[str, str, Optional[float]]:
    fund_name = str(ws.cell(row=1, column=2).value or "").strip()
    label_code = str(ws.cell(row=2, column=1).value or "").strip()
    code = normalize_fund_code(ws.cell(row=2, column=2).value)
    latest_nav = to_float(ws.cell(row=3, column=2).value, default=None)
    if label_code != META_CODE_LABEL:
        return fund_name, "", latest_nav
    return fund_name, code, latest_nav


def is_transaction_sheet(ws) -> bool:
    _, code, _ = parse_sheet_meta(ws)
    if not code:
        return False
    base_headers = row_headers(ws, DATA_HEADER_ROW, len(BASE_DATA_HEADERS))
    if base_headers != BASE_DATA_HEADERS:
        return False
    full_headers = row_headers(ws, DATA_HEADER_ROW, len(DATA_HEADERS))
    return full_headers == DATA_HEADERS or full_headers[: len(BASE_DATA_HEADERS)] == BASE_DATA_HEADERS


def init_template_sheet(ws) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.cell(row=1, column=1).value = META_NAME_LABEL
    ws.cell(row=1, column=2).value = "示例基金"
    ws.cell(row=2, column=1).value = META_CODE_LABEL
    ws.cell(row=2, column=2).value = ""
    ws.cell(row=3, column=1).value = META_NAV_LABEL
    ws.cell(row=3, column=2).value = ""
    for i, h in enumerate(DATA_HEADERS, start=1):
        ws.cell(row=DATA_HEADER_ROW, column=i).value = h


def resolve_transaction_sheet_names(wb: openpyxl.Workbook) -> List[str]:
    return [
        name
        for name in wb.sheetnames
        if name not in SYSTEM_SHEETS and name != "transactions" and is_transaction_sheet(wb[name])
    ]


def load_config(path: Path) -> Dict:
    if not path.exists():
        return deep_merge(DEFAULT_CONFIG, {})
    with path.open("r", encoding="utf-8") as f:
        raw = yaml.safe_load(f) or {}
    return deep_merge(DEFAULT_CONFIG, raw)


def normalize_ntfy_url(ntfy_url: str) -> str:
    url = str(ntfy_url or "").strip()
    if not url:
        return ""
    if not re.match(r"^[A-Za-z][A-Za-z0-9+.\-]*://", url):
        url = f"https://{url.lstrip('/')}"
    return url


def ensure_transactions_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        has_transaction_data_sheet = any(
            name not in SYSTEM_SHEETS and name != "transactions" and is_transaction_sheet(wb[name])
            for name in wb.sheetnames
        )
        if has_transaction_data_sheet:
            return

        ws = wb[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.create_sheet(TEMPLATE_SHEET_NAME)
        init_template_sheet(ws)
        wb.save(path)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TEMPLATE_SHEET_NAME
    init_template_sheet(ws)
    wb.save(path)


def load_transactions(xlsx_path: Path) -> List[Dict]:
    ensure_transactions_workbook(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    target_sheet_names = resolve_transaction_sheet_names(wb)

    txns: List[Dict] = []
    uid = 1
    for sheet_name in target_sheet_names:
        ws = wb[sheet_name]
        _, code, _ = parse_sheet_meta(ws)
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            date_str = parse_date(row[0]) if len(row) > 0 else ""
            shares_raw = to_float(row[1] if len(row) > 1 else None, default=None)
            nav = to_float(row[2] if len(row) > 2 else None, default=None)
            note = str(row[3] if len(row) > 3 and row[3] is not None else "").strip()
            if shares_raw is None or abs(shares_raw) <= 0:
                continue
            action = "SELL" if shares_raw < 0 else "BUY"
            shares = abs(float(shares_raw))

            txns.append(
                {
                    "uid": uid,
                    "date": date_str or dt.date.today().isoformat(),
                    "fund_code": code,
                    "action": action,
                    "shares": shares,
                    "nav": float(nav) if nav is not None else None,
                    "note": note,
                }
            )
            uid += 1

    txns.sort(key=lambda x: (x["date"], x["uid"]))
    return txns


def load_fund_name_map(xlsx_path: Path) -> Dict[str, str]:
    if not xlsx_path.exists():
        return {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    name_map: Dict[str, str] = {}
    for name in resolve_transaction_sheet_names(wb):
        ws = wb[name]
        fund_name, code, _ = parse_sheet_meta(ws)
        if code:
            name_map[code] = fund_name or name
    return name_map


def fetch_latest_nav(code: str, session: requests.Session) -> Tuple[str, float]:
    params = {
        "fundCode": code,
        "pageIndex": 1,
        "pageSize": 1,
        "startDate": "",
        "endDate": "",
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://fundf10.eastmoney.com/",
        "Accept": "application/json, text/plain, */*",
    }

    resp = session.get(EASTMONEY_LSJZ_URL, params=params, headers=headers, timeout=20)
    resp.raise_for_status()
    text = resp.text.strip()

    if text.startswith(("jQuery", "jsonp", "callback")) and "(" in text:
        text = re.sub(r"^[^(]+\(", "", text)
        text = re.sub(r"\)\s*;?\s*$", "", text)

    data = json.loads(text)
    item = data["Data"]["LSJZList"][0]
    return str(item["FSRQ"]), float(item["DWJZ"])


def build_positions(transactions: List[Dict]) -> Tuple[List[Dict], float]:
    lots: List[Dict] = []
    oversold = 0.0

    for txn in sorted(transactions, key=lambda x: (x["date"], x["uid"])):
        if txn["action"] == "BUY":
            lots.append(
                {
                    "source_uid": txn["uid"],
                    "buy_date": txn["date"],
                    "buy_nav": float(txn["nav"] or 0.0),
                    "remaining_shares": float(txn["shares"]),
                }
            )
            continue

        remain_to_sell = float(txn["shares"])
        for lot in lots:
            if remain_to_sell <= 0:
                break
            if lot["remaining_shares"] <= 0:
                continue
            can_consume = min(lot["remaining_shares"], remain_to_sell)
            lot["remaining_shares"] -= can_consume
            remain_to_sell -= can_consume

        if remain_to_sell > 1e-12:
            oversold += remain_to_sell

    clean_lots = [x for x in lots if x["remaining_shares"] > 1e-12]
    return clean_lots, oversold


def summarize_fund(transactions: List[Dict], nav_date: str, latest_nav: float) -> Dict:
    lots, oversold = build_positions(transactions)

    holding_shares = sum(x["remaining_shares"] for x in lots)
    total_cost = sum(x["remaining_shares"] * x["buy_nav"] for x in lots)
    market_value = holding_shares * latest_nav
    profit_amount = market_value - total_cost
    profit_rate = (profit_amount / total_cost) if total_cost > 0 else 0.0

    for lot in lots:
        lot_cost = lot["remaining_shares"] * lot["buy_nav"]
        lot_profit = lot["remaining_shares"] * (latest_nav - lot["buy_nav"])
        lot["cost"] = lot_cost
        lot["latest_nav"] = latest_nav
        lot["profit_amount"] = lot_profit
        lot["profit_rate"] = (lot_profit / lot_cost) if lot_cost > 0 else 0.0

    return {
        "nav_date": nav_date,
        "latest_nav": latest_nav,
        "holding_shares": holding_shares,
        "total_cost": total_cost,
        "market_value": market_value,
        "profit_amount": profit_amount,
        "profit_rate": profit_rate,
        "oversold": oversold,
        "lots": lots,
    }


def write_output(input_path: Path, output_path: Path, summaries: Dict[str, Dict]) -> None:
    ensure_transactions_workbook(input_path)
    wb = openpyxl.load_workbook(input_path)

    # Keep workbook focused on transaction sheets only.
    _ = summaries
    for name in ("summary", "positions", "汇总", "持仓明细"):
        if name in wb.sheetnames:
            wb.remove(wb[name])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def send_ntfy(ntfy_url: str, title: str, message: str, tags: str = "moneybag") -> None:
    target = normalize_ntfy_url(ntfy_url)
    if not target:
        return
    safe_title = str(title or "").strip() or "Fund Alert"
    safe_tags = str(tags or "").strip() or "moneybag"
    try:
        safe_title.encode("latin-1")
    except UnicodeEncodeError:
        safe_title = "Fund Alert"
    try:
        safe_tags.encode("latin-1")
    except UnicodeEncodeError:
        safe_tags = "moneybag"
    headers = {
        "Title": safe_title,
        "Tags": safe_tags,
        "Priority": "max",
        "Content-Type": "text/plain; charset=utf-8",
    }
    resp = requests.post(target, data=message.encode("utf-8"), headers=headers, timeout=20)
    resp.raise_for_status()


def parse_github_datetime(value: str) -> Optional[dt.datetime]:
    s = str(value or "").strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        parsed = dt.datetime.fromisoformat(s)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.timezone.utc)
    return parsed.astimezone(dt.timezone.utc)


def read_keepalive_time(keepalive_path: Path) -> Optional[dt.datetime]:
    if not keepalive_path.exists():
        return None
    try:
        text = keepalive_path.read_text(encoding="utf-8")
    except Exception:
        return None
    for line in text.splitlines():
        if line.startswith("last_keepalive_utc="):
            raw = line.split("=", 1)[1].strip()
            return parse_github_datetime(raw)
    return None


def build_github_headers(token: str) -> Dict[str, str]:
    headers = {
        "Accept": "application/vnd.github+json",
        "User-Agent": "jijin-bot",
    }
    safe_token = str(token or "").strip()
    if safe_token:
        headers["Authorization"] = f"Bearer {safe_token}"
    return headers


def fetch_repo_last_activity(repo: str, token: str, session: requests.Session) -> Tuple[dt.datetime, str]:
    headers = build_github_headers(token)
    events_url = f"https://api.github.com/repos/{repo}/events"
    events_resp = session.get(events_url, headers=headers, params={"per_page": 1}, timeout=20)
    if events_resp.ok:
        events = events_resp.json()
        if isinstance(events, list) and events:
            event_time = parse_github_datetime(events[0].get("created_at"))
            if event_time is not None:
                return event_time, "events.created_at"

    repo_url = f"https://api.github.com/repos/{repo}"
    repo_resp = session.get(repo_url, headers=headers, timeout=20)
    repo_resp.raise_for_status()
    payload = repo_resp.json()
    candidates: List[Tuple[dt.datetime, str]] = []
    for key in ("pushed_at", "updated_at"):
        ts = parse_github_datetime(payload.get(key))
        if ts is not None:
            candidates.append((ts, key))
    if not candidates:
        raise RuntimeError("No usable activity timestamp from GitHub API.")
    candidates.sort(key=lambda x: x[0])
    return candidates[-1]


def run_git(args: List[str], repo_root: Path) -> subprocess.CompletedProcess:
    return subprocess.run(["git", *args], cwd=str(repo_root), capture_output=True, text=True, check=False)


def ensure_git_identity(repo_root: Path) -> None:
    name_proc = run_git(["config", "--get", "user.name"], repo_root)
    if name_proc.returncode != 0 or not name_proc.stdout.strip():
        run_git(["config", "user.name", "github-actions[bot]"], repo_root)
    email_proc = run_git(["config", "--get", "user.email"], repo_root)
    if email_proc.returncode != 0 or not email_proc.stdout.strip():
        run_git(["config", "user.email", "github-actions[bot]@users.noreply.github.com"], repo_root)


def detect_keepalive_branch(app: Dict) -> str:
    branch = str(app.get("repo_keepalive_branch") or "").strip()
    if branch:
        return branch
    branch = str(os.getenv("GITHUB_REF_NAME", "")).strip()
    if branch:
        return branch
    ref = str(os.getenv("GITHUB_REF", "")).strip()
    if ref.startswith("refs/heads/"):
        return ref.split("/", 2)[-1]
    return ""


def try_keep_repo_activity(app: Dict, repo_root: Path) -> Tuple[bool, str]:
    if not (repo_root / ".git").exists():
        return False, ".git not found in current workspace."

    keepalive_rel = str(app.get("repo_keepalive_file") or "data/.repo_activity_keepalive.txt").strip()
    keepalive_path = Path(keepalive_rel)
    if not keepalive_path.is_absolute():
        keepalive_path = repo_root / keepalive_path
    keepalive_path.parent.mkdir(parents=True, exist_ok=True)

    interval_days = int(to_float(app.get("repo_keepalive_interval_days"), default=30) or 30)
    interval_days = max(1, interval_days)
    now_utc = dt.datetime.now(dt.timezone.utc).replace(microsecond=0)
    last_keepalive = read_keepalive_time(keepalive_path)
    if last_keepalive is not None:
        elapsed = now_utc - last_keepalive
        if elapsed < dt.timedelta(days=interval_days):
            next_due = last_keepalive + dt.timedelta(days=interval_days)
            return True, (
                f"Keepalive skipped: last update at {last_keepalive.strftime('%Y-%m-%d %H:%M:%S')} UTC, "
                f"next due after {next_due.strftime('%Y-%m-%d %H:%M:%S')} UTC."
            )

    keepalive_path.write_text(f"last_keepalive_utc={now_utc.isoformat()}\n", encoding="utf-8")

    try:
        ensure_git_identity(repo_root)

        try:
            add_target = str(keepalive_path.relative_to(repo_root))
        except ValueError:
            return False, "repo_keepalive_file must be inside repository root."

        add_proc = run_git(["add", add_target], repo_root)
        if add_proc.returncode != 0:
            return False, f"git add failed: {add_proc.stderr.strip() or add_proc.stdout.strip()}"

        commit_msg = f"chore: repo keepalive {now_utc.strftime('%Y-%m-%d %H:%M:%SZ')} [skip ci]"
        commit_proc = run_git(["commit", "-m", commit_msg], repo_root)
        if commit_proc.returncode != 0:
            merged = (commit_proc.stderr + "\n" + commit_proc.stdout).lower()
            if "nothing to commit" in merged:
                return True, "No changes to commit."
            return False, f"git commit failed: {commit_proc.stderr.strip() or commit_proc.stdout.strip()}"

        branch = detect_keepalive_branch(app)
        push_args = ["push", "origin", f"HEAD:{branch}"] if branch else ["push", "origin", "HEAD"]
        push_proc = run_git(push_args, repo_root)
        if push_proc.returncode != 0:
            return False, f"git push failed: {push_proc.stderr.strip() or push_proc.stdout.strip()}"

        sha_proc = run_git(["rev-parse", "--short", "HEAD"], repo_root)
        sha = sha_proc.stdout.strip() if sha_proc.returncode == 0 else ""
        return True, f"Keepalive commit pushed successfully. sha={sha or 'unknown'}"
    except FileNotFoundError:
        return False, "git command not found in current runtime."


def handle_repo_activity(app: Dict, ntfy_url: str, session: requests.Session) -> None:
    repo = str(app.get("github_repo") or os.getenv("GITHUB_REPOSITORY", "")).strip()
    if not repo:
        return

    keepalive_enabled = to_bool(app.get("repo_keepalive"), default=True)
    inactivity_notify = to_bool(app.get("repo_inactivity_notify"), default=True)
    warn_days = int(to_float(app.get("repo_inactive_warn_days"), default=60) or 60)
    now_utc = dt.datetime.now(dt.timezone.utc).replace(microsecond=0)
    keepalive_ok = False
    keepalive_msg = ""

    if keepalive_enabled:
        try:
            keepalive_ok, keepalive_msg = try_keep_repo_activity(app, Path.cwd())
        except Exception as exc:
            keepalive_ok, keepalive_msg = False, f"unexpected keepalive error: {exc}"
        if keepalive_ok:
            print(f"[repo-activity] {keepalive_msg}")
        else:
            print(f"[repo-activity] keepalive failed: {keepalive_msg}")

    if keepalive_ok or not inactivity_notify or not ntfy_url:
        return

    token = str(app.get("github_token") or os.getenv("GITHUB_TOKEN", "")).strip()
    try:
        last_activity, source = fetch_repo_last_activity(repo, token, session)
        inactive_days = max(0, int((now_utc - last_activity).total_seconds() // 86400))
        remaining_days = warn_days - inactive_days
        if remaining_days > 0:
            status = f"仓库已 {inactive_days} 天没有活动，距 {warn_days} 天阈值还剩 {remaining_days} 天。"
            tags = "warning,github"
        else:
            status = f"仓库已 {inactive_days} 天没有活动，已超过 {warn_days} 天阈值，定时任务可能被禁用。"
            tags = "warning,rotating_light,github"

        lines = [
            f"仓库: {repo}",
            f"当前时间(UTC): {now_utc.strftime('%Y-%m-%d %H:%M:%S')}",
            f"最近活动(UTC): {last_activity.strftime('%Y-%m-%d %H:%M:%S')}",
            f"活动时间来源: {source}",
            status,
        ]
        if keepalive_enabled and keepalive_msg:
            lines.append(f"保活失败原因: {keepalive_msg}")
        send_ntfy(ntfy_url, title="Repo Activity Warning", message="\n".join(lines), tags=tags)
    except Exception as exc:
        print(f"[repo-activity] failed to query GitHub activity: {exc}")


def run_job(config: Dict, xlsx_override: str, out_override: str, threshold_override: Optional[float], ntfy_override: str) -> None:
    app = config["app"]
    xlsx_path = Path(xlsx_override or app["xlsx_path"])
    output_path = Path(out_override or app.get("output_path") or app["xlsx_path"])
    threshold = float(threshold_override if threshold_override is not None else app["threshold"])
    ntfy_url = ntfy_override or os.getenv("NTFY_URL", "").strip() or app.get("ntfy_url", "")
    session = requests.Session()

    transactions = load_transactions(xlsx_path)
    if not transactions:
        write_output(xlsx_path, output_path, summaries={})
        print(f"No transactions found. Empty summary saved to: {output_path}")
        handle_repo_activity(app, ntfy_url, session)
        return

    by_fund: Dict[str, List[Dict]] = {}
    for txn in transactions:
        by_fund.setdefault(txn["fund_code"], []).append(txn)
    fund_name_map = load_fund_name_map(xlsx_path)

    summaries: Dict[str, Dict] = {}
    errors: List[str] = []

    for code, fund_txns in sorted(by_fund.items()):
        try:
            nav_date, latest_nav = fetch_latest_nav(code, session)
            summaries[code] = summarize_fund(fund_txns, nav_date, latest_nav)
        except Exception as exc:
            errors.append(f"{code}: {exc}")

    if not summaries:
        handle_repo_activity(app, ntfy_url, session)
        raise RuntimeError("All fund NAV queries failed: " + " | ".join(errors))

    write_output(xlsx_path, output_path, summaries)

    if ntfy_url:
        has_threshold_alert = False
        for code, fund_txns in sorted(by_fund.items()):
            s = summaries.get(code)
            if not s:
                continue
            latest_nav = float(s["latest_nav"])
            fund_name = str(fund_name_map.get(code, "") or "").strip() or code
            for txn in sorted(fund_txns, key=lambda x: (x["date"], x["uid"])):
                if str(txn.get("action", "")).upper() != "BUY":
                    continue
                nav = to_float(txn.get("nav"), default=None)
                if nav is None or nav <= 0:
                    continue
                rate = (latest_nav - nav) / nav
                if rate < threshold:
                    continue
                profit = float(txn["shares"]) * (latest_nav - nav)
                lines = [
                    f"基金名称: {fund_name}",
                    f"购买日期: {txn['date']}",
                    f"份额: {txn['shares']:.2f}",
                    f"成交净值: {nav:.4f}",
                    f"最新净值: {latest_nav:.4f}",
                    f"盈利利率: {rate * 100:.2f}%",
                    f"盈利额: {profit:.2f}",
                ]
                send_ntfy(ntfy_url, title="基金收益提醒", message="\n".join(lines), tags="moneybag,rotating_light")
                has_threshold_alert = True

        if not has_threshold_alert and summaries:
            for code, fund_txns in sorted(by_fund.items()):
                s = summaries.get(code)
                if not s:
                    continue

                latest_nav = float(s["latest_nav"])
                fund_name = str(fund_name_map.get(code, "") or "").strip() or code

                best_txn = None
                best_nav = None
                best_rate = None
                for txn in sorted(fund_txns, key=lambda x: (x["date"], x["uid"])):
                    if str(txn.get("action", "")).upper() != "BUY":
                        continue
                    nav = to_float(txn.get("nav"), default=None)
                    if nav is None or nav <= 0:
                        continue
                    rate = (latest_nav - nav) / nav
                    if best_rate is None or rate > best_rate:
                        best_txn = txn
                        best_nav = nav
                        best_rate = rate

                if best_txn is not None and best_nav is not None and best_rate is not None:
                    profit = float(best_txn["shares"]) * (latest_nav - best_nav)
                    lines = [
                        f"基金名称: {fund_name}",
                        f"基金代码: {code}",
                        f"购买日期: {best_txn['date']}",
                        f"份额: {best_txn['shares']:.2f}",
                        f"成交净值: {best_nav:.4f}",
                        f"最新净值: {latest_nav:.4f}",
                        f"盈利利率: {best_rate * 100:.2f}%",
                        f"盈利额: {profit:.2f}",
                    ]
                else:
                    lines = [
                        f"基金名称: {fund_name}",
                        f"基金代码: {code}",
                        f"净值日期: {s['nav_date']}",
                        f"持仓份额: {float(s['holding_shares']):.2f}",
                        f"最新净值: {latest_nav:.4f}",
                        f"盈利利率: {float(s['profit_rate']) * 100:.2f}%",
                        f"盈利额: {float(s['profit_amount']):.2f}",
                    ]

                send_ntfy(
                    ntfy_url,
                    title="基金收益概览",
                    message="\n".join(lines),
                    tags="moneybag,bar_chart",
                )

    handle_repo_activity(app, ntfy_url, session)

    print(f"Saved: {output_path}")
    for code, s in sorted(summaries.items()):
        print(
            f"{code} nav@{s['nav_date']}={s['latest_nav']:.4f} holding={s['holding_shares']:.2f} "
            f"profit_rate={s['profit_rate'] * 100:.2f}% profit={s['profit_amount']:.2f}"
        )

    if errors:
        print("Errors:")
        for err in errors:
            print("  -", err)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="config/funds.yaml", help="config file path")
    parser.add_argument("--xlsx", default="", help="override xlsx path")
    parser.add_argument("--out", default="", help="override output xlsx path")
    parser.add_argument("--threshold", type=float, default=None, help="override threshold, e.g. 0.40")
    parser.add_argument("--ntfy", default="", help="override ntfy url")
    args = parser.parse_args()

    config = load_config(Path(args.config))
    run_job(
        config=config,
        xlsx_override=args.xlsx,
        out_override=args.out,
        threshold_override=args.threshold,
        ntfy_override=args.ntfy,
    )


if __name__ == "__main__":
    main()
