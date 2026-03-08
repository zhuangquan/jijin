import argparse
import datetime as dt
import json
import math
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import requests
import yaml

from jijin import (
    build_positions,
    fetch_latest_nav,
    load_config,
    load_transactions,
    normalize_fund_code,
    parse_date,
    send_ntfy,
    to_float,
)
from strategy import DualTriggerParams, DualTriggerState, decide, stock_basket_index_equal_weights


BOND_SHEET_NAME = "债卷基金"
BOND_HEADERS = ["基金代码", "基金名称", "份额", "成交净值", "备注"]
STRATEGY_UI_STATE_FILE = "data/strategy_ui_state.json"
STRATEGY_PEAK_FILE = "strategy_peak.yaml"
STRATEGY_FIXED_NEW_HIGH_EPS = 1e-10


def _to_bool(value: object, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value or "").strip().lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "on", "y", "是", "开启"}:
        return True
    if text in {"0", "false", "no", "off", "n", "否", "关闭"}:
        return False
    return default


def _row_headers(ws, row_idx: int, width: int) -> List[str]:
    return [str(ws.cell(row=row_idx, column=i + 1).value or "").strip() for i in range(width)]


def load_bond_transactions(xlsx_path: Path) -> List[Dict]:
    if not xlsx_path.exists():
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if BOND_SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[BOND_SHEET_NAME]
    if _row_headers(ws, 1, len(BOND_HEADERS)) != BOND_HEADERS:
        return []

    rows: List[Dict] = []
    uid = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = normalize_fund_code(row[0] if len(row) > 0 else "")
        shares = to_float(row[2] if len(row) > 2 else None, default=None)
        nav = to_float(row[3] if len(row) > 3 else None, default=None)
        if not code or shares is None or shares <= 0 or nav is None or nav <= 0:
            continue
        rows.append(
            {
                "uid": uid,
                "fund_code": code,
                "fund_name": str(row[1] if len(row) > 1 and row[1] is not None else "").strip(),
                "shares": float(shares),
                "nav": float(nav),
                "note": str(row[4] if len(row) > 4 and row[4] is not None else "").strip(),
            }
        )
        uid += 1
    return rows


def _decode_json_payload(resp: requests.Response) -> Dict:
    text = resp.text.strip()
    if text.startswith(("jQuery", "jsonp", "callback")) and "(" in text:
        text = text[text.find("(") + 1 :]
        text = text.rstrip("); \n\r\t")
    return json.loads(text)


def fetch_nav_history_series(code: str, session: requests.Session, max_pages: int = 120) -> pd.Series:
    code = normalize_fund_code(code)
    if not code:
        raise ValueError("invalid fund code")
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://fundf10.eastmoney.com/",
        "Accept": "application/json, text/plain, */*",
    }
    points: Dict[pd.Timestamp, float] = {}
    page_size = 20
    for page_idx in range(1, max_pages + 1):
        params = {
            "fundCode": code,
            "pageIndex": page_idx,
            "pageSize": page_size,
            "startDate": "",
            "endDate": "",
        }
        resp = session.get("https://api.fund.eastmoney.com/f10/lsjz", params=params, headers=headers, timeout=20)
        resp.raise_for_status()
        payload = _decode_json_payload(resp)
        rows = ((payload.get("Data") or {}).get("LSJZList") or [])
        if not rows:
            break
        for row in rows:
            date_text = parse_date(row.get("FSRQ"))
            nav = to_float(row.get("DWJZ"), default=None)
            if date_text and nav is not None and nav > 0:
                points[pd.Timestamp(date_text)] = float(nav)
        if len(rows) < page_size:
            break

    if not points:
        raise ValueError(f"cannot fetch nav history for {code}")
    s = pd.Series(points).sort_index()
    s.name = code
    return s


def build_nav_df(funds: List[str], history_days: int, session: requests.Session) -> pd.DataFrame:
    need_days = max(1, int(history_days))
    fetch_days = max(need_days + 120, need_days * 2)
    max_pages = max(5, min(120, int(math.ceil(fetch_days / 20.0))))

    series_map: Dict[str, pd.Series] = {}
    errors: List[str] = []
    for code in funds:
        try:
            series_map[code] = fetch_nav_history_series(code, session=session, max_pages=max_pages)
        except Exception as exc:
            errors.append(f"{code}: {exc}")
    if errors:
        raise RuntimeError("获取净值历史失败: " + " | ".join(errors[:5]))

    nav_df = pd.concat({k: series_map[k] for k in funds if k in series_map}, axis=1).sort_index()
    nav_df = nav_df.ffill().dropna(how="any")
    nav_df = nav_df.tail(need_days)
    if nav_df.empty:
        raise RuntimeError("可用历史净值数据为空")
    nav_df.columns = funds
    return nav_df


def current_stock_shares(transactions: List[Dict], funds: List[str]) -> Dict[str, float]:
    shares_map: Dict[str, float] = {code: 0.0 for code in funds}
    by_fund: Dict[str, List[Dict]] = {}
    for txn in transactions:
        code = normalize_fund_code(txn.get("fund_code"))
        if code in shares_map:
            by_fund.setdefault(code, []).append(txn)
    for code in funds:
        lots, _ = build_positions(sorted(by_fund.get(code, []), key=lambda x: (x["date"], x["uid"])))
        shares_map[code] = float(sum(float(x["remaining_shares"]) for x in lots))
    return shares_map


def compute_bond_market_value(bond_txns: List[Dict], latest_nav_map: Dict[str, float]) -> float:
    by_code_shares: Dict[str, float] = {}
    for txn in bond_txns:
        code = normalize_fund_code(txn.get("fund_code"))
        shares = to_float(txn.get("shares"), default=None)
        if not code or shares is None or shares <= 0:
            continue
        by_code_shares[code] = by_code_shares.get(code, 0.0) + float(shares)
    total = 0.0
    for code, shares in by_code_shares.items():
        nav = to_float(latest_nav_map.get(code), default=None)
        if nav is None or nav <= 0:
            raise RuntimeError(f"债基 {code} 缺少最新净值，无法执行决策")
        total += float(shares) * float(nav)
    return float(total)


def load_strategy_ui_payload(path: Path) -> Dict:
    if not path.exists():
        raise FileNotFoundError(f"未找到策略状态文件: {path}")
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise RuntimeError("策略状态文件格式错误：根节点必须为对象")
    return raw


def _load_params_input(payload: Dict) -> Dict:
    parsed = payload.get("params_parsed")
    if isinstance(parsed, dict):
        return dict(parsed)
    raw_json = payload.get("params_json")
    if isinstance(raw_json, str) and raw_json.strip():
        data = json.loads(raw_json)
        if isinstance(data, dict):
            return dict(data)
    return {}


def _load_state_input(payload: Dict) -> Dict:
    raw_json = payload.get("state_json")
    if isinstance(raw_json, str) and raw_json.strip():
        data = json.loads(raw_json)
        if isinstance(data, dict):
            return dict(data)
    return {}


def parse_strategy_params(funds: List[str], data_input: Dict) -> DualTriggerParams:
    data = dict(data_input or {})
    if "dd_stages" in data:
        data["dd_stages"] = tuple(float(x) for x in list(data["dd_stages"]))
    if "gap_fill" in data:
        data["gap_fill"] = tuple(float(x) for x in list(data["gap_fill"]))
    if "peak_rolling_days" in data and data["peak_rolling_days"] in ("", None):
        data["peak_rolling_days"] = None
    if "target_stock_fund_weights" in data and data["target_stock_fund_weights"] in ("", None):
        data["target_stock_fund_weights"] = None

    allowed_keys = set(getattr(DualTriggerParams, "__dataclass_fields__", {}).keys())
    if "eps" in allowed_keys and "eps" not in data:
        data["eps"] = STRATEGY_FIXED_NEW_HIGH_EPS
    if "new_high_epsilon" in allowed_keys and "new_high_epsilon" not in data:
        data["new_high_epsilon"] = STRATEGY_FIXED_NEW_HIGH_EPS
    if "new_high_epsilon" in data and "eps" in allowed_keys and "eps" not in data:
        data["eps"] = data.get("new_high_epsilon")
    if allowed_keys:
        data = {k: v for k, v in data.items() if k in allowed_keys}

    params = DualTriggerParams(**data)
    if params.target_stock_fund_weights is not None:
        mapped: Dict[str, float] = {}
        for k, v in params.target_stock_fund_weights.items():
            code = normalize_fund_code(k)
            num = to_float(v, default=None)
            if code in funds and num is not None and num > 0:
                mapped[code] = float(num)
        params.target_stock_fund_weights = mapped if mapped else None
    return params


def parse_strategy_state(data_input: Dict) -> DualTriggerState:
    data = dict(data_input or {})
    allowed_keys = set(getattr(DualTriggerState, "__dataclass_fields__", {}).keys())
    if allowed_keys:
        data = {k: v for k, v in data.items() if k in allowed_keys}
    if "dd_triggered" in data:
        data["dd_triggered"] = [float(x) for x in list(data["dd_triggered"])]
    return DualTriggerState(**data)


def load_peak_config(path: Path) -> Dict:
    if not path.exists():
        return {}
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return raw if isinstance(raw, dict) else {}


def prepare_basket_with_peak_config(basket: pd.Series, today: pd.Timestamp, peak_cfg: Dict) -> pd.Series:
    peak_date = str(peak_cfg.get("peak_date", "") or "").strip()
    if not peak_date:
        return basket
    try:
        start_ts = pd.Timestamp(peak_date)
    except Exception:
        return basket
    eligible = basket.index[basket.index >= start_ts]
    if len(eligible) > 0 and eligible[0] <= today:
        return basket.loc[eligible[0] :]
    return basket


def update_peak_config_after_decision(
    path: Path,
    cfg_before: Dict,
    basket_used: pd.Series,
    today: pd.Timestamp,
    params: DualTriggerParams,
    result: Dict,
) -> None:
    cfg = dict(cfg_before or {})
    metrics = (result or {}).get("metrics") or {}
    new_state = (result or {}).get("new_state") or {}

    window = basket_used.loc[:today]
    if params.peak_rolling_days is not None:
        window = window.tail(int(params.peak_rolling_days))
    decision_peak_date = ""
    decision_peak_value = None
    if not window.empty:
        peak_idx = window.idxmax()
        decision_peak_date = peak_idx.date().isoformat()
        decision_peak_value = float(window.loc[peak_idx])

    manual_reset_date = str(metrics.get("manual_reset_applied_date", "") or "").strip()
    manual_reset_peak = to_float(metrics.get("manual_reset_applied_peak"), default=None)
    if manual_reset_date and manual_reset_peak is not None:
        cfg["peak_value"] = float(manual_reset_peak)
        cfg["peak_date"] = manual_reset_date
        cfg["reason"] = "manual_reset_applied"
    elif decision_peak_value is not None:
        old_peak_val = to_float(cfg.get("peak_value"), default=None)
        old_peak_date = str(cfg.get("peak_date", "") or "")
        should_init = (not old_peak_date) and (old_peak_val is None)
        should_new_high = (
            old_peak_val is not None
            and decision_peak_value >= float(old_peak_val) - 1e-12
            and bool(decision_peak_date)
        )
        if should_init or should_new_high:
            cfg["peak_value"] = float(decision_peak_value)
            cfg["peak_date"] = decision_peak_date
            cfg["reason"] = "new_high_update" if should_new_high else "init_from_decision_peak"

    trade_peak_val = to_float(metrics.get("peak_trade"), default=None)
    if trade_peak_val is not None:
        cfg["trade_peak_value"] = float(trade_peak_val)
        cfg["trade_peak_update_date"] = today.date().isoformat()

    cfg["pending_reset_date"] = str(new_state.get("pending_reset_date", "") or "")
    cfg["updated_at"] = dt.datetime.now().isoformat(timespec="seconds")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(cfg, allow_unicode=True, sort_keys=False), encoding="utf-8")


def save_strategy_ui_payload(
    path: Path,
    payload: Dict,
    params_input: Dict,
    result: Dict,
    today: pd.Timestamp,
    history_days: int,
    cash_value: float,
) -> None:
    out = dict(payload or {})
    base = out.get("base") if isinstance(out.get("base"), dict) else {}
    base["today"] = today.date().isoformat()
    base["history_days"] = str(int(history_days))
    base["cash_value"] = f"{float(cash_value):g}"
    out["base"] = base
    out["saved_at"] = dt.datetime.now().isoformat(timespec="seconds")

    params_clean = json.loads(json.dumps(params_input or {}, ensure_ascii=False))
    out["params_parsed"] = params_clean
    out["params_json"] = json.dumps(params_clean, ensure_ascii=False, indent=2)
    out["state_json"] = json.dumps((result or {}).get("new_state") or {}, ensure_ascii=False, indent=2)

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")


def build_decision_notify_message(today: pd.Timestamp, result: Dict, stock_values: Dict[str, float], bond_value: float, cash_value: float) -> str:
    metrics = (result or {}).get("metrics") or {}
    lines = [
        f"决策日: {today.date().isoformat()}",
        f"动作: {result.get('action', '-')}",
        f"原因: {result.get('reason', '-')}",
        f"股票金额: {sum(stock_values.values()):.2f}",
        f"债基金额: {float(bond_value):.2f}",
        f"现金金额: {float(cash_value):.2f}",
        f"当前仓位W: {float(to_float(metrics.get('W'), default=0.0) or 0.0) * 100:.2f}%",
    ]
    return "\n".join(lines)


def run_decision(
    config_path: Path,
    state_path: Path,
    today_override: str,
    history_days_override: Optional[int],
    cash_override: Optional[float],
    ntfy_override: str,
) -> None:
    config = load_config(config_path)
    app = config.get("app", {}) if isinstance(config.get("app", {}), dict) else {}
    xlsx_path = Path(str(app.get("xlsx_path", "data/jijin.xlsx")))

    transactions = load_transactions(xlsx_path)
    bond_txns = load_bond_transactions(xlsx_path)
    funds = sorted({normalize_fund_code(txn.get("fund_code")) for txn in transactions if normalize_fund_code(txn.get("fund_code"))})
    if not funds:
        raise RuntimeError("当前没有可用于决策的股票基金持仓")

    payload = load_strategy_ui_payload(state_path)
    params_input = _load_params_input(payload)
    state_input = _load_state_input(payload)
    params = parse_strategy_params(funds, params_input)
    state = parse_strategy_state(state_input)

    base = payload.get("base") if isinstance(payload.get("base"), dict) else {}
    history_days = int(history_days_override if history_days_override is not None else (to_float(base.get("history_days"), default=260) or 260))
    history_days = max(1, history_days)
    cash_value = float(cash_override if cash_override is not None else (to_float(base.get("cash_value"), default=0.0) or 0.0))

    desired_today_text = parse_date(today_override.strip()) if str(today_override or "").strip() else dt.datetime.now(
        dt.timezone(dt.timedelta(hours=8))
    ).date().isoformat()
    desired_today = pd.Timestamp(desired_today_text) if desired_today_text else None

    with requests.Session() as sess:
        all_codes = sorted(set(funds) | {normalize_fund_code(x.get("fund_code")) for x in bond_txns if normalize_fund_code(x.get("fund_code"))})
        latest_nav_map: Dict[str, float] = {}
        nav_errors: List[str] = []
        for code in all_codes:
            try:
                _nav_date, latest_nav = fetch_latest_nav(code, sess)
                latest_nav_map[code] = float(latest_nav)
            except Exception as exc:
                nav_errors.append(f"{code}: {exc}")
        if nav_errors:
            raise RuntimeError("更新最新净值失败: " + " | ".join(nav_errors[:5]))

        shares_map = current_stock_shares(transactions, funds)
        nav_df = build_nav_df(funds, history_days=history_days, session=sess)
        bond_value = compute_bond_market_value(bond_txns, latest_nav_map)

    if desired_today is not None:
        eligible = nav_df.index[nav_df.index <= desired_today]
        if len(eligible) <= 0:
            raise RuntimeError("决策日早于历史数据范围")
        today = eligible[-1]
    else:
        today = nav_df.index[-1]

    nav_today = nav_df.loc[today]
    stock_values = {code: float(shares_map.get(code, 0.0)) * float(nav_today[code]) for code in funds}

    basket = stock_basket_index_equal_weights(nav_df, funds)
    peak_cfg_path = config_path.parent / STRATEGY_PEAK_FILE
    peak_cfg_before = load_peak_config(peak_cfg_path)
    basket_for_decision = prepare_basket_with_peak_config(basket, today, peak_cfg_before)

    result = decide(
        today=today,
        basket_idx=basket_for_decision,
        bond_value=float(bond_value),
        cash_value=float(cash_value),
        stock_values=stock_values,
        funds=funds,
        params=params,
        state=state,
    )

    update_peak_config_after_decision(
        path=peak_cfg_path,
        cfg_before=peak_cfg_before,
        basket_used=basket_for_decision,
        today=today,
        params=params,
        result=result,
    )
    save_strategy_ui_payload(
        path=state_path,
        payload=payload,
        params_input=params_input,
        result=result,
        today=today,
        history_days=history_days,
        cash_value=cash_value,
    )

    notify_enabled = _to_bool(app.get("decision_notify"), default=False)
    ntfy_url = str(ntfy_override or "").strip() or str(os.getenv("NTFY_URL", "")).strip() or str(app.get("ntfy_url", "")).strip()
    if notify_enabled and ntfy_url:
        message = build_decision_notify_message(today, result, stock_values, bond_value, cash_value)
        send_ntfy(ntfy_url, title="Strategy Decision Success", message=message, tags="moneybag,memo")

    print(
        f"Decision done: action={result.get('action')} reason={result.get('reason')} "
        f"today={today.date().isoformat()} stock={sum(stock_values.values()):.2f} bond={bond_value:.2f} cash={cash_value:.2f}"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="config/funds.yaml", help="config file path")
    parser.add_argument("--state", default=STRATEGY_UI_STATE_FILE, help="strategy ui state json path")
    parser.add_argument("--today", default="", help="decision date (YYYY-MM-DD), default CN today")
    parser.add_argument("--history-days", type=int, default=None, help="override history days")
    parser.add_argument("--cash", type=float, default=None, help="override cash value")
    parser.add_argument("--ntfy", default="", help="override ntfy url")
    args = parser.parse_args()

    config_path = Path(args.config)
    state_path = Path(args.state)
    try:
        run_decision(
            config_path=config_path,
            state_path=state_path,
            today_override=args.today,
            history_days_override=args.history_days,
            cash_override=args.cash,
            ntfy_override=args.ntfy,
        )
    except Exception as exc:
        try:
            cfg = load_config(config_path)
            app = cfg.get("app", {}) if isinstance(cfg.get("app", {}), dict) else {}
            ntfy_url = str(args.ntfy or "").strip() or str(os.getenv("NTFY_URL", "")).strip() or str(app.get("ntfy_url", "")).strip()
            notify_enabled = _to_bool(app.get("decision_notify"), default=False)
            if notify_enabled and ntfy_url:
                send_ntfy(ntfy_url, title="Strategy Decision Failed", message=f"执行失败: {exc}", tags="warning,rotating_light")
        except Exception:
            pass
        raise


if __name__ == "__main__":
    main()
