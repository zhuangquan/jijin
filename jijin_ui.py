import argparse
import concurrent.futures
import copy
import datetime as dt
import json
import math
import re
import threading
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import requests
import tkinter as tk
import tkinter.font as tkfont
from tkinter import filedialog, messagebox, simpledialog, ttk
import yaml

from strategy import DualTriggerParams, DualTriggerState, decide, stock_basket_index_equal_weights


EASTMONEY_LSJZ_URL = "https://api.fund.eastmoney.com/f10/lsjz"
EASTMONEY_FUND_INFO_URL = "https://fund.eastmoney.com/pingzhongdata/{code}.js"
TX_HEADERS = ["date", "fund_code", "action", "shares", "nav", "note"]
SYSTEM_SHEETS = {"summary", "positions", "汇总", "持仓明细"}
TEMPLATE_SHEET_NAME = "模板"
META_NAME_LABEL = "基金名称"
META_CODE_LABEL = "基金代码"
META_NAV_LABEL = "最新净值"
BASE_DATA_HEADERS = ["日期", "份额", "购买净值", "备注"]
DATA_HEADERS = BASE_DATA_HEADERS + ["最新净值", "盈利利率", "盈利额"]
DATA_HEADER_ROW = 5
DATA_START_ROW = DATA_HEADER_ROW + 1
BOND_SHEET_NAME = "债卷基金"
BOND_HEADERS = ["基金代码", "基金名称", "份额", "成交净值", "备注"]
DECISION_FREQ_LABEL_TO_VALUE = {
    "每日": "daily",
    "每周": "weekly",
    "每月": "monthly",
    "每季度": "quarterly",
}
DECISION_FREQ_VALUE_TO_LABEL = {v: k for k, v in DECISION_FREQ_LABEL_TO_VALUE.items()}
STRATEGY_PEAK_CONFIG_FILE = "strategy_peak.yaml"
STRATEGY_UI_STATE_FILE = "strategy_ui_state.json"
STRATEGY_FIXED_NEW_HIGH_EPS = 1e-10

DEFAULT_CONFIG = {
    "app": {
        "xlsx_path": "data/jijin.xlsx",
        "threshold": 0.40,
        "ntfy_url": "",
    },
    "funds": [],
}


def deep_merge(base: Dict, patch: Dict) -> Dict:
    result = copy.deepcopy(base)
    for key, value in (patch or {}).items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = deep_merge(result[key], value)
        else:
            result[key] = value
    return result


def normalize_fund_code(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if s.isdigit():
        s = s.zfill(6)
    if re.fullmatch(r"\d{6}", s):
        return s
    return ""


def parse_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def to_float(value, default: Optional[float] = None) -> Optional[float]:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def row_headers(ws, row_idx: int, width: int) -> List[str]:
    return [str(ws.cell(row=row_idx, column=i + 1).value or "").strip() for i in range(width)]


def parse_sheet_meta(ws) -> Tuple[str, str, Optional[float]]:
    fund_name = str(ws.cell(row=1, column=2).value or "").strip()
    label_code = str(ws.cell(row=2, column=1).value or "").strip()
    code = normalize_fund_code(ws.cell(row=2, column=2).value)
    latest_nav = to_float(ws.cell(row=3, column=2).value, default=None)
    if label_code != META_CODE_LABEL:
        return fund_name, "", latest_nav
    return fund_name, code, latest_nav


def is_transaction_sheet(ws) -> bool:
    _, code, _ = parse_sheet_meta(ws)
    if not code:
        return False
    base_headers = row_headers(ws, DATA_HEADER_ROW, len(BASE_DATA_HEADERS))
    if base_headers != BASE_DATA_HEADERS:
        return False
    full_headers = row_headers(ws, DATA_HEADER_ROW, len(DATA_HEADERS))
    return full_headers == DATA_HEADERS or full_headers[: len(BASE_DATA_HEADERS)] == BASE_DATA_HEADERS


def init_template_sheet(ws) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.cell(row=1, column=1).value = META_NAME_LABEL
    ws.cell(row=1, column=2).value = "示例基金"
    ws.cell(row=2, column=1).value = META_CODE_LABEL
    ws.cell(row=2, column=2).value = ""
    ws.cell(row=3, column=1).value = META_NAV_LABEL
    ws.cell(row=3, column=2).value = ""
    for i, h in enumerate(DATA_HEADERS, start=1):
        ws.cell(row=DATA_HEADER_ROW, column=i).value = h


def sanitize_sheet_title(title: str) -> str:
    invalid = set('\\/*?:[]')
    return "".join(ch for ch in str(title or "").strip() if ch not in invalid).strip()


def normalize_fund_display_name(name: str, code: str = "") -> str:
    s = str(name or "").strip()
    s = s.replace("（LOF）", "").replace("(LOF)", "")
    s = re.sub(r"\([^)]*\)", "", s).strip()
    s = re.sub(r"[A-Za-z]+", "", s).strip()
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return f"基金{code}" if code else "基金"
    return s


def alloc_sheet_title(base: str, used_titles: set) -> str:
    base_clean = sanitize_sheet_title(base) or "基金"
    if len(base_clean) > 31:
        base_clean = base_clean[:31]
    title = base_clean
    idx = 1
    while title in used_titles:
        suffix = f"_{idx}"
        keep = 31 - len(suffix)
        title = f"{base_clean[:keep]}{suffix}"
        idx += 1
    used_titles.add(title)
    return title


def resolve_transaction_sheet_names(wb: openpyxl.Workbook) -> List[str]:
    return [
        name
        for name in wb.sheetnames
        if name not in SYSTEM_SHEETS and name != "transactions" and is_transaction_sheet(wb[name])
    ]


def load_latest_nav_map(xlsx_path: Path) -> Dict[str, float]:
    if not xlsx_path.exists():
        return {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    nav_map: Dict[str, float] = {}
    for name in resolve_transaction_sheet_names(wb):
        ws = wb[name]
        _, code, nav = parse_sheet_meta(ws)
        if code and nav is not None and nav > 0:
            nav_map[code] = float(nav)
    return nav_map


def load_fund_name_map(xlsx_path: Path) -> Dict[str, str]:
    if not xlsx_path.exists():
        return {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    name_map: Dict[str, str] = {}
    for name in resolve_transaction_sheet_names(wb):
        ws = wb[name]
        fund_name, code, _ = parse_sheet_meta(ws)
        if code:
            name_map[code] = fund_name or name
    return name_map


def load_or_default(path: Path) -> Dict:
    if not path.exists():
        return copy.deepcopy(DEFAULT_CONFIG)
    with path.open("r", encoding="utf-8") as f:
        raw = yaml.safe_load(f) or {}
    return deep_merge(DEFAULT_CONFIG, raw)


def save_config(path: Path, config: Dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        yaml.safe_dump(config, f, allow_unicode=True, sort_keys=False)


def normalize_ntfy_url(ntfy_url: str) -> str:
    url = str(ntfy_url or "").strip()
    if not url:
        return ""
    if not re.match(r"^[A-Za-z][A-Za-z0-9+.\-]*://", url):
        url = f"https://{url.lstrip('/')}"
    return url


def send_ntfy(ntfy_url: str, title: str, message: str, tags: str = "moneybag") -> None:
    target = normalize_ntfy_url(ntfy_url)
    if not target:
        return
    safe_title = str(title or "").strip() or "Fund Alert"
    safe_tags = str(tags or "").strip() or "moneybag"
    try:
        safe_title.encode("latin-1")
    except UnicodeEncodeError:
        safe_title = "Fund Alert"
    try:
        safe_tags.encode("latin-1")
    except UnicodeEncodeError:
        safe_tags = "moneybag"
    headers = {
        "Title": safe_title,
        "Tags": safe_tags,
        "Priority": "max",
        "Content-Type": "text/plain; charset=utf-8",
    }
    resp = requests.post(target, data=message.encode("utf-8"), headers=headers, timeout=20)
    resp.raise_for_status()


def ensure_transactions_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        has_transaction_data_sheet = any(
            name not in SYSTEM_SHEETS and name != "transactions" and is_transaction_sheet(wb[name])
            for name in wb.sheetnames
        )
        if has_transaction_data_sheet:
            return

        ws = wb[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.create_sheet(TEMPLATE_SHEET_NAME)
        init_template_sheet(ws)
        wb.save(path)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TEMPLATE_SHEET_NAME
    init_template_sheet(ws)
    wb.save(path)


def load_transactions(xlsx_path: Path) -> List[Dict]:
    ensure_transactions_workbook(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    target_sheet_names = resolve_transaction_sheet_names(wb)

    txns = []
    uid = 1
    for sheet_name in target_sheet_names:
        ws = wb[sheet_name]
        _, code, _ = parse_sheet_meta(ws)
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            date_str = parse_date(row[0]) if len(row) > 0 else ""
            shares_raw = to_float(row[1] if len(row) > 1 else None, default=None)
            nav = to_float(row[2] if len(row) > 2 else None, default=None)
            note = str(row[3] if len(row) > 3 and row[3] is not None else "").strip()
            if shares_raw is None or abs(shares_raw) <= 0:
                continue
            action = "SELL" if shares_raw < 0 else "BUY"
            shares = abs(float(shares_raw))

            txns.append(
                {
                    "uid": uid,
                    "date": date_str or dt.date.today().isoformat(),
                    "fund_code": code,
                    "action": action,
                    "shares": shares,
                    "nav": float(nav) if nav is not None else None,
                    "note": note,
                }
            )
            uid += 1

    txns.sort(key=lambda x: (x["date"], x["uid"]))
    return txns


def load_bond_transactions(xlsx_path: Path) -> List[Dict]:
    ensure_transactions_workbook(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if BOND_SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[BOND_SHEET_NAME]
    headers = row_headers(ws, 1, len(BOND_HEADERS))
    if headers != BOND_HEADERS:
        return []

    txns: List[Dict] = []
    uid = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = normalize_fund_code(row[0] if len(row) > 0 else "")
        name = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()
        shares = to_float(row[2] if len(row) > 2 else None, default=None)
        nav = to_float(row[3] if len(row) > 3 else None, default=None)
        note = str(row[4] if len(row) > 4 and row[4] is not None else "").strip()
        if not code:
            continue
        if shares is None or shares <= 0:
            continue
        if nav is None or nav <= 0:
            continue
        txns.append(
            {
                "uid": uid,
                "fund_code": code,
                "fund_name": name,
                "shares": float(shares),
                "nav": float(nav),
                "note": note,
            }
        )
        uid += 1
    return txns


def save_transactions(
    xlsx_path: Path,
    transactions: List[Dict],
    latest_nav_map: Optional[Dict[str, float]] = None,
    fund_name_map: Optional[Dict[str, str]] = None,
) -> None:
    ensure_transactions_workbook(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    merged_nav_map = load_latest_nav_map(xlsx_path)
    merged_name_map = load_fund_name_map(xlsx_path)
    for code, value in (latest_nav_map or {}).items():
        norm_code = normalize_fund_code(code)
        nav = to_float(value, default=None)
        if norm_code and nav is not None and nav > 0:
            merged_nav_map[norm_code] = float(nav)
    for code, name in (fund_name_map or {}).items():
        norm_code = normalize_fund_code(code)
        if norm_code and str(name or "").strip():
            merged_name_map[norm_code] = str(name).strip()

    tx_sheet_names = [
        name
        for name in wb.sheetnames
        if name not in SYSTEM_SHEETS and (name == TEMPLATE_SHEET_NAME or name == "transactions" or is_transaction_sheet(wb[name]))
    ]
    for name in tx_sheet_names:
        wb.remove(wb[name])

    if "Sheet" in wb.sheetnames:
        ws_default = wb["Sheet"]
        if ws_default.max_row <= 1 and ws_default.max_column <= 1 and ws_default.cell(row=1, column=1).value in (None, ""):
            wb.remove(ws_default)

    grouped: Dict[str, List[Dict]] = {}
    for txn in sorted(transactions, key=lambda x: (x["date"], x["uid"]), reverse=True):
        code = normalize_fund_code(txn.get("fund_code"))
        if not code:
            continue
        grouped.setdefault(code, []).append(txn)

    if not grouped:
        ws = wb.create_sheet(TEMPLATE_SHEET_NAME)
        init_template_sheet(ws)
    else:
        used_titles = set(wb.sheetnames)
        for code in sorted(grouped.keys()):
            fund_name = normalize_fund_display_name(merged_name_map.get(code, ""), code=code)
            title = alloc_sheet_title(fund_name or code, used_titles)
            ws = wb.create_sheet(title)
            ws.cell(row=1, column=1).value = META_NAME_LABEL
            ws.cell(row=1, column=2).value = fund_name or code
            ws.cell(row=2, column=1).value = META_CODE_LABEL
            ws.cell(row=2, column=2).value = code
            ws.cell(row=3, column=1).value = META_NAV_LABEL
            latest_nav = merged_nav_map.get(code)
            ws.cell(row=3, column=2).value = "" if latest_nav is None else float(latest_nav)
            for i, h in enumerate(DATA_HEADERS, start=1):
                ws.cell(row=DATA_HEADER_ROW, column=i).value = h
            write_txns = sorted(grouped[code], key=lambda x: (x["date"], x["uid"]), reverse=True)
            for txn in write_txns:
                shares = float(txn["shares"])
                action = str(txn.get("action", "BUY")).upper()
                if action == "SELL":
                    shares = -shares
                nav = to_float(txn["nav"], default=None)
                rate_text = ""
                profit_amount = ""
                if action == "BUY":
                    if latest_nav is not None and nav is not None and nav > 0:
                        rate_text = f"{((latest_nav - nav) / nav) * 100:.2f}%"
                        profit_amount = round(float(txn["shares"]) * (latest_nav - nav), 2)
                ws.append(
                    [
                        txn["date"],
                        shares,
                        "" if nav is None else float(nav),
                        txn.get("note", ""),
                        "" if latest_nav is None else float(latest_nav),
                        rate_text,
                        profit_amount,
                    ]
                )

    wb.save(xlsx_path)


def save_bond_transactions(xlsx_path: Path, transactions: List[Dict]) -> None:
    ensure_transactions_workbook(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    if BOND_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[BOND_SHEET_NAME])
    ws = wb.create_sheet(BOND_SHEET_NAME)
    ws.append(BOND_HEADERS)

    for txn in transactions:
        code = normalize_fund_code(txn.get("fund_code"))
        if not code:
            continue
        shares = to_float(txn.get("shares"), default=None)
        nav = to_float(txn.get("nav"), default=None)
        if shares is None or shares <= 0:
            continue
        if nav is None or nav <= 0:
            continue
        ws.append(
            [
                code,
                str(txn.get("fund_name", "") or "").strip(),
                float(shares),
                float(nav),
                str(txn.get("note", "") or "").strip(),
            ]
        )
    wb.save(xlsx_path)


def fetch_latest_nav(code: str, session: requests.Session) -> Tuple[str, float]:
    params = {
        "fundCode": code,
        "pageIndex": 1,
        "pageSize": 1,
        "startDate": "",
        "endDate": "",
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://fundf10.eastmoney.com/",
        "Accept": "application/json, text/plain, */*",
    }

    resp = session.get(EASTMONEY_LSJZ_URL, params=params, headers=headers, timeout=20)
    resp.raise_for_status()
    data = resp.json()
    item = data["Data"]["LSJZList"][0]
    return str(item["FSRQ"]), float(item["DWJZ"])


def fetch_nav_history_series(code: str, session: requests.Session, max_pages: int = 120) -> pd.Series:
    code = normalize_fund_code(code)
    if not code:
        raise ValueError("invalid fund code")

    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://fundf10.eastmoney.com/",
        "Accept": "application/json, text/plain, */*",
    }
    page_size = 20
    points: Dict[pd.Timestamp, float] = {}
    for page_idx in range(1, max_pages + 1):
        params = {
            "fundCode": code,
            "pageIndex": page_idx,
            "pageSize": page_size,
            "startDate": "",
            "endDate": "",
        }
        resp = session.get(EASTMONEY_LSJZ_URL, params=params, headers=headers, timeout=20)
        resp.raise_for_status()
        payload = resp.json()
        rows = ((payload.get("Data") or {}).get("LSJZList") or [])
        if not rows:
            break
        for row in rows:
            date_text = parse_date(row.get("FSRQ"))
            nav = to_float(row.get("DWJZ"), default=None)
            if not date_text or nav is None or nav <= 0:
                continue
            points[pd.Timestamp(date_text)] = float(nav)
        if len(rows) < page_size:
            break

    if not points:
        raise ValueError(f"cannot fetch nav history for {code}")
    s = pd.Series(points).sort_index()
    s.name = code
    return s


def fetch_fund_name(code: str, session: requests.Session) -> str:
    url = EASTMONEY_FUND_INFO_URL.format(code=code)
    params = {"v": str(int(dt.datetime.now().timestamp()))}
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://fundf10.eastmoney.com/",
        "Accept": "*/*",
    }

    resp = session.get(url, params=params, headers=headers, timeout=20)
    resp.raise_for_status()
    text = resp.text
    matched = re.search(r'fS_name\s*=\s*"([^"]+)"', text)
    if not matched:
        raise ValueError(f"Cannot parse fund name for {code}")
    return matched.group(1).strip()


def build_positions(transactions: List[Dict]) -> Tuple[List[Dict], Dict[int, float], float]:
    lots: List[Dict] = []
    oversold = 0.0

    for txn in sorted(transactions, key=lambda x: (x["date"], x["uid"])):
        if txn["action"] == "BUY":
            lots.append(
                {
                    "source_uid": txn["uid"],
                    "buy_date": txn["date"],
                    "buy_nav": float(txn["nav"] or 0.0),
                    "remaining_shares": float(txn["shares"]),
                }
            )
            continue

        remain_to_sell = float(txn["shares"])
        for lot in lots:
            if remain_to_sell <= 0:
                break
            if lot["remaining_shares"] <= 0:
                continue
            consume = min(lot["remaining_shares"], remain_to_sell)
            lot["remaining_shares"] -= consume
            remain_to_sell -= consume

        if remain_to_sell > 1e-12:
            oversold += remain_to_sell

    remain_map = {lot["source_uid"]: lot["remaining_shares"] for lot in lots}
    clean_lots = [x for x in lots if x["remaining_shares"] > 1e-12]
    return clean_lots, remain_map, oversold


class JijinUI(tk.Tk):
    def __init__(self, config_path: Path):
        super().__init__()
        self.title("\u57fa\u91d1\u8bb0\u5f55\u7ba1\u7406")
        self.geometry("680x500")
        self.minsize(640, 460)
        self._init_theme()

        self.config_path = config_path
        self.config: Dict = load_or_default(config_path)
        self.transactions: List[Dict] = []
        self.bond_transactions: List[Dict] = []
        self.sheet_fund_codes: List[str] = []
        self.next_uid = 1
        self.bond_next_uid = 1
        self.latest_nav_cache: Dict[str, Tuple[str, float]] = {}
        self.fund_name_cache: Dict[str, str] = {}
        self.session = requests.Session()
        self.tabs: Optional[ttk.Notebook] = None
        self.general_tab: Optional[ttk.Frame] = None
        self.stock_tab: Optional[ttk.Frame] = None
        self.bond_tab: Optional[ttk.Frame] = None
        self.analysis_tab: Optional[ttk.Frame] = None
        self.strategy_tab: Optional[ttk.Frame] = None
        self._last_tab_id = ""
        self._is_switching_tab = False
        self.saved_app_state: Dict = copy.deepcopy(self.config.get("app", {}))

        self.selected_fund_var = tk.StringVar()
        self.sort_column: Optional[str] = None
        self.sort_descending = True
        self.sortable_columns = {"date", "rate", "profit", "amount"}
        self.header_labels: Dict[str, str] = {}
        self.bond_sort_column: Optional[str] = None
        self.bond_sort_descending = True
        self.bond_sortable_columns = {"rate", "profit", "amount"}
        self.bond_header_labels: Dict[str, str] = {}
        self.fund_display_to_code: Dict[str, str] = {}
        self.rate_cell_meta: Dict[str, Optional[float]] = {}
        self.rate_overlay_labels: Dict[str, tk.Label] = {}
        self._overlay_job: Optional[str] = None
        self.bond_rate_cell_meta: Dict[str, Optional[float]] = {}
        self.ntfy_sent_cache: set[str] = set()
        self._startup_notify_done = False
        self.strategy_status_var = tk.StringVar(value="-")
        self.strategy_funds_var = tk.StringVar(value="")
        self.strategy_today_var = tk.StringVar(value="")
        self.strategy_history_days_var = tk.StringVar(value="260")
        self.strategy_bond_value_var = tk.StringVar(value="0")
        self.strategy_cash_value_var = tk.StringVar(value="0")
        self.strategy_stock_values_var = tk.StringVar(value="")
        self.strategy_trade_peak_var = tk.StringVar(value="-")
        self.strategy_trade_peak_date_var = tk.StringVar(value="-")
        self.strategy_param_vars: Dict[str, tk.Variable] = {}
        self.strategy_weights_preview_var = tk.StringVar(value="未配置（默认等权）")
        self.strategy_dd_gap_preview_var = tk.StringVar(value="未配置")
        self.strategy_reset_year_var = tk.StringVar(value="")
        self.strategy_reset_month_var = tk.StringVar(value="")
        self.strategy_reset_day_var = tk.StringVar(value="")
        self.strategy_params_text: Optional[tk.Text] = None
        self.strategy_state_text: Optional[tk.Text] = None
        self.strategy_output_cn_text: Optional[tk.Text] = None
        self.strategy_output_text: Optional[tk.Text] = None
        self.strategy_params_json_cache = "{}"
        self.strategy_state_json_cache = "{}"
        self.strategy_output_json_cache = ""
        self.strategy_run_btn: Optional[ttk.Button] = None
        self._strategy_running = False
        self._strategy_nav_series_cache: Dict[str, pd.Series] = {}
        self._strategy_nav_series_cached_at: Dict[str, dt.datetime] = {}
        self.vars: Dict[str, tk.Variable] = {
            "app.xlsx_path": tk.StringVar(),
            "app.threshold": tk.StringVar(),
            "app.ntfy_url": tk.StringVar(),
        }

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self.on_window_close)
        self.apply_config_to_form()
        self.load_transactions_from_doc(silent=True, fetch_latest_bond=False)
        self.after(80, self._run_startup_sync_and_notify_once)

    def _init_theme(self) -> None:
        self.colors = {
            "bg": "#F3F6FB",
            "surface": "#FFFFFF",
            "accent": "#2B5DB9",
            "accent_active": "#214D9D",
            "danger": "#C84B4B",
            "danger_active": "#A93B3B",
            "text": "#20283A",
            "muted": "#5F6C86",
            "tab": "#E6ECF7",
            "tab_active": "#FFFFFF",
            "table_alt": "#F8FAFF",
        }

        self.fonts = {
            "title": tkfont.Font(family="Microsoft YaHei UI", size=13, weight="bold"),
            "body": tkfont.Font(family="Microsoft YaHei UI", size=9),
            "small": tkfont.Font(family="Microsoft YaHei UI", size=8),
            "head": tkfont.Font(family="Microsoft YaHei UI", size=9, weight="bold"),
        }

        self.configure(bg=self.colors["bg"])
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure("App.TFrame", background=self.colors["bg"])
        style.configure("Card.TFrame", background=self.colors["surface"])

        style.configure("Body.TLabel", background=self.colors["bg"], foreground=self.colors["text"], font=self.fonts["body"])
        style.configure("Title.TLabel", background=self.colors["bg"], foreground=self.colors["text"], font=self.fonts["title"])
        style.configure(
            "Subtle.TLabel",
            background=self.colors["bg"],
            foreground=self.colors["muted"],
            font=self.fonts["small"],
        )
        style.configure(
            "Summary.TLabel",
            background=self.colors["surface"],
            foreground=self.colors["text"],
            font=self.fonts["body"],
        )
        style.configure(
            "SummaryHead.TLabel",
            background=self.colors["surface"],
            foreground=self.colors["muted"],
            font=self.fonts["small"],
        )
        style.configure(
            "SummaryValue.TLabel",
            background=self.colors["surface"],
            foreground=self.colors["text"],
            font=self.fonts["small"],
        )
        style.configure(
            "SummaryRateNeutral.TLabel",
            background=self.colors["surface"],
            foreground=self.colors["text"],
            font=self.fonts["small"],
        )
        style.configure(
            "SummaryRatePos.TLabel",
            background="#ECA9A7",
            foreground="#FFFFFF",
            font=self.fonts["small"],
        )
        style.configure(
            "SummaryRateNeg.TLabel",
            background="#97CFAD",
            foreground="#FFFFFF",
            font=self.fonts["small"],
        )

        style.configure("TNotebook", background=self.colors["bg"], borderwidth=0)
        style.configure("TNotebook", tabmargins=(0, 0, 0, 0))
        style.configure(
            "TNotebook.Tab",
            background=self.colors["tab"],
            foreground=self.colors["text"],
            padding=(10, 6),
            font=self.fonts["body"],
            relief="flat",
            borderwidth=0,
            width=9,
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", self.colors["tab"]), ("!selected", self.colors["tab"])],
            foreground=[("selected", self.colors["text"]), ("!selected", self.colors["text"])],
            relief=[("selected", "flat"), ("!selected", "flat")],
            font=[("selected", self.fonts["body"]), ("!selected", self.fonts["body"])],
            padding=[("selected", (10, 6)), ("!selected", (10, 6))],
            borderwidth=[("selected", 0), ("!selected", 0)],
        )

        style.configure("TEntry", padding=(4, 2), font=self.fonts["body"])
        style.configure("TCombobox", padding=(4, 2), font=self.fonts["body"])

        style.configure("TButton", padding=(6, 3), font=self.fonts["body"])
        style.configure("Toolbar.TButton", padding=(3, 1), font=self.fonts["small"])
        style.configure("Primary.TButton", background=self.colors["accent"], foreground="#FFFFFF")
        style.map("Primary.TButton", background=[("active", self.colors["accent_active"])])
        style.configure("Danger.TButton", background=self.colors["danger"], foreground="#FFFFFF")
        style.map("Danger.TButton", background=[("active", self.colors["danger_active"])])
        style.configure("ToolbarPrimary.TButton", padding=(3, 1), font=self.fonts["small"], background=self.colors["accent"], foreground="#FFFFFF")
        style.map("ToolbarPrimary.TButton", background=[("active", self.colors["accent_active"])])
        style.configure("ToolbarDanger.TButton", padding=(3, 1), font=self.fonts["small"], background=self.colors["danger"], foreground="#FFFFFF")
        style.map("ToolbarDanger.TButton", background=[("active", self.colors["danger_active"])])

        style.configure(
            "Summary.TLabelframe",
            background=self.colors["surface"],
            relief="solid",
            borderwidth=1,
        )
        style.configure(
            "Summary.TLabelframe.Label",
            background=self.colors["surface"],
            foreground=self.colors["text"],
            font=self.fonts["head"],
        )

        style.configure(
            "Data.Treeview",
            background=self.colors["surface"],
            fieldbackground=self.colors["surface"],
            foreground=self.colors["text"],
            rowheight=21,
            font=self.fonts["body"],
        )
        style.configure(
            "Data.Treeview.Heading",
            background="#E9EFFA",
            foreground=self.colors["text"],
            font=self.fonts["head"],
            relief="flat",
            padding=(4, 3),
        )
        style.map("Data.Treeview", background=[("selected", "#DCE7FF")], foreground=[("selected", self.colors["text"])])

    def _build_ui(self) -> None:
        root = ttk.Frame(self, style="App.TFrame", padding=(6, 5, 6, 5))
        root.pack(fill=tk.BOTH, expand=True)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)

        self.tabs = ttk.Notebook(root)
        self.tabs.grid(row=0, column=0, sticky="nsew")

        self.stock_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=6)
        self.bond_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=6)
        self.analysis_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=6)
        self.strategy_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=6)
        self.general_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=6)
        self.tabs.add(self.stock_tab, text="\u80a1\u7968\u57fa\u91d1")
        self.tabs.add(self.bond_tab, text="\u503a\u5377\u57fa\u91d1")
        self.tabs.add(self.analysis_tab, text="\u6c47\u603b\u5206\u6790")
        self.tabs.add(self.strategy_tab, text="\u7b56\u7565")
        self.tabs.add(self.general_tab, text="\u901a\u7528\u8bbe\u7f6e")

        self._build_fund_tab(self.stock_tab)
        self._build_bond_tab(self.bond_tab)
        self._build_analysis_tab(self.analysis_tab)
        self._build_strategy_tab(self.strategy_tab)
        self._build_general_tab(self.general_tab)
        self._last_tab_id = self.tabs.select()
        self.tabs.bind("<<NotebookTabChanged>>", self.on_tab_changed)

    def _build_fund_tab(self, frame: ttk.Frame) -> None:
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)

        bar = ttk.Frame(frame, style="Card.TFrame", padding=(6, 5))
        bar.grid(row=0, column=0, sticky="w", pady=(0, 5))
        bar.columnconfigure(5, weight=0)

        ttk.Label(bar, text="\u80a1\u7968\u57fa\u91d1", style="Body.TLabel").grid(row=0, column=0, sticky="w")
        self.fund_combo = ttk.Combobox(bar, textvariable=self.selected_fund_var, state="normal", width=20)
        self.fund_combo.grid(row=0, column=1, sticky="w", padx=(3, 2))
        self.fund_combo.bind("<<ComboboxSelected>>", lambda _: self.on_fund_selected())

        ttk.Button(bar, text="\u65b0\u589e\u57fa\u91d1", command=self.add_fund, style="Toolbar.TButton").grid(row=0, column=2, padx=0)
        ttk.Button(bar, text="\u65b0\u589e\u4e70\u5165", command=self.add_buy, style="ToolbarPrimary.TButton").grid(row=0, column=3, padx=0)
        ttk.Button(bar, text="\u5220\u9664\u8bb0\u5f55", command=self.delete_selected_records, style="ToolbarDanger.TButton").grid(row=0, column=4, padx=0)

        table_wrap = ttk.Frame(frame, style="Card.TFrame", padding=(5, 5))
        table_wrap.grid(row=1, column=0, sticky="nsew")
        table_wrap.columnconfigure(0, weight=1)
        table_wrap.rowconfigure(0, weight=1)

        cols = ("date", "shares", "nav", "rate", "profit", "amount")
        self.tx_tree = ttk.Treeview(table_wrap, columns=cols, show="headings", height=13, style="Data.Treeview")
        self.header_labels = {
            "date": "\u65e5\u671f",
            "shares": "\u4efd\u989d",
            "nav": "\u6210\u4ea4\u51c0\u503c",
            "rate": "\u76c8\u5229\u5229\u7387",
            "profit": "\u76c8\u5229\u989d",
            "amount": "\u91d1\u989d",
        }
        widths = {
            "date": 100,
            "shares": 85,
            "nav": 85,
            "rate": 85,
            "profit": 100,
            "amount": 100,
        }
        for c in cols:
            self.tx_tree.column(c, width=widths[c], anchor="center", stretch=True, minwidth=40)
        self._refresh_tree_headings()

        self.tx_tree.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(table_wrap, orient=tk.VERTICAL, command=self.tx_tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.tx_tree.config(yscrollcommand=scroll.set)
        self.tx_tree.tag_configure("row_even", background=self.colors["surface"])
        self.tx_tree.tag_configure("row_odd", background=self.colors["table_alt"])
        self.tx_tree.bind("<Button-3>", self.on_tree_right_click)
        self.tx_tree.bind("<Configure>", lambda _e: self._queue_rate_overlay_refresh())
        self.tx_tree.bind("<MouseWheel>", lambda _e: self._queue_rate_overlay_refresh())
        self.tx_tree.bind("<ButtonRelease-1>", lambda _e: self._queue_rate_overlay_refresh())
        scroll.bind("<B1-Motion>", lambda _e: self._queue_rate_overlay_refresh())
        scroll.bind("<ButtonRelease-1>", lambda _e: self._queue_rate_overlay_refresh())

        self.tx_menu = tk.Menu(self, tearoff=0)
        self.tx_menu.add_command(label="\u4fee\u6539", command=self.edit_selected_record)
        self.tx_menu.add_command(label="\u5356\u51fa", command=self.sell_selected_records)

        summary_box = ttk.LabelFrame(frame, text="\u5f53\u524d\u57fa\u91d1\u6c47\u603b", padding=2, style="Summary.TLabelframe")
        summary_box.grid(row=2, column=0, sticky="ew", pady=(4, 0))
        for i in range(6):
            summary_box.columnconfigure(i, weight=1, uniform="summary")

        headers = ["\u57fa\u91d1\u540d\u79f0", "\u6700\u65b0\u51c0\u503c", "\u6301\u4ed3\u4efd\u989d", "\u6210\u672c", "\u76c8\u5229\u989d", "\u76c8\u5229\u5229\u7387"]
        col_width_chars = [10, 10, 10, 10, 10, 10]
        for i, h in enumerate(headers):
            ttk.Label(summary_box, text=h, style="SummaryHead.TLabel", anchor="center", width=col_width_chars[i]).grid(row=0, column=i, sticky="ew", padx=0, pady=(0, 0))

        self.summary_fund_var = tk.StringVar(value="-")
        self.summary_latest_var = tk.StringVar(value="-")
        self.summary_shares_var = tk.StringVar(value="-")
        self.summary_cost_var = tk.StringVar(value="-")
        self.summary_profit_var = tk.StringVar(value="-")
        self.summary_rate_var = tk.StringVar(value="-")

        ttk.Label(summary_box, textvariable=self.summary_fund_var, style="SummaryValue.TLabel", anchor="center", width=col_width_chars[0]).grid(row=1, column=0, sticky="ew", padx=0, pady=(0, 0))
        ttk.Label(summary_box, textvariable=self.summary_latest_var, style="SummaryValue.TLabel", anchor="center", width=col_width_chars[1]).grid(row=1, column=1, sticky="ew", padx=0, pady=(0, 0))
        ttk.Label(summary_box, textvariable=self.summary_shares_var, style="SummaryValue.TLabel", anchor="center", width=col_width_chars[2]).grid(row=1, column=2, sticky="ew", padx=0, pady=(0, 0))
        ttk.Label(summary_box, textvariable=self.summary_cost_var, style="SummaryValue.TLabel", anchor="center", width=col_width_chars[3]).grid(row=1, column=3, sticky="ew", padx=0, pady=(0, 0))
        ttk.Label(summary_box, textvariable=self.summary_profit_var, style="SummaryValue.TLabel", anchor="center", width=col_width_chars[4]).grid(row=1, column=4, sticky="ew", padx=0, pady=(0, 0))
        self.summary_rate_label = ttk.Label(summary_box, textvariable=self.summary_rate_var, style="SummaryRateNeutral.TLabel", anchor="center")
        self.summary_rate_label.configure(width=col_width_chars[5])
        self.summary_rate_label.grid(row=1, column=5, sticky="ew", padx=0, pady=(0, 0))

    def _build_bond_tab(self, frame: ttk.Frame) -> None:
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)

        bar = ttk.Frame(frame, style="Card.TFrame", padding=(6, 5))
        bar.grid(row=0, column=0, sticky="w", pady=(0, 5))
        ttk.Label(bar, text="\u503a\u5377\u57fa\u91d1", style="Body.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Button(bar, text="\u65b0\u589e\u503a\u5377\u57fa\u91d1", command=self.add_bond_buy, style="ToolbarPrimary.TButton").grid(row=0, column=1, padx=(4, 0))
        ttk.Button(bar, text="\u5237\u65b0\u6536\u76ca", command=lambda: self.refresh_bond_view(fetch_latest=True), style="Toolbar.TButton").grid(row=0, column=2, padx=(2, 0))
        ttk.Button(bar, text="\u5220\u9664\u8bb0\u5f55", command=self.delete_selected_bond_records, style="ToolbarDanger.TButton").grid(row=0, column=3, padx=(2, 0))

        table_wrap = ttk.Frame(frame, style="Card.TFrame", padding=(5, 5))
        table_wrap.grid(row=1, column=0, sticky="nsew")
        table_wrap.columnconfigure(0, weight=1)
        table_wrap.rowconfigure(0, weight=1)

        cols = ("fund_name", "shares", "nav", "rate", "profit", "amount")
        self.bond_tree = ttk.Treeview(table_wrap, columns=cols, show="headings", height=13, style="Data.Treeview")
        self.bond_header_labels = {
            "fund_name": "\u57fa\u91d1\u540d\u79f0",
            "shares": "\u4efd\u989d",
            "nav": "\u6210\u4ea4\u51c0\u503c",
            "rate": "\u76c8\u5229\u5229\u7387",
            "profit": "\u76c8\u5229\u989d",
            "amount": "\u91d1\u989d",
        }
        widths = {
            "fund_name": 170,
            "shares": 110,
            "nav": 110,
            "rate": 120,
            "profit": 120,
            "amount": 120,
        }
        for c in cols:
            self.bond_tree.column(c, width=widths[c], anchor="center", stretch=True, minwidth=40)
        self._refresh_bond_tree_headings()
        self.bond_tree.grid(row=0, column=0, sticky="nsew")

        scroll = ttk.Scrollbar(table_wrap, orient=tk.VERTICAL, command=self.bond_tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.bond_tree.config(yscrollcommand=scroll.set)
        self.bond_tree.tag_configure("row_even", background=self.colors["surface"])
        self.bond_tree.tag_configure("row_odd", background=self.colors["table_alt"])
        self.bond_tree.bind("<Button-3>", self.on_bond_tree_right_click)

        self.bond_menu = tk.Menu(self, tearoff=0)
        self.bond_menu.add_command(label="\u4fee\u6539", command=self.edit_selected_bond_record)
        self.bond_menu.add_command(label="\u5356\u51fa", command=self.sell_selected_bond_records)

    def _build_analysis_tab(self, frame: ttk.Frame) -> None:
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(1, weight=1)

        ttk.Label(frame, text="\u8d44\u4ea7\u6c47\u603b\u5206\u6790", style="Title.TLabel").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 4))

        self.analysis_stock_amount_var = tk.StringVar(value="-")
        self.analysis_stock_rate_var = tk.StringVar(value="-")
        self.analysis_stock_profit_var = tk.StringVar(value="-")
        self.analysis_stock_ratio_var = tk.StringVar(value="-")
        self.analysis_bond_amount_var = tk.StringVar(value="-")
        self.analysis_bond_rate_var = tk.StringVar(value="-")
        self.analysis_bond_profit_var = tk.StringVar(value="-")
        self.analysis_bond_ratio_var = tk.StringVar(value="-")

        stock_box = ttk.LabelFrame(frame, text="\u80a1\u7968\u57fa\u91d1", padding=8, style="Summary.TLabelframe")
        stock_box.grid(row=1, column=0, sticky="nsew", padx=(0, 4))
        stock_box.columnconfigure(1, weight=1)
        ttk.Label(stock_box, text="\u603b\u91d1\u989d", style="Body.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(stock_box, textvariable=self.analysis_stock_amount_var, style="SummaryValue.TLabel").grid(row=0, column=1, sticky="e")
        ttk.Label(stock_box, text="\u603b\u76c8\u5229\u5229\u7387", style="Body.TLabel").grid(row=1, column=0, sticky="w", pady=(4, 0))
        ttk.Label(stock_box, textvariable=self.analysis_stock_rate_var, style="SummaryValue.TLabel").grid(row=1, column=1, sticky="e", pady=(4, 0))
        ttk.Label(stock_box, text="\u603b\u76c8\u5229\u989d", style="Body.TLabel").grid(row=2, column=0, sticky="w", pady=(4, 0))
        ttk.Label(stock_box, textvariable=self.analysis_stock_profit_var, style="SummaryValue.TLabel").grid(row=2, column=1, sticky="e", pady=(4, 0))
        ttk.Label(stock_box, text="\u5360\u6bd4", style="Body.TLabel").grid(row=3, column=0, sticky="w", pady=(4, 0))
        ttk.Label(stock_box, textvariable=self.analysis_stock_ratio_var, style="SummaryValue.TLabel").grid(row=3, column=1, sticky="e", pady=(4, 0))
        ttk.Button(stock_box, text="\u5c55\u5f00\u660e\u7ec6", command=self.show_stock_amount_details, style="Toolbar.TButton").grid(
            row=4, column=0, columnspan=2, sticky="w", pady=(8, 0)
        )

        bond_box = ttk.LabelFrame(frame, text="\u503a\u5377\u57fa\u91d1", padding=8, style="Summary.TLabelframe")
        bond_box.grid(row=1, column=1, sticky="nsew", padx=(4, 0))
        bond_box.columnconfigure(1, weight=1)
        ttk.Label(bond_box, text="\u603b\u91d1\u989d", style="Body.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(bond_box, textvariable=self.analysis_bond_amount_var, style="SummaryValue.TLabel").grid(row=0, column=1, sticky="e")
        ttk.Label(bond_box, text="\u603b\u76c8\u5229\u5229\u7387", style="Body.TLabel").grid(row=1, column=0, sticky="w", pady=(4, 0))
        ttk.Label(bond_box, textvariable=self.analysis_bond_rate_var, style="SummaryValue.TLabel").grid(row=1, column=1, sticky="e", pady=(4, 0))
        ttk.Label(bond_box, text="\u603b\u76c8\u5229\u989d", style="Body.TLabel").grid(row=2, column=0, sticky="w", pady=(4, 0))
        ttk.Label(bond_box, textvariable=self.analysis_bond_profit_var, style="SummaryValue.TLabel").grid(row=2, column=1, sticky="e", pady=(4, 0))
        ttk.Label(bond_box, text="\u5360\u6bd4", style="Body.TLabel").grid(row=3, column=0, sticky="w", pady=(4, 0))
        ttk.Label(bond_box, textvariable=self.analysis_bond_ratio_var, style="SummaryValue.TLabel").grid(row=3, column=1, sticky="e", pady=(4, 0))

    def _build_strategy_tab(self, frame: ttk.Frame) -> None:
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)

        base_box = ttk.LabelFrame(frame, text="\u8f93\u5165\uff08\u57fa\u7840\uff09", padding=8, style="Summary.TLabelframe")
        base_box.grid(row=0, column=0, sticky="ew")
        for i in range(6):
            base_box.columnconfigure(i, weight=1 if i in (1, 3, 5) else 0)

        ttk.Label(
            base_box,
            text="\u6807\u7684\u57fa\u91d1\u4f1a\u81ea\u52a8\u4ece\u201c\u80a1\u7968\u57fa\u91d1\u201d\u6301\u4ed3\u83b7\u53d6\uff1b\u503a\u57fa\u91d1\u989d\u4f1a\u81ea\u52a8\u4ece\u201c\u503a\u5377\u57fa\u91d1\u201d\u603b\u5e02\u503c\u8bfb\u53d6\u3002",
            style="Subtle.TLabel",
        ).grid(row=0, column=0, columnspan=6, sticky="w")
        ttk.Label(base_box, text="\u51b3\u7b56\u65e5", style="Body.TLabel").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(base_box, textvariable=self.strategy_today_var, width=16).grid(row=1, column=1, sticky="w", padx=(6, 0), pady=(6, 0))
        ttk.Label(base_box, text="\u56de\u770b\u5929\u6570", style="Body.TLabel").grid(row=1, column=2, sticky="w", pady=(6, 0), padx=(8, 0))
        ttk.Entry(base_box, textvariable=self.strategy_history_days_var, width=10).grid(row=1, column=3, sticky="w", padx=(6, 0), pady=(6, 0))
        ttk.Label(base_box, text="\u73b0\u91d1\u91d1\u989d", style="Body.TLabel").grid(row=1, column=4, sticky="w", pady=(6, 0), padx=(8, 0))
        ttk.Entry(base_box, textvariable=self.strategy_cash_value_var, width=16).grid(row=1, column=5, sticky="ew", padx=(6, 0), pady=(6, 0))

        btns = ttk.Frame(base_box, style="Card.TFrame")
        btns.grid(row=2, column=0, columnspan=6, sticky="ew", pady=(8, 0))
        btns.columnconfigure(1, weight=1)
        self.strategy_run_btn = ttk.Button(btns, text="\u6267\u884c\u51b3\u7b56", command=self._run_strategy_decision, style="ToolbarPrimary.TButton")
        self.strategy_run_btn.grid(row=0, column=0, sticky="w")
        ttk.Entry(btns, textvariable=self.strategy_status_var, state="readonly").grid(row=0, column=1, padx=(10, 0), sticky="ew")

        ttk.Label(base_box, text="交易峰值", style="Body.TLabel").grid(row=3, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(base_box, textvariable=self.strategy_trade_peak_var, state="readonly", width=16).grid(row=3, column=1, sticky="w", padx=(6, 0), pady=(6, 0))
        ttk.Label(base_box, text="更新日期", style="Body.TLabel").grid(row=3, column=2, sticky="w", pady=(6, 0), padx=(8, 0))
        ttk.Entry(base_box, textvariable=self.strategy_trade_peak_date_var, state="readonly", width=16).grid(row=3, column=3, sticky="w", padx=(6, 0), pady=(6, 0))

        params_box = ttk.LabelFrame(frame, text="策略参数（表单）", padding=6, style="Summary.TLabelframe")
        params_box.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
        params_box.columnconfigure(0, weight=1)
        params_box.rowconfigure(0, weight=1)
        self._build_strategy_params_form(params_box)

        self._strategy_reset_defaults(force=True)
        self._strategy_fill_from_portfolio()
        if self._strategy_load_ui_state():
            self.strategy_status_var.set("已加载上次“策略”参数配置。")
        else:
            self.strategy_status_var.set("点击“执行决策”后，计算完成会自动弹出中文解读结果窗口。")
        self._strategy_refresh_trade_peak_display_from_config()

    def _strategy_get_var(self, key: str, default, kind: str = "str") -> tk.Variable:
        if key in self.strategy_param_vars:
            return self.strategy_param_vars[key]
        if kind == "double":
            var = tk.DoubleVar(value=float(default))
        else:
            var = tk.StringVar(value=str(default))
        self.strategy_param_vars[key] = var
        return var

    def _strategy_add_ratio_field(
        self,
        frame: ttk.Frame,
        row: int,
        label: str,
        key: str,
        default: float,
        from_: float = 0.0,
        to: float = 1.0,
    ) -> None:
        _ = (from_, to)
        var = self._strategy_get_var(key, default, kind="str")
        ttk.Label(frame, text=label, style="Body.TLabel").grid(row=row, column=0, sticky="w", pady=(3, 0))
        ttk.Entry(frame, textvariable=var, width=10).grid(row=row, column=1, sticky="w", padx=(6, 0), pady=(3, 0))
        ttk.Label(frame, text="%", style="Subtle.TLabel").grid(row=row, column=2, sticky="w", padx=(4, 0), pady=(3, 0))

    def _strategy_add_entry_field(self, frame: ttk.Frame, row: int, label: str, key: str, default: str, width: int = 20) -> None:
        var = self._strategy_get_var(key, default, kind="str")
        ttk.Label(frame, text=label, style="Body.TLabel").grid(row=row, column=0, sticky="w", pady=(3, 0))
        ttk.Entry(frame, textvariable=var, width=width).grid(row=row, column=1, columnspan=2, sticky="ew", padx=(6, 0), pady=(3, 0))

    @staticmethod
    def _strategy_format_percent_num(value: float) -> str:
        return f"{float(value):.6f}".rstrip("0").rstrip(".")

    def _strategy_refresh_weights_preview(self) -> None:
        if self.strategy_weights_preview_var is None:
            return
        raw = ""
        var = self.strategy_param_vars.get("target_stock_fund_weights")
        if var is not None:
            raw = str(var.get() or "").strip()
        if not raw:
            self.strategy_weights_preview_var.set("未配置（默认等权）")
            return

        parsed = self._strategy_parse_code_float_map(raw)
        if not parsed:
            self.strategy_weights_preview_var.set("未配置（默认等权）")
            return

        parts: List[str] = []
        for code, val in sorted(parsed.items()):
            pct = float(val) * 100.0 if float(val) <= 1.0 else float(val)
            name = str(self.fund_name_cache.get(code, "") or "").strip()
            title = f"{name}({code})" if name else code
            parts.append(f"{title} {self._strategy_format_percent_num(pct)}%")
        self.strategy_weights_preview_var.set("；".join(parts))

    def _strategy_open_weights_dialog(self) -> None:
        funds = self.get_all_fund_codes()
        if not funds:
            messagebox.showwarning("提示", "当前“股票基金”暂无可配置的基金")
            return

        existing_raw = ""
        var = self.strategy_param_vars.get("target_stock_fund_weights")
        if var is not None:
            existing_raw = str(var.get() or "").strip()
        existing_map = self._strategy_parse_code_float_map(existing_raw) if existing_raw else {}
        existing_pct_map: Dict[str, float] = {}
        for code, val in existing_map.items():
            existing_pct_map[code] = float(val) * 100.0 if float(val) <= 1.0 else float(val)

        dialog = tk.Toplevel(self)
        dialog.title("配置股票子基权重")
        dialog.geometry("680x560")
        dialog.minsize(620, 420)
        dialog.transient(self)

        root = ttk.Frame(dialog, style="App.TFrame", padding=8)
        root.pack(fill=tk.BOTH, expand=True)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(1, weight=1)

        ttk.Label(root, text="请按百分比填写各基金权重，保存时合计必须为 100%", style="Subtle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 6))

        table = ttk.Frame(root, style="Card.TFrame")
        table.grid(row=1, column=0, sticky="nsew")
        table.columnconfigure(0, weight=1)
        table.rowconfigure(0, weight=1)

        canvas = tk.Canvas(table, highlightthickness=0, bg=self.colors["surface"])
        canvas.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(table, orient=tk.VERTICAL, command=canvas.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scroll.set)

        body = ttk.Frame(canvas, style="Card.TFrame")
        body_id = canvas.create_window((0, 0), window=body, anchor="nw")

        def _on_body_configure(_e) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(e) -> None:
            canvas.itemconfigure(body_id, width=e.width)

        body.bind("<Configure>", _on_body_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        ttk.Label(body, text="基金", style="SummaryHead.TLabel").grid(row=0, column=0, sticky="w", padx=(4, 0))
        ttk.Label(body, text="权重(%)", style="SummaryHead.TLabel").grid(row=0, column=1, sticky="w", padx=(8, 0))

        entry_vars: Dict[str, tk.StringVar] = {}
        for idx, code in enumerate(funds, start=1):
            name = str(self.fund_name_cache.get(code, "") or "").strip()
            title = f"{name}({code})" if name else code
            ttk.Label(body, text=title, style="Body.TLabel").grid(row=idx, column=0, sticky="w", padx=(4, 0), pady=(4, 0))
            value = existing_pct_map.get(code)
            var_pct = tk.StringVar(value="" if value is None else self._strategy_format_percent_num(value))
            entry_vars[code] = var_pct
            ttk.Entry(body, textvariable=var_pct, width=12).grid(row=idx, column=1, sticky="w", padx=(8, 0), pady=(4, 0))

        def fill_equal() -> None:
            if not funds:
                return
            each = 100.0 / float(len(funds))
            vals = [round(each, 6) for _ in funds]
            vals[-1] = round(100.0 - sum(vals[:-1]), 6)
            for code, v in zip(funds, vals):
                entry_vars[code].set(self._strategy_format_percent_num(v))

        def clear_default() -> None:
            tgt_var = self.strategy_param_vars.get("target_stock_fund_weights")
            if tgt_var is not None:
                tgt_var.set("")
            self._strategy_refresh_weights_preview()
            self._strategy_save_ui_state()
            self.strategy_status_var.set("已清空股票子基权重，后续按默认等权执行")
            dialog.destroy()

        def save_weights() -> None:
            weights_pct: Dict[str, float] = {}
            total = 0.0
            for code in funds:
                raw_text = str(entry_vars[code].get() or "").strip()
                if not raw_text:
                    continue
                num = to_float(raw_text, default=None)
                if num is None:
                    messagebox.showwarning("提示", f"{code} 的权重不是数字")
                    return
                val = float(num)
                if val < 0:
                    messagebox.showwarning("提示", f"{code} 的权重不能为负数")
                    return
                if val > 0:
                    weights_pct[code] = val
                    total += val

            if abs(total - 100.0) > 1e-6:
                messagebox.showwarning("提示", f"权重合计必须为 100%，当前为 {self._strategy_format_percent_num(total)}%")
                return

            tgt_var = self.strategy_param_vars.get("target_stock_fund_weights")
            if tgt_var is not None:
                tgt_var.set(json.dumps(weights_pct, ensure_ascii=False))
            self._strategy_refresh_weights_preview()
            self._strategy_save_ui_state()
            self.strategy_status_var.set("已更新股票子基权重（合计100%）")
            dialog.destroy()

        btns = ttk.Frame(root, style="App.TFrame")
        btns.grid(row=2, column=0, sticky="e", pady=(8, 0))
        ttk.Button(btns, text="均分", command=fill_equal, style="Toolbar.TButton").grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btns, text="默认等权", command=clear_default, style="Toolbar.TButton").grid(row=0, column=1, padx=(0, 6))
        ttk.Button(btns, text="保存", command=save_weights, style="ToolbarPrimary.TButton").grid(row=0, column=2, padx=(0, 6))
        ttk.Button(btns, text="取消", command=dialog.destroy, style="Toolbar.TButton").grid(row=0, column=3)

        self._center_dialog(dialog)
        dialog.wait_window()

    def _strategy_refresh_dd_gap_preview(self) -> None:
        if self.strategy_dd_gap_preview_var is None:
            return
        dd_var = self.strategy_param_vars.get("dd_stages")
        gap_var = self.strategy_param_vars.get("gap_fill")
        if dd_var is None or gap_var is None:
            self.strategy_dd_gap_preview_var.set("未配置")
            return

        dd_raw = str(dd_var.get() or "").strip()
        gap_raw = str(gap_var.get() or "").strip()
        if not dd_raw or not gap_raw:
            self.strategy_dd_gap_preview_var.set("未配置")
            return
        try:
            dd_vals = self._strategy_parse_percent_list(dd_raw, "回撤触发档位")
            gap_vals = self._strategy_parse_percent_list(gap_raw, "每档补差额比例")
        except Exception:
            self.strategy_dd_gap_preview_var.set("未配置")
            return
        if len(dd_vals) != 3 or len(gap_vals) != 3:
            self.strategy_dd_gap_preview_var.set("未配置（需3档）")
            return

        parts: List[str] = []
        for dd, gp in zip(dd_vals, gap_vals):
            dd_pct = self._strategy_format_percent_num(float(dd) * 100.0)
            gp_pct = self._strategy_format_percent_num(float(gp) * 100.0)
            parts.append(f"回撤{dd_pct}%→补差额{gp_pct}%")
        self.strategy_dd_gap_preview_var.set("；".join(parts))

    def _strategy_open_dd_gap_dialog(self) -> None:
        dd_var = self.strategy_param_vars.get("dd_stages")
        gap_var = self.strategy_param_vars.get("gap_fill")
        if dd_var is None or gap_var is None:
            return

        def _parse_pct_list_or_default(raw_text: str, defaults: List[float], field_name: str) -> List[float]:
            text = str(raw_text or "").strip()
            if not text:
                return list(defaults)
            try:
                vals = self._strategy_parse_percent_list(text, field_name)
                if len(vals) != 3:
                    return list(defaults)
                return [float(x) * 100.0 for x in vals]
            except Exception:
                return list(defaults)

        dd_defaults = [20.0, 30.0, 40.0]
        gap_defaults = [40.0, 70.0, 100.0]
        dd_init = _parse_pct_list_or_default(str(dd_var.get()), dd_defaults, "回撤触发档位")
        gap_init = _parse_pct_list_or_default(str(gap_var.get()), gap_defaults, "每档补差额比例")

        dialog = tk.Toplevel(self)
        dialog.title("配置回撤与补差额")
        dialog.geometry("520x300")
        dialog.minsize(460, 260)
        dialog.transient(self)

        root = ttk.Frame(dialog, style="App.TFrame", padding=10)
        root.pack(fill=tk.BOTH, expand=True)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(1, weight=1)

        ttk.Label(root, text="请填写3档回撤触发与补差额，单位均为百分比", style="Subtle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
        table = ttk.Frame(root, style="Card.TFrame")
        table.grid(row=1, column=0, sticky="nsew")
        for c in range(6):
            table.columnconfigure(c, weight=0)

        dd_entries: List[tk.StringVar] = []
        gap_entries: List[tk.StringVar] = []
        for i in range(3):
            dd_text = self._strategy_format_percent_num(dd_init[i])
            gap_text = self._strategy_format_percent_num(gap_init[i])
            v_dd = tk.StringVar(value=dd_text)
            v_gp = tk.StringVar(value=gap_text)
            dd_entries.append(v_dd)
            gap_entries.append(v_gp)

            ttk.Label(table, text=f"提示回撤{i + 1}", style="Body.TLabel").grid(row=i, column=0, sticky="w", padx=(8, 4), pady=(8 if i == 0 else 6, 0))
            ttk.Entry(table, textvariable=v_dd, width=10).grid(row=i, column=1, sticky="w", pady=(8 if i == 0 else 6, 0))
            ttk.Label(table, text="%", style="Subtle.TLabel").grid(row=i, column=2, sticky="w", padx=(4, 12), pady=(8 if i == 0 else 6, 0))
            ttk.Label(table, text="补差额", style="Body.TLabel").grid(row=i, column=3, sticky="w", pady=(8 if i == 0 else 6, 0))
            ttk.Entry(table, textvariable=v_gp, width=10).grid(row=i, column=4, sticky="w", pady=(8 if i == 0 else 6, 0))
            ttk.Label(table, text="%", style="Subtle.TLabel").grid(row=i, column=5, sticky="w", padx=(4, 0), pady=(8 if i == 0 else 6, 0))

        def save_pair_config() -> None:
            dd_pct: List[float] = []
            gap_pct: List[float] = []

            for idx in range(3):
                dd_raw = str(dd_entries[idx].get() or "").strip()
                gp_raw = str(gap_entries[idx].get() or "").strip()
                if not dd_raw or not gp_raw:
                    messagebox.showwarning("提示", f"第{idx + 1}档请完整填写")
                    return
                dd_num = to_float(dd_raw, default=None)
                gp_num = to_float(gp_raw, default=None)
                if dd_num is None or gp_num is None:
                    messagebox.showwarning("提示", f"第{idx + 1}档必须填写数字")
                    return
                dd_val = float(dd_num)
                gp_val = float(gp_num)
                if dd_val < 0 or gp_val < 0:
                    messagebox.showwarning("提示", f"第{idx + 1}档不能为负数")
                    return
                if dd_val > 100 or gp_val > 100:
                    messagebox.showwarning("提示", f"第{idx + 1}档不能大于100%")
                    return
                dd_pct.append(dd_val)
                gap_pct.append(gp_val)

            if not (dd_pct[0] < dd_pct[1] < dd_pct[2]):
                messagebox.showwarning("提示", "3档回撤触发值需要从小到大递增")
                return

            dd_var.set(",".join(f"{self._strategy_format_percent_num(x)}%" for x in dd_pct))
            gap_var.set(",".join(f"{self._strategy_format_percent_num(x)}%" for x in gap_pct))
            self._strategy_refresh_dd_gap_preview()
            self._strategy_save_ui_state()
            self.strategy_status_var.set("已更新回撤触发与补差额（3档）")
            dialog.destroy()

        btns = ttk.Frame(root, style="App.TFrame")
        btns.grid(row=2, column=0, sticky="e", pady=(10, 0))
        ttk.Button(btns, text="确定", command=save_pair_config, style="ToolbarPrimary.TButton").grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btns, text="取消", command=dialog.destroy, style="Toolbar.TButton").grid(row=0, column=1)

        self._center_dialog(dialog)
        dialog.wait_window()

    @staticmethod
    def _strategy_parse_percent_to_ratio(raw_value, field_name: str) -> float:
        text = str(raw_value or "").strip()
        has_pct = text.endswith("%")
        if has_pct:
            text = text[:-1].strip()
        num = to_float(text, default=None)
        if num is None:
            raise ValueError(f"{field_name} 必须是数字")
        val = float(num)
        if val < 0:
            raise ValueError(f"{field_name} 不能为负数")
        if val > 100:
            raise ValueError(f"{field_name} 不能大于 100%")
        # 占比类统一按“百分数输入”解析：1 => 1%
        return val / 100.0

    @staticmethod
    def _strategy_ratio_to_percent_text(value: object) -> str:
        num = to_float(value, default=None)
        if num is None:
            return ""
        pct = float(num) * 100.0
        s = f"{pct:.6f}".rstrip("0").rstrip(".")
        return f"{s}%"

    @staticmethod
    def _strategy_ratio_to_percent_value_text(value: object) -> str:
        num = to_float(value, default=None)
        if num is None:
            return ""
        pct = float(num) * 100.0
        return f"{pct:.6f}".rstrip("0").rstrip(".")

    @staticmethod
    def _strategy_parse_zh_date_to_iso(raw_value: object, field_name: str) -> Optional[str]:
        text = str(raw_value or "").strip()
        if (not text) or (text == "____年__月__日"):
            return None
        m = re.fullmatch(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return dt.date(y, mo, d).isoformat()
            except ValueError:
                raise ValueError(f"{field_name}不是有效日期")
        parsed = parse_date(text)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", parsed):
            return parsed
        raise ValueError(f"{field_name}格式应为 YYYY年MM月DD日")

    @staticmethod
    def _strategy_iso_date_to_zh_text(raw_value: object) -> str:
        parsed = parse_date(raw_value)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", parsed):
            return ""
        return f"{parsed[0:4]}年{parsed[5:7]}月{parsed[8:10]}日"

    def _strategy_collect_reset_date_iso(self) -> Optional[str]:
        y = str(self.strategy_reset_year_var.get() or "").strip()
        m = str(self.strategy_reset_month_var.get() or "").strip()
        d = str(self.strategy_reset_day_var.get() or "").strip()
        if not y and not m and not d:
            return None
        if not y or not m or not d:
            raise ValueError("仓位转化日期请完整填写 年/月/日")
        if (not y.isdigit()) or (not m.isdigit()) or (not d.isdigit()):
            raise ValueError("仓位转化日期必须是数字")
        try:
            return dt.date(int(y), int(m), int(d)).isoformat()
        except ValueError:
            raise ValueError("仓位转化日期不是有效日期")

    def _strategy_set_reset_date_from_value(self, raw_value: object) -> None:
        iso = ""
        parsed = parse_date(raw_value)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", parsed):
            iso = parsed
        else:
            try:
                alt = self._strategy_parse_zh_date_to_iso(raw_value, "仓位转化日期")
                if alt and re.fullmatch(r"\d{4}-\d{2}-\d{2}", alt):
                    iso = alt
            except Exception:
                iso = ""
        if iso:
            self.strategy_reset_year_var.set(iso[0:4])
            self.strategy_reset_month_var.set(str(int(iso[5:7])))
            self.strategy_reset_day_var.set(str(int(iso[8:10])))
        else:
            self.strategy_reset_year_var.set("")
            self.strategy_reset_month_var.set("")
            self.strategy_reset_day_var.set("")

    def _strategy_parse_percent_list(self, raw: str, field_name: str) -> List[float]:
        vals: List[float] = []
        for seg in re.split(r"[,;，；\s]+", str(raw or "").strip()):
            if not seg:
                continue
            vals.append(self._strategy_parse_percent_to_ratio(seg, field_name))
        return vals

    def _build_strategy_params_form(self, frame: ttk.Frame) -> None:
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        form = ttk.Frame(frame, style="App.TFrame", padding=(2, 2))
        form.grid(row=0, column=0, sticky="nsew")
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)

        left = ttk.LabelFrame(form, text="\u80a1\u7968\u4ed3\u4f4d/\u89e6\u53d1", style="Summary.TLabelframe", padding=6)
        right = ttk.LabelFrame(form, text="\u9ad8\u7ea7/\u7ea6\u675f", style="Summary.TLabelframe", padding=6)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        right.grid(row=0, column=1, sticky="nsew", padx=(4, 0))
        for col in (0, 1, 2):
            left.columnconfigure(col, weight=1 if col == 1 else 0)
            right.columnconfigure(col, weight=1 if col == 1 else 0)

        self._strategy_add_ratio_field(left, 0, "\u76ee\u6807\u80a1\u7968\u4ed3\u4f4d", "target_stock_w", "30")
        self._strategy_add_ratio_field(left, 1, "\u4ed3\u4f4d\u4e0b\u9650", "force_low_w", "20")
        self._strategy_add_ratio_field(left, 2, "\u4ed3\u4f4d\u4e0a\u9650", "force_high_w", "40")
        self._strategy_add_ratio_field(left, 3, "\u9ad8\u4ed3\u56de\u843d\u76ee\u6807", "force_high_target_w", "35")
        self._strategy_get_var("dd_stages", "20%,30%,40%", kind="str")
        self._strategy_get_var("gap_fill", "40%,70%,100%", kind="str")
        ttk.Label(left, text="回撤触发与每档补差额", style="Body.TLabel").grid(row=4, column=0, sticky="w", pady=(3, 0))
        ttk.Button(left, text="配置", command=self._strategy_open_dd_gap_dialog, style="Toolbar.TButton").grid(
            row=4, column=1, columnspan=2, sticky="w", padx=(6, 0), pady=(3, 0)
        )
        self._strategy_refresh_dd_gap_preview()

        freq_var = self._strategy_get_var("decision_freq", DECISION_FREQ_VALUE_TO_LABEL.get("weekly", "每周"), kind="str")
        ttk.Label(left, text="\u51b3\u7b56\u9891\u7387", style="Body.TLabel").grid(row=5, column=0, sticky="w", pady=(3, 0))
        ttk.Combobox(left, textvariable=freq_var, values=tuple(DECISION_FREQ_LABEL_TO_VALUE.keys()), state="readonly", width=16).grid(
            row=5, column=1, columnspan=2, sticky="w", padx=(6, 0), pady=(3, 0)
        )

        self._strategy_add_entry_field(right, 0, "\u56de\u64a4\u6eda\u52a8\u7a97\u53e3(\u7a7a=\u5168\u90e8)", "peak_rolling_days", "252")
        self._strategy_add_entry_field(right, 1, "\u4ed3\u4f4d\u8fc7\u4f4e\u51b7\u9759\u671f(\u5929)", "force_low_cooldown_days", "60")
        self._strategy_add_entry_field(right, 2, "\u503a\u57fa\u6700\u4f4e\u4fdd\u7559", "min_bond_value", "0")
        self._strategy_add_entry_field(right, 3, "\u80a1\u7968\u6700\u4f4e\u4fdd\u7559", "min_stock_total_value", "0")
        self._strategy_get_var("last_stage_reset_date", "", kind="str")
        ttk.Label(right, text="仓位转化日期", style="Body.TLabel").grid(row=4, column=0, sticky="w", pady=(3, 0))
        date_row = ttk.Frame(right, style="App.TFrame")
        date_row.grid(row=4, column=1, columnspan=2, sticky="w", padx=(6, 0), pady=(3, 0))
        ttk.Entry(date_row, textvariable=self.strategy_reset_year_var, width=4).grid(row=0, column=0, sticky="w")
        ttk.Label(date_row, text="年", style="Body.TLabel").grid(row=0, column=1, sticky="w", padx=(2, 6))
        ttk.Entry(date_row, textvariable=self.strategy_reset_month_var, width=2).grid(row=0, column=2, sticky="w")
        ttk.Label(date_row, text="月", style="Body.TLabel").grid(row=0, column=3, sticky="w", padx=(2, 6))
        ttk.Entry(date_row, textvariable=self.strategy_reset_day_var, width=2).grid(row=0, column=4, sticky="w")
        ttk.Label(date_row, text="日", style="Body.TLabel").grid(row=0, column=5, sticky="w", padx=(2, 0))
        self._strategy_get_var("target_stock_fund_weights", "", kind="str")
        ttk.Label(right, text="股票子基权重", style="Body.TLabel").grid(row=5, column=0, sticky="w", pady=(3, 0))
        ttk.Button(right, text="配置", command=self._strategy_open_weights_dialog, style="Toolbar.TButton").grid(
            row=5, column=1, columnspan=2, sticky="w", padx=(6, 0), pady=(3, 0)
        )
        self._strategy_refresh_weights_preview()

    @staticmethod
    def _strategy_parse_num_list(raw: str) -> List[float]:
        vals: List[float] = []
        for seg in re.split(r"[,;，；\s]+", str(raw or "").strip()):
            if not seg:
                continue
            num = to_float(seg, default=None)
            if num is None:
                raise ValueError(f"\u65e0\u6cd5\u89e3\u6790\u6570\u503c: {seg}")
            vals.append(float(num))
        return vals

    def _strategy_collect_params_from_form(self) -> Dict:
        if not self.strategy_param_vars:
            raw = self._strategy_text_get(self.strategy_params_text) or self.strategy_params_json_cache
            data = json.loads(raw) if raw else {}
            if not isinstance(data, dict):
                raise ValueError("params JSON \u5fc5\u987b\u662f\u5bf9\u8c61")
            return data

        def num(key: str, default: Optional[float] = None) -> float:
            val = self.strategy_param_vars[key].get()
            n = to_float(val, default=default)
            if n is None:
                raise ValueError(f"\u53c2\u6570 {key} \u5fc5\u987b\u662f\u6570\u5b57")
            return float(n)

        dd_stages = self._strategy_parse_percent_list(str(self.strategy_param_vars["dd_stages"].get()), "回撤触发档位")
        gap_fill = self._strategy_parse_percent_list(str(self.strategy_param_vars["gap_fill"].get()), "每档补差额比例")
        if len(dd_stages) != 3:
            raise ValueError("回撤触发档位需要 3 个数值，例如 20,30,40")
        if len(gap_fill) != 3:
            raise ValueError("每档补差额比例需要 3 个数值，例如 40,70,100")

        peak_text = str(self.strategy_param_vars["peak_rolling_days"].get()).strip()
        cooldown_text = str(self.strategy_param_vars["force_low_cooldown_days"].get()).strip()
        freq_label = str(self.strategy_param_vars["decision_freq"].get()).strip() or DECISION_FREQ_VALUE_TO_LABEL.get("weekly", "每周")
        freq = DECISION_FREQ_LABEL_TO_VALUE.get(freq_label, freq_label)
        if freq not in {"daily", "weekly", "monthly", "quarterly"}:
            raise ValueError("决策频率必须是 每日/每周/每月/每季度")

        weights_text = str(self.strategy_param_vars["target_stock_fund_weights"].get()).strip()
        weights = self._strategy_parse_code_float_map(weights_text) if weights_text else None

        peak_value: Optional[int] = None
        if peak_text != "":
            peak_num = int(to_float(peak_text, default=0) or 0)
            if peak_num <= 0:
                raise ValueError("回撤滚动窗口必须为正整数，或者留空")
            peak_value = peak_num
        cooldown_value = int(to_float(cooldown_text, default=60) or 60)
        if cooldown_value < 0:
            raise ValueError("低仓冷静期天数不能为负数")
        reset_date_value = self._strategy_collect_reset_date_iso()
        self.strategy_param_vars["last_stage_reset_date"].set(reset_date_value or "")

        target_stock_w = self._strategy_parse_percent_to_ratio(self.strategy_param_vars["target_stock_w"].get(), "目标股票仓位")
        force_low_w = self._strategy_parse_percent_to_ratio(self.strategy_param_vars["force_low_w"].get(), "仓位下限")
        force_high_w = self._strategy_parse_percent_to_ratio(self.strategy_param_vars["force_high_w"].get(), "仓位上限")
        force_high_target_w = self._strategy_parse_percent_to_ratio(
            self.strategy_param_vars["force_high_target_w"].get(),
            "高仓回落目标",
        )
        for ratio_name, ratio_val in [
            ("目标股票仓位", target_stock_w),
            ("仓位下限", force_low_w),
            ("仓位上限", force_high_w),
            ("高仓回落目标", force_high_target_w),
        ]:
            if ratio_val > 1:
                raise ValueError(f"{ratio_name} 不能大于 100%")

        if weights:
            # 股票子基权重统一按百分数输入解析：1 => 1%
            weights = {k: (v / 100.0) for k, v in weights.items() if v > 0}

        return {
            "target_stock_w": target_stock_w,
            "force_low_w": force_low_w,
            "force_high_w": force_high_w,
            "force_high_target_w": force_high_target_w,
            "dd_stages": [float(x) for x in dd_stages],
            "gap_fill": [float(x) for x in gap_fill],
            "peak_rolling_days": peak_value,
            "force_low_cooldown_days": cooldown_value,
            "decision_freq": freq,
            "min_bond_value": num("min_bond_value", default=0.0),
            "min_stock_total_value": num("min_stock_total_value", default=0.0),
            "last_stage_reset_date": reset_date_value,
            "target_stock_fund_weights": weights if weights else None,
        }

    def _strategy_apply_params_to_form(self, params_dict: Dict) -> None:
        if not self.strategy_param_vars:
            return

        ratio_keys = ["target_stock_w", "force_low_w", "force_high_w", "force_high_target_w"]
        for key in ratio_keys:
            if key in params_dict and params_dict[key] not in ("", None):
                try:
                    self.strategy_param_vars[key].set(self._strategy_ratio_to_percent_value_text(params_dict[key]))
                except Exception:
                    pass

        num_keys = ["min_bond_value", "min_stock_total_value"]
        for key in num_keys:
            if key in params_dict and params_dict[key] not in ("", None):
                try:
                    self.strategy_param_vars[key].set(float(params_dict[key]))
                except Exception:
                    pass

        if "dd_stages" in params_dict and params_dict["dd_stages"] is not None:
            self.strategy_param_vars["dd_stages"].set(",".join(self._strategy_ratio_to_percent_text(x) for x in list(params_dict["dd_stages"])))
        if "gap_fill" in params_dict and params_dict["gap_fill"] is not None:
            self.strategy_param_vars["gap_fill"].set(",".join(self._strategy_ratio_to_percent_text(x) for x in list(params_dict["gap_fill"])))
        if "peak_rolling_days" in params_dict:
            val = params_dict["peak_rolling_days"]
            self.strategy_param_vars["peak_rolling_days"].set("" if val in ("", None) else str(int(float(val))))
        if "force_low_cooldown_days" in params_dict and params_dict["force_low_cooldown_days"] not in ("", None):
            self.strategy_param_vars["force_low_cooldown_days"].set(str(int(float(params_dict["force_low_cooldown_days"]))))
        if "decision_freq" in params_dict and params_dict["decision_freq"]:
            raw_freq = str(params_dict["decision_freq"])
            self.strategy_param_vars["decision_freq"].set(DECISION_FREQ_VALUE_TO_LABEL.get(raw_freq, raw_freq))
        if "last_stage_reset_date" in params_dict:
            reset_iso = parse_date(params_dict.get("last_stage_reset_date"))
            self.strategy_param_vars["last_stage_reset_date"].set(reset_iso if re.fullmatch(r"\d{4}-\d{2}-\d{2}", reset_iso) else "")
            self._strategy_set_reset_date_from_value(params_dict.get("last_stage_reset_date"))
        else:
            self.strategy_param_vars["last_stage_reset_date"].set("")
            self._strategy_set_reset_date_from_value(None)
        if "target_stock_fund_weights" in params_dict:
            weights = params_dict["target_stock_fund_weights"]
            if isinstance(weights, dict) and weights:
                self.strategy_param_vars["target_stock_fund_weights"].set(json.dumps(weights, ensure_ascii=False))
            else:
                self.strategy_param_vars["target_stock_fund_weights"].set("")
        self._strategy_refresh_dd_gap_preview()
        self._strategy_refresh_weights_preview()

    def _strategy_sync_params_text_from_form(self) -> None:
        try:
            data = self._strategy_collect_params_from_form()
            self.strategy_params_json_cache = json.dumps(data, ensure_ascii=False, indent=2)
            self._strategy_text_set(self.strategy_params_text, self.strategy_params_json_cache)
            self.strategy_status_var.set("\u5df2\u5c06\u8868\u5355\u53c2\u6570\u540c\u6b65\u5230 params JSON")
        except Exception as exc:
            self.strategy_status_var.set(f"\u540c\u6b65\u5931\u8d25: {exc}")

    def _strategy_load_params_form_from_text(self) -> None:
        try:
            raw = self._strategy_text_get(self.strategy_params_text) or self.strategy_params_json_cache
            data = json.loads(raw) if raw else {}
            if not isinstance(data, dict):
                raise ValueError("params JSON \u5fc5\u987b\u662f\u5bf9\u8c61")
            self._strategy_apply_params_to_form(data)
            self.strategy_params_json_cache = json.dumps(data, ensure_ascii=False, indent=2)
            self.strategy_status_var.set("\u5df2\u4ece params JSON \u52a0\u8f7d\u5230\u8868\u5355")
        except Exception as exc:
            self.strategy_status_var.set(f"\u52a0\u8f7d\u5931\u8d25: {exc}")

    @staticmethod
    def _strategy_text_get(widget: Optional[tk.Text]) -> str:
        if widget is None:
            return ""
        return widget.get("1.0", tk.END).strip()

    @staticmethod
    def _strategy_text_set(widget: Optional[tk.Text], value: str) -> None:
        if widget is None:
            return
        widget.delete("1.0", tk.END)
        widget.insert("1.0", value)

    @staticmethod
    def _strategy_parse_funds(raw: str) -> List[str]:
        codes: List[str] = []
        for token in re.split(r"[\s,;，；]+", str(raw or "").strip()):
            code = normalize_fund_code(token)
            if code and code not in codes:
                codes.append(code)
        return codes

    def _strategy_reset_defaults(self, force: bool = False) -> None:
        params_text = self._strategy_text_get(self.strategy_params_text) or self.strategy_params_json_cache
        state_text = self._strategy_text_get(self.strategy_state_text) or self.strategy_state_json_cache
        params_dict = {
            "target_stock_w": 0.30,
            "force_low_w": 0.20,
            "force_high_w": 0.40,
            "force_high_target_w": 0.35,
            "dd_stages": [0.20, 0.30, 0.40],
            "gap_fill": [0.40, 0.70, 1.00],
            "peak_rolling_days": 252,
            "force_low_cooldown_days": 60,
            "decision_freq": "weekly",
            "min_bond_value": 0.0,
            "min_stock_total_value": 0.0,
            "last_stage_reset_date": None,
            "target_stock_fund_weights": None,
            "new_high_epsilon": 1e-10,
        }
        if force or not params_text:
            self.strategy_params_json_cache = json.dumps(params_dict, ensure_ascii=False, indent=2)
            self._strategy_text_set(self.strategy_params_text, self.strategy_params_json_cache)
            self._strategy_apply_params_to_form(params_dict)
        else:
            try:
                parsed = json.loads(params_text)
                if isinstance(parsed, dict):
                    self._strategy_apply_params_to_form(parsed)
                    self.strategy_params_json_cache = json.dumps(parsed, ensure_ascii=False, indent=2)
                else:
                    self._strategy_apply_params_to_form(params_dict)
                    self.strategy_params_json_cache = json.dumps(params_dict, ensure_ascii=False, indent=2)
            except Exception:
                self._strategy_apply_params_to_form(params_dict)
                self.strategy_params_json_cache = json.dumps(params_dict, ensure_ascii=False, indent=2)
        if force or not state_text:
            state_dict = {
                "dd_triggered": [],
                "cooldown_until_date": None,
                "last_decision_date": None,
                "pending_reset_date": None,
                "last_stage_wait_conversion": False,
                "use_trade_peak_for_dd": False,
            }
            self.strategy_state_json_cache = json.dumps(state_dict, ensure_ascii=False, indent=2)
            self._strategy_text_set(self.strategy_state_text, self.strategy_state_json_cache)
        else:
            try:
                parsed_state = json.loads(state_text)
                if isinstance(parsed_state, dict):
                    self.strategy_state_json_cache = json.dumps(parsed_state, ensure_ascii=False, indent=2)
                else:
                    raise ValueError("state JSON \u5fc5\u987b\u662f\u5bf9\u8c61")
            except Exception:
                self.strategy_state_json_cache = json.dumps(
                    {
                        "dd_triggered": [],
                        "cooldown_until_date": None,
                        "last_decision_date": None,
                        "pending_reset_date": None,
                        "last_stage_wait_conversion": False,
                        "use_trade_peak_for_dd": False,
                    },
                    ensure_ascii=False,
                    indent=2,
                )

    @staticmethod
    def _strategy_parse_code_float_map(raw: str) -> Dict[str, float]:
        def parse_num_allow_percent(value_text: object) -> Optional[float]:
            s = str(value_text or "").strip()
            if not s:
                return None
            has_pct = s.endswith("%")
            if has_pct:
                s = s[:-1].strip()
            n = to_float(s, default=None)
            if n is None:
                return None
            v = float(n)
            if has_pct:
                return v / 100.0
            return v

        text = str(raw or "").strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
            if not isinstance(parsed, dict):
                raise ValueError("JSON must be an object")
            result: Dict[str, float] = {}
            for k, v in parsed.items():
                code = normalize_fund_code(k)
                num = parse_num_allow_percent(v)
                if code and num is not None:
                    result[code] = float(num)
            return result
        except json.JSONDecodeError:
            result = {}
            for seg in re.split(r"[,;，；]+", text):
                if not seg.strip():
                    continue
                if ":" in seg:
                    k, v = seg.split(":", 1)
                elif "=" in seg:
                    k, v = seg.split("=", 1)
                else:
                    continue
                code = normalize_fund_code(k.strip())
                num = parse_num_allow_percent(v.strip())
                if code and num is not None:
                    result[code] = float(num)
            return result

    def _strategy_current_stock_shares(self, funds: List[str]) -> Dict[str, float]:
        shares_map: Dict[str, float] = {code: 0.0 for code in funds}
        by_fund: Dict[str, List[Dict]] = {}
        for txn in self.transactions:
            code = normalize_fund_code(txn.get("fund_code"))
            if code in shares_map:
                by_fund.setdefault(code, []).append(txn)
        for code in funds:
            lots, _, _ = build_positions(sorted(by_fund.get(code, []), key=lambda x: (x["date"], x["uid"])))
            shares_map[code] = float(sum(float(x["remaining_shares"]) for x in lots))
        return shares_map

    def _strategy_fill_from_portfolio(self) -> None:
        funds = self.get_all_fund_codes()
        if not self.strategy_today_var.get().strip():
            self.strategy_today_var.set(dt.date.today().isoformat())
        bond_market = float(self._compute_bond_portfolio_summary().get("market", 0.0))
        self.strategy_bond_value_var.set(f"{bond_market:.2f}")
        if not self.strategy_cash_value_var.get().strip():
            self.strategy_cash_value_var.set("0")

        shares_map = self._strategy_current_stock_shares(funds)
        stock_values = {}
        for code in funds:
            latest = self.get_latest_nav_cached(code)
            nav = float(latest[1]) if latest and latest[1] is not None else 0.0
            stock_values[code] = float(shares_map.get(code, 0.0)) * nav
        self.strategy_stock_values_var.set(json.dumps(stock_values, ensure_ascii=False))
        self.strategy_status_var.set(f"\u5df2\u6309\u5f53\u524d\u6301\u4ed3\u586b\u5145\u7b56\u7565\u8f93\u5165\uff08\u503a\u57fa={bond_market:.2f}\uff09")

    @staticmethod
    def _strategy_ui_state_path() -> Path:
        return Path("data") / STRATEGY_UI_STATE_FILE

    def _strategy_save_ui_state(self) -> None:
        try:
            path = self._strategy_ui_state_path()
            path.parent.mkdir(parents=True, exist_ok=True)
            param_raw = {k: str(v.get()) for k, v in self.strategy_param_vars.items()}
            payload: Dict = {
                "saved_at": dt.datetime.now().isoformat(timespec="seconds"),
                "base": {
                    "today": str(self.strategy_today_var.get() or "").strip(),
                    "history_days": str(self.strategy_history_days_var.get() or "").strip(),
                    "cash_value": str(self.strategy_cash_value_var.get() or "").strip(),
                },
                "date_parts": {
                    "year": str(self.strategy_reset_year_var.get() or "").strip(),
                    "month": str(self.strategy_reset_month_var.get() or "").strip(),
                    "day": str(self.strategy_reset_day_var.get() or "").strip(),
                },
                "param_raw": param_raw,
                "state_json": self._strategy_text_get(self.strategy_state_text) or self.strategy_state_json_cache,
                "params_json": self._strategy_text_get(self.strategy_params_text) or self.strategy_params_json_cache,
            }
            try:
                payload["params_parsed"] = self._strategy_collect_params_from_form()
            except Exception:
                pass
            path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass

    def _strategy_load_ui_state(self) -> bool:
        path = self._strategy_ui_state_path()
        if not path.exists():
            return False
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(raw, dict):
                return False
        except Exception:
            return False

        try:
            base = raw.get("base") or {}
            if isinstance(base, dict):
                today = str(base.get("today", "") or "").strip()
                hist = str(base.get("history_days", "") or "").strip()
                cash = str(base.get("cash_value", "") or "").strip()
                if today:
                    self.strategy_today_var.set(today)
                if hist:
                    self.strategy_history_days_var.set(hist)
                if cash:
                    self.strategy_cash_value_var.set(cash)

            param_raw = raw.get("param_raw") or {}
            if isinstance(param_raw, dict):
                for k, v in param_raw.items():
                    if k in self.strategy_param_vars:
                        self.strategy_param_vars[k].set(str(v))

            date_parts = raw.get("date_parts") or {}
            if isinstance(date_parts, dict):
                self.strategy_reset_year_var.set(str(date_parts.get("year", "") or "").strip())
                self.strategy_reset_month_var.set(str(date_parts.get("month", "") or "").strip())
                self.strategy_reset_day_var.set(str(date_parts.get("day", "") or "").strip())
            if not (
                self.strategy_reset_year_var.get().strip()
                or self.strategy_reset_month_var.get().strip()
                or self.strategy_reset_day_var.get().strip()
            ):
                if isinstance(param_raw, dict):
                    self._strategy_set_reset_date_from_value(param_raw.get("last_stage_reset_date"))

            state_json = raw.get("state_json")
            if isinstance(state_json, str) and state_json.strip():
                self.strategy_state_json_cache = state_json
                self._strategy_text_set(self.strategy_state_text, state_json)
            params_json = raw.get("params_json")
            if isinstance(params_json, str) and params_json.strip():
                self.strategy_params_json_cache = params_json
                self._strategy_text_set(self.strategy_params_text, params_json)

            self._strategy_refresh_dd_gap_preview()
            self._strategy_refresh_weights_preview()
            return True
        except Exception:
            return False

    def _strategy_peak_config_path(self) -> Path:
        base_dir = self.config_path.parent if self.config_path else Path("config")
        return base_dir / STRATEGY_PEAK_CONFIG_FILE

    def _strategy_set_trade_peak_display(self, peak_value: Optional[float], peak_date: str) -> None:
        if peak_value is None:
            self.strategy_trade_peak_var.set("-")
        else:
            self.strategy_trade_peak_var.set(f"{float(peak_value):.6f}")
        d = parse_date(peak_date)
        self.strategy_trade_peak_date_var.set(d if re.fullmatch(r"\d{4}-\d{2}-\d{2}", d) else "-")

    def _strategy_refresh_trade_peak_display_from_config(self) -> None:
        cfg = self._strategy_load_peak_config()
        peak_value = to_float(cfg.get("trade_peak_value"), default=None)
        peak_date = str(cfg.get("trade_peak_update_date", "") or "")
        if peak_value is None:
            peak_value = to_float(cfg.get("peak_value"), default=None)
            peak_date = str(cfg.get("peak_date", "") or peak_date)
        self._strategy_set_trade_peak_display(peak_value, peak_date)

    def _strategy_load_peak_config(self) -> Dict:
        path = self._strategy_peak_config_path()
        if not path.exists():
            return {}
        try:
            raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
            if not isinstance(raw, dict):
                return {}
            peak_value = to_float(raw.get("peak_value"), default=None)
            peak_date = parse_date(raw.get("peak_date"))
            return {
                "peak_value": float(peak_value) if peak_value is not None else None,
                "peak_date": peak_date if peak_date else "",
                "pending_reset_date": parse_date(raw.get("pending_reset_date")),
                "trade_peak_value": to_float(raw.get("trade_peak_value"), default=None),
                "trade_peak_update_date": parse_date(raw.get("trade_peak_update_date")),
                "reason": str(raw.get("reason", "") or ""),
                "updated_at": str(raw.get("updated_at", "") or ""),
            }
        except Exception:
            return {}

    def _strategy_save_peak_config(self, data: Dict) -> None:
        path = self._strategy_peak_config_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "peak_value": float(to_float(data.get("peak_value"), default=0.0) or 0.0),
            "peak_date": str(data.get("peak_date", "") or ""),
            "pending_reset_date": str(data.get("pending_reset_date", "") or ""),
            "trade_peak_value": to_float(data.get("trade_peak_value"), default=None),
            "trade_peak_update_date": str(data.get("trade_peak_update_date", "") or ""),
            "reason": str(data.get("reason", "") or ""),
            "updated_at": str(data.get("updated_at", "") or ""),
        }
        path.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")

    def _strategy_prepare_basket_with_peak_config(self, basket: pd.Series, today: pd.Timestamp) -> Tuple[pd.Series, Dict, str]:
        cfg = self._strategy_load_peak_config()
        source = "rolling_peak"
        basket_used = basket
        peak_date = str(cfg.get("peak_date", "") or "").strip()
        if peak_date:
            start_ts = pd.Timestamp(peak_date)
            eligible = basket.index[basket.index >= start_ts]
            if len(eligible) > 0 and eligible[0] <= today:
                effective_ts = eligible[0]
                basket_used = basket.loc[effective_ts:]
                cfg["effective_peak_date"] = effective_ts.date().isoformat()
                source = "peak_config"
            else:
                source = "peak_config_ignored"
        return basket_used, cfg, source

    def _strategy_update_peak_config(
        self,
        result: Dict,
        basket_used: pd.Series,
        today: pd.Timestamp,
        params: DualTriggerParams,
        cfg_before: Dict,
        source: str,
    ) -> Dict:
        metrics = (result or {}).get("metrics") or {}
        action = str((result or {}).get("action", "") or "")

        window = basket_used.loc[:today]
        if params.peak_rolling_days is not None:
            window = window.tail(int(params.peak_rolling_days))

        decision_peak_date = ""
        decision_peak_value: Optional[float] = None
        if not window.empty:
            peak_idx = window.idxmax()
            decision_peak_date = peak_idx.date().isoformat()
            decision_peak_value = float(window.loc[peak_idx])

        cfg_after = dict(cfg_before or {})
        config_updated = False
        update_reason = ""
        new_state = (result or {}).get("new_state") or {}
        pending_reset_date = str(new_state.get("pending_reset_date", "") or "").strip()
        manual_reset_applied_date = str(metrics.get("manual_reset_applied_date", "") or "").strip()
        manual_reset_applied_peak = to_float(metrics.get("manual_reset_applied_peak"), default=None)
        trade_peak_val = to_float(metrics.get("peak_trade"), default=None)
        trade_peak_date = today.date().isoformat()

        if trade_peak_val is not None:
            cfg_after["trade_peak_value"] = float(trade_peak_val)
            cfg_after["trade_peak_update_date"] = trade_peak_date
            cfg_after["updated_at"] = dt.datetime.now().isoformat(timespec="seconds")
            self._strategy_save_peak_config(cfg_after)

        if manual_reset_applied_date and manual_reset_applied_peak is not None:
            cfg_after = {
                "peak_value": float(manual_reset_applied_peak),
                "peak_date": manual_reset_applied_date,
                "pending_reset_date": "",
                "trade_peak_value": cfg_after.get("trade_peak_value"),
                "trade_peak_update_date": cfg_after.get("trade_peak_update_date", trade_peak_date),
                "reason": "manual_reset_applied",
                "updated_at": dt.datetime.now().isoformat(timespec="seconds"),
            }
            self._strategy_save_peak_config(cfg_after)
            config_updated = True
            update_reason = f"按仓位转化日期 {self._strategy_iso_date_to_zh_text(manual_reset_applied_date) or manual_reset_applied_date} 完成峰值重置"

        cfg_pending_before = str(cfg_after.get("pending_reset_date", "") or "").strip()
        if (not config_updated) and (pending_reset_date != cfg_pending_before):
            cfg_after = dict(cfg_after or {})
            cfg_after["pending_reset_date"] = pending_reset_date
            cfg_after["reason"] = "dd_last_stage_wait_manual_reset" if pending_reset_date else "pending_reset_cleared"
            cfg_after["updated_at"] = dt.datetime.now().isoformat(timespec="seconds")
            self._strategy_save_peak_config(cfg_after)
            config_updated = True
            if pending_reset_date:
                update_reason = (
                    "已记录仓位转化日期 "
                    f"{self._strategy_iso_date_to_zh_text(pending_reset_date) or pending_reset_date}，到期后生效"
                )
            else:
                update_reason = "已清空待执行仓位转化日期"

        last_stage = max(list(params.dd_stages) or [0.0])
        fired_stage = to_float(metrics.get("dd_stage_fired"), default=None)
        is_last_stage_fire = action == "DD_ADD" and fired_stage is not None and abs(float(fired_stage) - float(last_stage)) <= 1e-12

        if is_last_stage_fire and not config_updated:
            if pending_reset_date:
                cfg_after = dict(cfg_after or {})
                cfg_after["pending_reset_date"] = pending_reset_date
                cfg_after["reason"] = "dd_last_stage_wait_manual_reset"
                cfg_after["updated_at"] = dt.datetime.now().isoformat(timespec="seconds")
                self._strategy_save_peak_config(cfg_after)
                config_updated = True
                update_reason = (
                    "已触发最后回撤档，已记录仓位转化日期 "
                    f"{self._strategy_iso_date_to_zh_text(pending_reset_date) or pending_reset_date}，到期后生效"
                )
            else:
                update_reason = "已触发最后回撤档，但未填写仓位转化日期，暂不重置峰值"

        if not config_updated and not is_last_stage_fire:
            old_peak = to_float(cfg_after.get("peak_value"), default=None)
            old_date = str(cfg_after.get("peak_date", "") or "")
            should_init = (not old_date) and (decision_peak_value is not None) and bool(decision_peak_date)
            should_new_high = (
                decision_peak_value is not None
                and old_peak is not None
                and decision_peak_value > old_peak + 1e-12
                and bool(decision_peak_date)
            )
            if should_init or should_new_high:
                cfg_after = {
                    "peak_value": float(decision_peak_value or 0.0),
                    "peak_date": decision_peak_date,
                    "reason": "new_high_update" if should_new_high else "init_from_decision_peak",
                    "trade_peak_value": cfg_after.get("trade_peak_value"),
                    "trade_peak_update_date": cfg_after.get("trade_peak_update_date", trade_peak_date),
                    "updated_at": dt.datetime.now().isoformat(timespec="seconds"),
                }
                self._strategy_save_peak_config(cfg_after)
                config_updated = True
                update_reason = "检测到峰值更新，已写入配置文件"

        effective_pending_reset = pending_reset_date or str(cfg_after.get("pending_reset_date", "") or "")
        return {
            "source": source,
            "decision_peak_value": decision_peak_value,
            "decision_peak_date": decision_peak_date,
            "config_peak_value": to_float(cfg_after.get("peak_value"), default=None),
            "config_peak_date": str(cfg_after.get("peak_date", "") or ""),
            "config_pending_reset_date": str(cfg_after.get("pending_reset_date", "") or ""),
            "trade_peak_value": to_float(cfg_after.get("trade_peak_value"), default=trade_peak_val),
            "trade_peak_update_date": str(cfg_after.get("trade_peak_update_date", "") or trade_peak_date),
            "config_reason": str(cfg_after.get("reason", "") or ""),
            "config_updated_at": str(cfg_after.get("updated_at", "") or ""),
            "config_updated": config_updated,
            "update_reason": update_reason,
            "pending_reset_date": effective_pending_reset,
            "manual_reset_applied_date": manual_reset_applied_date,
            "config_path": str(self._strategy_peak_config_path()),
        }

    def _strategy_build_nav_df(self, funds: List[str], history_days: int, session: Optional[requests.Session] = None) -> pd.DataFrame:
        series_map: Dict[str, pd.Series] = {}
        errors: List[str] = []
        active_session = session or self.session
        need_days = max(1, int(history_days))
        fetch_days = max(need_days + 120, need_days * 2)
        max_pages = max(5, min(120, int(math.ceil(fetch_days / 20.0))))

        # Short TTL cache to speed up repeated "执行决策".
        now = dt.datetime.now()
        cache_ttl = dt.timedelta(minutes=10)
        pending_codes: List[str] = []
        for code in funds:
            cached = self._strategy_nav_series_cache.get(code)
            cached_at = self._strategy_nav_series_cached_at.get(code)
            if cached is not None and cached_at is not None and (now - cached_at) <= cache_ttl and len(cached) >= need_days:
                series_map[code] = cached.copy()
            else:
                pending_codes.append(code)

        if pending_codes:
            if len(pending_codes) == 1:
                code = pending_codes[0]
                try:
                    s = fetch_nav_history_series(code, active_session, max_pages=max_pages)
                    series_map[code] = s
                    self._strategy_nav_series_cache[code] = s
                    self._strategy_nav_series_cached_at[code] = now
                except Exception as exc:
                    errors.append(f"{code}: {exc}")
            else:
                workers = max(2, min(6, len(pending_codes)))

                def _fetch_one(code: str) -> Tuple[str, pd.Series]:
                    with requests.Session() as sess:
                        s = fetch_nav_history_series(code, sess, max_pages=max_pages)
                    return code, s

                with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as ex:
                    future_map = {ex.submit(_fetch_one, code): code for code in pending_codes}
                    for fut in concurrent.futures.as_completed(future_map):
                        code = future_map[fut]
                        try:
                            c, s = fut.result()
                            series_map[c] = s
                            self._strategy_nav_series_cache[c] = s
                            self._strategy_nav_series_cached_at[c] = now
                        except Exception as exc:
                            errors.append(f"{code}: {exc}")

        if errors:
            raise ValueError("\u83b7\u53d6\u51c0\u503c\u5386\u53f2\u5931\u8d25:\n" + "\n".join(errors[:5]))

        nav_df = pd.concat({k: series_map[k] for k in funds if k in series_map}, axis=1).sort_index()
        nav_df = nav_df.ffill().dropna(how="any")
        if history_days > 0:
            nav_df = nav_df.tail(history_days)
        if nav_df.empty:
            raise ValueError("\u53ef\u7528\u5386\u53f2\u51c0\u503c\u6570\u636e\u4e3a\u7a7a")
        nav_df.columns = funds
        return nav_df

    def _strategy_parse_params(self, funds: List[str], data_override: Optional[Dict] = None) -> DualTriggerParams:
        if data_override is None:
            raw = self._strategy_text_get(self.strategy_params_text) or self.strategy_params_json_cache
            data = json.loads(raw) if raw else {}
        else:
            data = dict(data_override)
        if not isinstance(data, dict):
            raise ValueError("params JSON \u5fc5\u987b\u662f\u5bf9\u8c61")
        if "dd_stages" in data:
            data["dd_stages"] = tuple(float(x) for x in list(data["dd_stages"]))
        if "gap_fill" in data:
            data["gap_fill"] = tuple(float(x) for x in list(data["gap_fill"]))
        if "peak_rolling_days" in data and data["peak_rolling_days"] in ("", None):
            data["peak_rolling_days"] = None
        if "target_stock_fund_weights" in data and data["target_stock_fund_weights"] in ("", None):
            data["target_stock_fund_weights"] = None
        allowed_keys = set(getattr(DualTriggerParams, "__dataclass_fields__", {}).keys())
        if "eps" in allowed_keys and "eps" not in data:
            data["eps"] = STRATEGY_FIXED_NEW_HIGH_EPS
        if "new_high_epsilon" in allowed_keys and "new_high_epsilon" not in data:
            data["new_high_epsilon"] = STRATEGY_FIXED_NEW_HIGH_EPS
        if "new_high_epsilon" in data and "eps" in allowed_keys and "eps" not in data:
            data["eps"] = data.get("new_high_epsilon")
        if allowed_keys:
            data = {k: v for k, v in data.items() if k in allowed_keys}
        params = DualTriggerParams(**data)
        if params.target_stock_fund_weights is not None:
            mapped: Dict[str, float] = {}
            for k, v in params.target_stock_fund_weights.items():
                code = normalize_fund_code(k)
                num = to_float(v, default=None)
                if code in funds and num is not None and num > 0:
                    mapped[code] = float(num)
            params.target_stock_fund_weights = mapped if mapped else None
        return params

    def _strategy_parse_state(self) -> DualTriggerState:
        raw = self._strategy_text_get(self.strategy_state_text) or self.strategy_state_json_cache
        data = json.loads(raw) if raw else {}
        if not isinstance(data, dict):
            raise ValueError("state JSON \u5fc5\u987b\u662f\u5bf9\u8c61")
        if "dd_triggered" in data:
            data["dd_triggered"] = [float(x) for x in list(data["dd_triggered"])]
        return DualTriggerState(**data)

    def _strategy_build_params_cn(self, params: Dict) -> str:
        p = params or {}
        lines: List[str] = []
        lines.append("【参数解读】")
        lines.append(f"目标股票仓位：{self._strategy_fmt_pct(p.get('target_stock_w'))}")
        lines.append(f"仓位下限：{self._strategy_fmt_pct(p.get('force_low_w'))}")
        lines.append(f"仓位上限：{self._strategy_fmt_pct(p.get('force_high_w'))}")
        lines.append(f"高仓回落目标：{self._strategy_fmt_pct(p.get('force_high_target_w'))}")

        dd_stages = list(p.get("dd_stages") or [])
        gap_fill = list(p.get("gap_fill") or [])
        if dd_stages:
            lines.append("回撤触发档位：" + " / ".join(self._strategy_fmt_pct(x) for x in dd_stages))
        else:
            lines.append("回撤触发档位：-")
        if gap_fill:
            lines.append("每档补差额比例：" + " / ".join(self._strategy_fmt_pct(x) for x in gap_fill))
        else:
            lines.append("每档补差额比例：-")

        lines.append(f"回撤滚动窗口：{p.get('peak_rolling_days', '-')}")
        lines.append(f"低仓冷静期天数：{p.get('force_low_cooldown_days', '-')}")
        lines.append(f"仓位转化日期：{self._strategy_iso_date_to_zh_text(p.get('last_stage_reset_date')) or '-'}")
        freq_raw = str(p.get("decision_freq", "-"))
        lines.append(f"决策频率：{DECISION_FREQ_VALUE_TO_LABEL.get(freq_raw, freq_raw)}")
        lines.append(f"债基最低保留：{self._strategy_fmt_money(p.get('min_bond_value'))}")
        lines.append(f"股票最低保留：{self._strategy_fmt_money(p.get('min_stock_total_value'))}")
        lines.append(f"新高重置容差：{p.get('new_high_epsilon', p.get('eps', '-'))}")

        weights = p.get("target_stock_fund_weights") or {}
        if isinstance(weights, dict) and weights:
            lines.append("子基金权重：")
            for code, val in sorted(weights.items()):
                lines.append(f"- {code}: {self._strategy_fmt_pct(val)}")
        else:
            lines.append("子基金权重：未设置（默认等权）")
        return "\n".join(lines)

    def _strategy_build_state_cn(self, state: Dict) -> str:
        s = state or {}
        lines: List[str] = []
        lines.append("【状态解读】")
        dd_triggered = list(s.get("dd_triggered") or [])
        if dd_triggered:
            lines.append("已触发回撤档位：" + " / ".join(self._strategy_fmt_pct(x) for x in dd_triggered))
        else:
            lines.append("已触发回撤档位：无")
        lines.append(f"冷静期截止日：{s.get('cooldown_until_date') or '-'}")
        lines.append(f"上次决策日：{s.get('last_decision_date') or '-'}")
        lines.append(f"待执行仓位转化日期：{self._strategy_iso_date_to_zh_text(s.get('pending_reset_date')) or '-'}")
        lines.append(f"是否等待仓位转化：{'是' if bool(s.get('last_stage_wait_conversion')) else '否'}")
        lines.append(f"回撤模式：{'trade_peak' if bool(s.get('use_trade_peak_for_dd')) else 'rolling_peak'}")
        return "\n".join(lines)

    def _show_strategy_result_dialog(self, params_cn: str, state_cn: str, output_cn: str) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("策略执行结果")
        dialog.geometry("960x576")
        dialog.minsize(720, 400)
        dialog.resizable(True, True)

        root = ttk.Frame(dialog, style="App.TFrame", padding=8)
        root.pack(fill=tk.BOTH, expand=True)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(1, weight=1)
        root.rowconfigure(2, weight=0)

        ttk.Label(root, text="本次执行完成：参数、状态、结果（中文解读）", style="Body.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 6))

        tabs = ttk.Notebook(root)
        tabs.grid(row=1, column=0, sticky="nsew")

        def add_text_tab(title: str, content: str) -> None:
            tab = ttk.Frame(tabs, style="App.TFrame")
            tab.columnconfigure(0, weight=1)
            tab.rowconfigure(0, weight=1)
            text = tk.Text(tab, wrap="word", font=self.fonts["small"])
            text.grid(row=0, column=0, sticky="nsew")
            y_scroll = ttk.Scrollbar(tab, orient=tk.VERTICAL, command=text.yview)
            y_scroll.grid(row=0, column=1, sticky="ns")
            text.configure(yscrollcommand=y_scroll.set)
            text.insert("1.0", content)
            text.configure(state="disabled")
            tabs.add(tab, text=title)

        add_text_tab("执行结果", output_cn)
        add_text_tab("参数解读", params_cn)
        add_text_tab("状态解读", state_cn)
        tabs.select(0)

        ttk.Button(root, text="\u5173\u95ed", command=dialog.destroy, style="ToolbarPrimary.TButton").grid(row=2, column=0, sticky="e", pady=(8, 0))
        self._center_dialog(dialog)

    @staticmethod
    def _strategy_reason_zh(reason: str) -> str:
        text = str(reason or "").strip()
        if not text:
            return "-"
        m = re.match(r"freq_gate\(([^)]+)\)", text)
        if m:
            freq = DECISION_FREQ_VALUE_TO_LABEL.get(m.group(1), m.group(1))
            return f"触发决策频率门控（{freq}）"
        m = re.match(r"FORCE_LOW:\s*W=([0-9.]+)\s*<\s*([0-9.]+)", text)
        if m:
            w_now = float(m.group(1))
            low = float(m.group(2))
            return f"仓位低于下限：当前{w_now * 100:.2f}% < 下限{low * 100:.2f}%"
        m = re.match(r"FORCE_HIGH:\s*W=([0-9.]+)\s*>\s*([0-9.]+);\s*sell to W=([0-9.]+)", text)
        if m:
            w_now = float(m.group(1))
            high = float(m.group(2))
            tgt = float(m.group(3))
            return f"仓位高于上限：当前{w_now * 100:.2f}% > 上限{high * 100:.2f}%，目标降至{tgt * 100:.2f}%"
        m = re.match(r"DD_ADD:\s*dd=([0-9.]+%)\s*crossed\s*([0-9.]+%),\s*fill_gap=([0-9.]+%)", text)
        if m:
            return f"触发回撤加仓：回撤{m.group(1)}，跨过档位{m.group(2)}，按{m.group(3)}补齐仓位缺口"
        if text == "no trigger":
            return "未触发任何调仓条件"
        if text == "no trigger (cooldown)":
            return "冷静期内，未触发调仓"
        return text

    def _strategy_error_zh(self, err: str) -> str:
        text = str(err or "").strip()
        if not text:
            return "未知错误"
        replacements = {
            "today not found in basket index series": "决策日不在组合指数序列中",
            "S + B must be > 0 (cash excluded from decision base).": "股票市值与债基市值之和必须大于0（现金不计入仓位基数）",
            "decision_freq must be one of: daily/weekly/monthly/quarterly": "决策频率必须是：每日/每周/每月/每季度",
            "force_high_target_w must be >0 and < force_high_w.": "高仓回落目标必须大于0且小于仓位上限",
            "invalid fund code": "基金代码无效",
            "cannot fetch nav history": "无法获取净值历史",
        }
        for src, dst in replacements.items():
            text = text.replace(src, dst)
        reason_zh = self._strategy_reason_zh(text)
        if reason_zh == "未命中已知规则，请检查参数和数据":
            return f"系统异常（原始信息：{text}）"
        return reason_zh

    @staticmethod
    def _strategy_action_zh(action: str) -> str:
        mapping = {
            "FORCE_REBALANCE": "\u5f3a\u5236\u518d\u5e73\u8861",
            "DD_ADD": "\u56de\u64a4\u52a0\u4ed3",
            "HOLD": "\u4fdd\u6301\u4e0d\u52a8",
        }
        return mapping.get(str(action or "").strip(), str(action or "-"))

    @staticmethod
    def _strategy_constraint_zh(code: str) -> str:
        mapping = {
            "BOND_FLOOR_BLOCKED": "\u53d7\u201c\u503a\u57fa\u6700\u4f4e\u4fdd\u7559\u91d1\u989d\u201d\u9650\u5236\uff0c\u65e0\u6cd5\u7ee7\u7eed\u5356\u503a\u4e70\u80a1",
            "STOCK_FLOOR_BLOCKED": "\u53d7\u201c\u80a1\u7968\u6700\u4f4e\u4fdd\u7559\u91d1\u989d\u201d\u9650\u5236\uff0c\u65e0\u6cd5\u7ee7\u7eed\u5356\u80a1\u4e70\u503a",
        }
        return mapping.get(str(code or "").strip(), str(code or "-"))

    @staticmethod
    def _strategy_fmt_money(value: object) -> str:
        num = to_float(value, default=0.0) or 0.0
        return f"{num:,.2f}"

    @staticmethod
    def _strategy_fmt_pct(value: object) -> str:
        num = to_float(value, default=0.0) or 0.0
        return f"{num * 100:.2f}%"

    def _strategy_build_cn_output(self, display_payload: Dict, result: Dict) -> str:
        inputs = display_payload.get("inputs") or {}
        output = result or {}
        metrics = output.get("metrics") or {}
        orders = output.get("orders") or {}
        summary = orders.get("summary") or {}
        per_fund = orders.get("per_fund") or {}

        action = str(output.get("action", "-"))
        action_zh = self._strategy_action_zh(action)
        reason_raw = str(output.get("reason", "-"))
        reason_zh = self._strategy_reason_zh(reason_raw)

        lines: List[str] = []
        lines.append("\u3010\u51b3\u7b56\u7ed3\u8bba\u3011")
        lines.append(f"\u52a8\u4f5c\uff1a{action_zh}")
        lines.append(f"\u539f\u56e0\uff1a{reason_zh}")
        lines.append(f"\u51b3\u7b56\u65e5\uff1a{inputs.get('today', '-')}")
        if str(reason_raw).startswith("freq_gate("):
            lines.append("\u8bf4\u660e\uff1a\u5f53\u5929\u89e6\u53d1\u4e86\u201c\u51b3\u7b56\u9891\u7387\u95e8\u63a7\u201d\uff0c\u672c\u6b21\u672a\u6267\u884c\u8c03\u4ed3\uff0c\u4ec5\u8f93\u51fa\u4ed3\u4f4d\u5feb\u7167\u3002")
        lines.append("")

        snap_s = to_float(metrics.get("S"), default=None)
        snap_b = to_float(metrics.get("B"), default=None)
        snap_c = to_float(metrics.get("C"), default=None)
        if snap_s is None:
            stock_values = inputs.get("stock_values") or {}
            snap_s = float(sum(to_float(v, default=0.0) or 0.0 for v in stock_values.values()))
        if snap_b is None:
            snap_b = to_float(inputs.get("bond_value"), default=0.0) or 0.0
        if snap_c is None:
            snap_c = to_float(inputs.get("cash_value"), default=0.0) or 0.0
        snap_w = to_float(metrics.get("stock_weight_W"), default=None)
        if snap_w is None:
            base = (snap_s or 0.0) + (snap_b or 0.0)
            snap_w = ((snap_s or 0.0) / base) if base > 0 else 0.0

        lines.append("\u3010\u5f53\u524d\u4ed3\u4f4d\u72b6\u6001\u3011")
        lines.append(f"\u80a1\u7968\u603b\u5e02\u503c(S)\uff1a{self._strategy_fmt_money(snap_s)}")
        lines.append(f"\u503a\u57fa\u603b\u5e02\u503c(B)\uff1a{self._strategy_fmt_money(snap_b)}")
        lines.append(f"\u73b0\u91d1(C)\uff1a{self._strategy_fmt_money(snap_c)}")
        lines.append(f"\u80a1\u7968\u4ed3\u4f4dW=S/(S+B)\uff1a{self._strategy_fmt_pct(snap_w)}")
        if "dd" in metrics:
            dd_val = to_float(metrics.get("dd"), default=0.0) or 0.0
            dd_display = -abs(dd_val) if abs(dd_val) > 1e-12 else 0.0
            lines.append(f"\u7ec4\u5408\u56de\u64a4DD\uff1a{self._strategy_fmt_pct(dd_display)}")
            lines.append(
                "\u8bf4\u660e\uff1aDD=(\u56de\u770b\u7a97\u53e3\u5185\u7ec4\u5408\u6307\u6570\u5cf0\u503c-\u5f53\u524d\u7ec4\u5408\u6307\u6570)/\u7ec4\u5408\u6307\u6570\u5cf0\u503c\uff0c"
                "\u8868\u793a\u8ddd\u79bb\u8fd1\u671f\u9ad8\u70b9\u7684\u56de\u64a4\u6bd4\u4f8b\u3002"
            )
            if "stock_basket_index" in metrics and "peak" in metrics:
                idx_today = to_float(metrics.get("stock_basket_index"), default=0.0) or 0.0
                peak = to_float(metrics.get("peak"), default=0.0) or 0.0
                lines.append(f"\u5f53\u524d\u7ec4\u5408\u6307\u6570\uff1a{idx_today:.6f} | \u7a97\u53e3\u5cf0\u503c\uff1a{peak:.6f}")
        peak_meta = inputs.get("peak_meta") or {}
        if peak_meta:
            cfg_peak_val = to_float(peak_meta.get("config_peak_value"), default=None)
            cfg_peak_date = str(peak_meta.get("config_peak_date", "") or "-")
            trade_peak_val = to_float(peak_meta.get("trade_peak_value"), default=None)
            trade_peak_date = str(peak_meta.get("trade_peak_update_date", "") or "-")
            decision_peak_val = to_float(peak_meta.get("decision_peak_value"), default=None)
            decision_peak_date = str(peak_meta.get("decision_peak_date", "") or "-")
            source = str(peak_meta.get("source", "") or "")
            source_zh = {
                "rolling_peak": "按滚动窗口峰值计算",
                "peak_config": "使用峰值配置作为起点",
                "peak_config_ignored": "峰值配置日期超出可用范围，已忽略",
            }.get(source, source or "-")
            lines.append("\u3010峰值基准信息\u3011")
            if trade_peak_val is not None:
                lines.append(f"交易峰值：{trade_peak_val:.6f}（更新日期：{trade_peak_date}）")
            else:
                lines.append("交易峰值：-")
            if cfg_peak_val is not None:
                lines.append(f"配置峰值：{cfg_peak_val:.6f}（根据 {cfg_peak_date} 净值更新）")
            else:
                lines.append("配置峰值：尚未写入")
            if decision_peak_val is not None:
                lines.append(f"本次决策峰值：{decision_peak_val:.6f}（根据 {decision_peak_date} 净值更新）")
            lines.append(f"峰值来源：{source_zh}")
            if peak_meta.get("config_updated"):
                lines.append(f"峰值配置已更新：{peak_meta.get('update_reason', '-')}")
            pending_reset_date = str(peak_meta.get("pending_reset_date", "") or "").strip()
            if pending_reset_date:
                lines.append(
                    "待执行仓位转化日期："
                    f"{self._strategy_iso_date_to_zh_text(pending_reset_date) or pending_reset_date}"
                )
            applied_reset_date = str(peak_meta.get("manual_reset_applied_date", "") or "").strip()
            if applied_reset_date:
                lines.append(
                    "本次已按仓位转化日期执行峰值重置："
                    f"{self._strategy_iso_date_to_zh_text(applied_reset_date) or applied_reset_date}"
                )
        lines.append("")

        lookback_rows = list(inputs.get("fund_lookback_rows") or [])
        if lookback_rows:
            lookback_days = inputs.get("lookback_days_used", "-")
            lookback_start = inputs.get("lookback_start", "-")
            latest_nav_date = inputs.get("latest_nav_date", "-")
            lines.append("\u3010\u5404\u80a1\u7968\u57fa\u91d1\u56de\u770b\u7ed3\u679c\u3011")
            lines.append(f"\u56de\u770b\u5929\u6570\uff1a{lookback_days} | \u56de\u770b\u8d77\u70b9\u65e5\uff1a{lookback_start} | \u6700\u65b0\u51c0\u503c\u65e5\uff1a{latest_nav_date}")
            lines.append("\u8bf4\u660e\uff1a\u201c\u56de\u770b\u7ed3\u679c\u201d=\u51b3\u7b56\u65e5\u76f8\u5bf9\u56de\u770b\u8d77\u70b9\u7684\u6536\u76ca\u7387\u3002")
            for row in lookback_rows:
                code = row.get("code", "-")
                name = str(row.get("name", "") or "").strip()
                start_nav = to_float(row.get("start_nav"), default=0.0) or 0.0
                decision_nav = to_float(row.get("decision_nav"), default=0.0) or 0.0
                latest_nav = to_float(row.get("latest_nav"), default=0.0) or 0.0
                lookback_return = to_float(row.get("lookback_return"), default=0.0) or 0.0
                latest_vs_decision = to_float(row.get("latest_vs_decision"), default=0.0) or 0.0
                latest_vs_start = to_float(row.get("latest_vs_start"), default=0.0) or 0.0
                fund_title = f"{name}({code})" if name else str(code)
                lines.append(
                    f"- {fund_title} | \u8d77\u70b9\u51c0\u503c:{start_nav:.4f} | \u51b3\u7b56\u65e5\u51c0\u503c:{decision_nav:.4f} | "
                    f"\u56de\u770b\u6536\u76ca:{self._strategy_fmt_pct(lookback_return)}"
                )
                lines.append(
                    f"  \u6700\u65b0\u51c0\u503c:{latest_nav:.4f} | \u6700\u65b0\u51c0\u503c\u4e0e\u56de\u770b\u7ed3\u679c\u5dee\u503c:{self._strategy_fmt_pct(latest_vs_decision)} | "
                    f"\u6700\u65b0\u76f8\u5bf9\u8d77\u70b9\u7d2f\u8ba1:{self._strategy_fmt_pct(latest_vs_start)}"
                )
            lines.append("")

        lines.append("\u3010\u4ea4\u6613\u6307\u4ee4\u6c47\u603b\u3011")
        lines.append(
            f"\u73b0\u91d1\u4e70\u80a1\uff1a{self._strategy_fmt_money(summary.get('cash_buy_stock_total'))} | "
            f"\u5356\u503a\u4e70\u80a1\uff1a{self._strategy_fmt_money(summary.get('sell_bond_buy_stock_total'))} | "
            f"\u5356\u80a1\u4e70\u503a\uff1a{self._strategy_fmt_money(summary.get('sell_stock_buy_bond_total'))}"
        )
        buy_calc = metrics.get("buy_calc") or {}
        if isinstance(buy_calc, dict):
            buy_total = to_float(buy_calc.get("buy_total"), default=0.0) or 0.0
            if buy_total > 1e-8:
                s_before = to_float(buy_calc.get("stock_total_before"), default=0.0) or 0.0
                s_after = to_float(buy_calc.get("stock_total_after"), default=s_before + buy_total) or (s_before + buy_total)
                lines.append("【买入分配计算过程】")
                lines.append("规则：按“买入后股票总市值(S+A)”计算目标金额，再按差额分配到各基金。")
                lines.append(
                    f"S(买入前股票总市值)：{self._strategy_fmt_money(s_before)} | "
                    f"A(本次买入总额)：{self._strategy_fmt_money(buy_total)} | "
                    f"S+A：{self._strategy_fmt_money(s_after)}"
                )
                for row in list(buy_calc.get("funds") or []):
                    code = str(row.get("code", "") or "").strip()
                    if not code:
                        continue
                    name = str(self.fund_name_cache.get(code, "") or "").strip()
                    fund_title = f"{name}({code})" if name else code
                    cur_val = to_float(row.get("current_value"), default=0.0) or 0.0
                    tgt_val = to_float(row.get("target_after_value"), default=0.0) or 0.0
                    gap_val = to_float(row.get("gap_value"), default=0.0) or 0.0
                    alloc_val = to_float(row.get("alloc_value"), default=0.0) or 0.0
                    lines.append(
                        f"- {fund_title} | 当前:{self._strategy_fmt_money(cur_val)} | 目标:{self._strategy_fmt_money(tgt_val)} | "
                        f"差额:{self._strategy_fmt_money(gap_val)} | 分配:{self._strategy_fmt_money(alloc_val)}"
                    )
                lines.append("")

        def append_nonzero(title: str, data: Dict, note: Optional[str] = None) -> None:
            nonzero = [(k, to_float(v, default=0.0) or 0.0) for k, v in (data or {}).items()]
            nonzero = [(k, v) for k, v in nonzero if abs(v) > 1e-8]
            if not nonzero:
                return
            lines.append(f"{title}\uff1a")
            if note:
                lines.append(f"  说明：{note}")
            for code, val in nonzero:
                name = str(self.fund_name_cache.get(str(code), "") or "").strip()
                fund_title = f"{name}({code})" if name else str(code)
                lines.append(f"  - {fund_title}: {self._strategy_fmt_money(val)}")

        append_nonzero("\u73b0\u91d1\u4e70\u5165\u80a1\u7968(\u5206\u57fa\u91d1)", per_fund.get("buy_stock_by_fund_cash") or {})
        append_nonzero("\u5356\u503a\u8f6c\u4e70\u80a1(\u5206\u57fa\u91d1)", per_fund.get("buy_stock_by_fund_bond") or {})
        append_nonzero(
            "\u5356\u80a1\u8f6c\u4e70\u503a(\u5206\u57fa\u91d1)",
            per_fund.get("sell_stock_by_fund_to_bond") or {},
            note="\u6309\u5404\u57fa\u91d1\u76f8\u5bf9\u76ee\u6807\u6743\u91cd\u7684\u8d85\u914d\u7a0b\u5ea6\u5206\u644a\u5356\u51fa\uff0c\u8d85\u914d\u8d8a\u9ad8\u5356\u51fa\u8d8a\u591a\uff1b\u5982\u679c\u6574\u4f53\u65e0\u8d85\u914d\uff0c\u5219\u6309\u5f53\u524d\u6301\u4ed3\u5360\u6bd4\u5206\u644a\u3002",
        )
        lines.append("")

        constraints = list(output.get("constraints_hit") or [])
        if constraints:
            lines.append("\u3010\u7ea6\u675f/\u963b\u585e\u8bf4\u660e\u3011")
            for item in constraints:
                lines.append(f"- {self._strategy_constraint_zh(item)}")
            lines.append(
                f"\u672a\u5b8c\u6210\u4e70\u5165\u91d1\u989d\uff1a{self._strategy_fmt_money(output.get('unfilled_buy_amount'))} | "
                f"\u672a\u5b8c\u6210\u5356\u51fa\u91d1\u989d\uff1a{self._strategy_fmt_money(output.get('unfilled_sell_amount'))}"
            )
            lines.append("")

        return "\n".join(lines)

    def _set_strategy_running(self, running: bool) -> None:
        self._strategy_running = running
        if self.strategy_run_btn is not None:
            self.strategy_run_btn.configure(state="disabled" if running else "normal")

    def _run_strategy_worker(
        self,
        funds: List[str],
        history_days: int,
        bond_value: float,
        cash_value: float,
        desired_today: Optional[pd.Timestamp],
        params: DualTriggerParams,
        state: DualTriggerState,
        shares_map: Dict[str, float],
        params_input: Dict,
        state_input: Dict,
    ) -> None:
        try:
            with requests.Session() as worker_session:
                nav_df = self._strategy_build_nav_df(funds, history_days=history_days, session=worker_session)

            if desired_today is not None:
                eligible = nav_df.index[nav_df.index <= desired_today]
                if len(eligible) <= 0:
                    raise ValueError("\u51b3\u7b56\u65e5\u65e9\u4e8e\u5386\u53f2\u6570\u636e\u8303\u56f4")
                today = eligible[-1]
            else:
                today = nav_df.index[-1]

            nav_today = nav_df.loc[today]
            stock_values = {code: float(shares_map.get(code, 0.0)) * float(nav_today[code]) for code in funds}
            nav_hist = nav_df.loc[:today].tail(history_days)
            if nav_hist.empty:
                nav_hist = nav_df.loc[:today]
            lookback_start_idx = nav_hist.index[0]
            latest_idx = nav_df.index[-1]
            lookback_rows: List[Dict] = []
            for code in funds:
                start_nav = float(nav_hist.loc[lookback_start_idx, code])
                decision_nav = float(nav_df.loc[today, code])
                latest_nav = float(nav_df.loc[latest_idx, code])
                fund_name = str(self.fund_name_cache.get(code, "") or "").strip()
                lookback_return = (decision_nav / start_nav - 1.0) if start_nav > 0 else 0.0
                latest_vs_decision = (latest_nav / decision_nav - 1.0) if decision_nav > 0 else 0.0
                latest_vs_start = (latest_nav / start_nav - 1.0) if start_nav > 0 else 0.0
                lookback_rows.append(
                    {
                        "code": code,
                        "name": fund_name,
                        "start_nav": start_nav,
                        "decision_nav": decision_nav,
                        "latest_nav": latest_nav,
                        "lookback_return": lookback_return,
                        "latest_vs_decision": latest_vs_decision,
                        "latest_vs_start": latest_vs_start,
                    }
                )

            basket = stock_basket_index_equal_weights(nav_df, funds)
            basket_for_decision, peak_cfg_before, peak_source = self._strategy_prepare_basket_with_peak_config(basket, today)
            result = decide(
                today=today,
                basket_idx=basket_for_decision,
                bond_value=float(bond_value),
                cash_value=float(cash_value),
                stock_values=stock_values,
                funds=funds,
                params=params,
                state=state,
            )
            peak_meta = self._strategy_update_peak_config(
                result=result,
                basket_used=basket_for_decision,
                today=today,
                params=params,
                cfg_before=peak_cfg_before,
                source=peak_source,
            )

            display_payload = {
                "inputs": {
                    "today": today.date().isoformat(),
                    "funds": funds,
                    "bond_value": float(bond_value),
                    "cash_value": float(cash_value),
                    "stock_values": stock_values,
                    "params": params_input,
                    "state": state_input,
                    "nav_df_rows": int(len(nav_df)),
                    "nav_df_start": nav_df.index[0].date().isoformat(),
                    "nav_df_end": nav_df.index[-1].date().isoformat(),
                    "lookback_days_used": int(len(nav_hist)),
                    "lookback_start": lookback_start_idx.date().isoformat(),
                    "lookback_end": today.date().isoformat(),
                    "latest_nav_date": latest_idx.date().isoformat(),
                    "fund_lookback_rows": lookback_rows,
                    "peak_meta": peak_meta,
                },
                "output": result,
            }
            self.after(0, lambda: self._run_strategy_done(today, funds, stock_values, result, display_payload))
        except Exception as exc:
            msg = str(exc)
            self.after(0, lambda: self._run_strategy_failed(msg))

    def _run_strategy_done(
        self,
        today: pd.Timestamp,
        funds: List[str],
        stock_values: Dict[str, float],
        result: Dict,
        display_payload: Dict,
    ) -> None:
        self.strategy_today_var.set(today.date().isoformat())
        self.strategy_stock_values_var.set(json.dumps(stock_values, ensure_ascii=False))
        output_cn = self._strategy_build_cn_output(display_payload, result)
        params_dict = display_payload.get("inputs", {}).get("params", {}) or {}
        state_dict = result.get("new_state", {}) or {}
        output_json = json.dumps(display_payload, ensure_ascii=False, indent=2, default=str)
        params_json = json.dumps(params_dict, ensure_ascii=False, indent=2, default=str)
        state_json = json.dumps(state_dict, ensure_ascii=False, indent=2, default=str)
        params_cn = self._strategy_build_params_cn(params_dict)
        state_cn = self._strategy_build_state_cn(state_dict)
        self.strategy_output_json_cache = output_json
        self.strategy_params_json_cache = params_json
        self.strategy_state_json_cache = state_json
        self._strategy_text_set(self.strategy_output_cn_text, output_cn)
        self._strategy_text_set(self.strategy_output_text, output_json)
        self._strategy_text_set(self.strategy_state_text, state_json)
        peak_meta = display_payload.get("inputs", {}).get("peak_meta", {}) or {}
        self._strategy_set_trade_peak_display(
            to_float(peak_meta.get("trade_peak_value"), default=None),
            str(peak_meta.get("trade_peak_update_date", "") or ""),
        )
        self._strategy_save_ui_state()
        action_zh = self._strategy_action_zh(str(result.get("action", "-")))
        self.strategy_status_var.set(
            f"\u6267\u884c\u6210\u529f\uff1a{action_zh}\uff0c\u51b3\u7b56\u65e5={today.date().isoformat()}\uff0c\u6807\u7684\u6570={len(funds)}"
        )
        self._show_strategy_result_dialog(params_cn, state_cn, output_cn)
        self._set_strategy_running(False)

    def _run_strategy_failed(self, err: str) -> None:
        err_zh = self._strategy_error_zh(err)
        self.strategy_status_var.set(f"\u6267\u884c\u5931\u8d25\uff1a{err_zh}")
        self._strategy_text_set(self.strategy_output_cn_text, f"\u6267\u884c\u5931\u8d25\uff1a{err_zh}")
        self._strategy_text_set(self.strategy_output_text, f"\u6267\u884c\u5931\u8d25\uff1a{err_zh}")
        self._strategy_save_ui_state()
        self._set_strategy_running(False)

    def _run_strategy_decision(self) -> None:
        if self._strategy_running:
            self.strategy_status_var.set("\u7b56\u7565\u6b63\u5728\u6267\u884c\uff0c\u8bf7\u7a0d\u5019\u2026")
            return
        try:
            funds = self.get_all_fund_codes()
            if not funds:
                raise ValueError("\u5f53\u524d\u201c\u80a1\u7968\u57fa\u91d1\u201d\u65e0\u53ef\u7528\u6301\u4ed3\uff0c\u65e0\u6cd5\u6267\u884c\u7b56\u7565")

            history_days = int(to_float(self.strategy_history_days_var.get(), default=260) or 260)
            history_days = max(1, history_days)
            bond_value = float(self._compute_bond_portfolio_summary().get("market", 0.0))
            self.strategy_bond_value_var.set(f"{bond_value:.2f}")
            cash_value = to_float(self.strategy_cash_value_var.get(), default=None)
            if cash_value is None:
                raise ValueError("\u73b0\u91d1\u91d1\u989d\u5fc5\u987b\u4e3a\u6570\u5b57")

            today_raw = parse_date(self.strategy_today_var.get().strip())
            desired_today = pd.Timestamp(today_raw) if today_raw else None

            params_input = self._strategy_collect_params_from_form()
            self.strategy_params_json_cache = json.dumps(params_input, ensure_ascii=False, indent=2)
            self._strategy_text_set(self.strategy_params_text, self.strategy_params_json_cache)
            self._strategy_save_ui_state()
            params = self._strategy_parse_params(funds, data_override=params_input)
            state = self._strategy_parse_state()
            state_input = json.loads((self._strategy_text_get(self.strategy_state_text) or self.strategy_state_json_cache or "{}"))
            if not isinstance(state_input, dict):
                raise ValueError("state JSON \u5fc5\u987b\u662f\u5bf9\u8c61")

            shares_map = self._strategy_current_stock_shares(funds)
            self._set_strategy_running(True)
            self.strategy_status_var.set(
                f"\u6267\u884c\u4e2d\uff1a\u6b63\u5728\u83b7\u53d6{len(funds)}\u53ea\u57fa\u91d1\u7684\u5386\u53f2\u51c0\u503c\u2026"
            )
            t = threading.Thread(
                target=self._run_strategy_worker,
                args=(
                    funds,
                    history_days,
                    float(bond_value),
                    float(cash_value),
                    desired_today,
                    params,
                    state,
                    shares_map,
                    params_input,
                    state_input,
                ),
                daemon=True,
            )
            t.start()
        except Exception as exc:
            err_zh = self._strategy_error_zh(str(exc))
            self._set_strategy_running(False)
            self.strategy_status_var.set(f"\u6267\u884c\u5931\u8d25\uff1a{err_zh}")
            self._strategy_text_set(self.strategy_output_cn_text, f"\u6267\u884c\u5931\u8d25\uff1a{err_zh}")
            self._strategy_text_set(self.strategy_output_text, f"\u6267\u884c\u5931\u8d25\uff1a{err_zh}")

    def _build_general_tab(self, frame: ttk.Frame) -> None:
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(3, weight=0)

        ttk.Label(frame, text="\u4ea4\u6613\u6587\u6863", style="Body.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame, textvariable=self.vars["app.xlsx_path"]).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(frame, text="\u9009\u62e9", command=self.choose_xlsx_path).grid(row=0, column=2, padx=2)
        ttk.Button(frame, text="\u52a0\u8f7d\u6587\u6863", command=self.load_transactions_from_doc, style="Primary.TButton").grid(row=0, column=3, padx=2)

        ttk.Label(frame, text="\u6536\u76ca\u9608\u503c", style="Body.TLabel").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frame, textvariable=self.vars["app.threshold"], width=20).grid(row=1, column=1, sticky="w", padx=6, pady=(8, 0))

        ttk.Label(frame, text="ntfy_url", style="Body.TLabel").grid(row=2, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frame, textvariable=self.vars["app.ntfy_url"]).grid(row=2, column=1, columnspan=3, sticky="ew", padx=6, pady=(8, 0))

        ttk.Button(frame, text="\u4fdd\u5b58\u901a\u7528\u8bbe\u7f6e", command=self.save_settings_only, style="Primary.TButton").grid(
            row=3, column=0, columnspan=4, pady=(12, 0), sticky="e"
        )

    def apply_config_to_form(self) -> None:
        app = self.config.get("app", {})
        self.vars["app.xlsx_path"].set(str(app.get("xlsx_path", "data/jijin.xlsx")))
        self.vars["app.threshold"].set(str(app.get("threshold", 0.40)))
        self.vars["app.ntfy_url"].set(str(app.get("ntfy_url", "")))

    def collect_form_to_config(self) -> None:
        threshold_text = self.vars["app.threshold"].get().strip() or "0.40"
        try:
            threshold = float(threshold_text)
        except ValueError as exc:
            raise ValueError("threshold 必须是数字") from exc

        self.config.setdefault("app", {})
        self.config["app"]["xlsx_path"] = self.vars["app.xlsx_path"].get().strip() or "data/jijin.xlsx"
        self.config["app"]["threshold"] = threshold
        self.config["app"]["ntfy_url"] = self.vars["app.ntfy_url"].get().strip()
        self.config["app"].pop("notify_all", None)

        funds = sorted(set(self.get_all_fund_codes()))
        self.config["funds"] = funds

    def choose_xlsx_path(self) -> None:
        path = filedialog.asksaveasfilename(
            title="\u9009\u62e9\u4ea4\u6613\u8bb0\u5f55\u6587\u4ef6",
            defaultextension=".xlsx",
            initialfile=Path(self.vars["app.xlsx_path"].get() or "jijin.xlsx").name,
            filetypes=[("Excel \u6587\u4ef6", "*.xlsx"), ("\u6240\u6709\u6587\u4ef6", "*.*")],
        )
        if path:
            self.vars["app.xlsx_path"].set(path)

    def load_transactions_from_doc(self, silent: bool = False, fetch_latest_bond: bool = True) -> None:
        try:
            self.collect_form_to_config()
            xlsx_path = Path(self.config["app"]["xlsx_path"])
            self.transactions = load_transactions(xlsx_path)
            self.bond_transactions = load_bond_transactions(xlsx_path)
            self.next_uid = max([x["uid"] for x in self.transactions], default=0) + 1
            self.bond_next_uid = max([x["uid"] for x in self.bond_transactions], default=0) + 1
            stored_nav_map = load_latest_nav_map(xlsx_path)
            self.latest_nav_cache = {code: ("workbook", nav) for code, nav in stored_nav_map.items()}
            sheet_name_map = load_fund_name_map(xlsx_path)
            self.fund_name_cache.update(sheet_name_map)
            for txn in self.bond_transactions:
                code = normalize_fund_code(txn.get("fund_code"))
                name = str(txn.get("fund_name", "") or "").strip()
                if code and name:
                    self.fund_name_cache[code] = name
            self.sheet_fund_codes = sorted(set(sheet_name_map.keys()) | {txn["fund_code"] for txn in self.transactions})
            self.refresh_fund_list(keep_current=False)
            self.refresh_fund_view()
            self.refresh_bond_view(fetch_latest=fetch_latest_bond)
            self.refresh_analysis_view()
            self._strategy_fill_from_portfolio()
        except Exception as exc:
            messagebox.showerror("\u52a0\u8f7d\u5931\u8d25", str(exc))

    def save_settings_only(self) -> None:
        self._save_settings(show_success=True)

    def save_all(self) -> None:
        try:
            self.collect_form_to_config()
            xlsx_path = Path(self.config["app"]["xlsx_path"])
            latest_nav_map = {code: val[1] for code, val in self.latest_nav_cache.items() if val and len(val) >= 2}
            fund_name_map = {normalize_fund_code(code): name for code, name in self.fund_name_cache.items() if normalize_fund_code(code)}
            save_transactions(xlsx_path, self.transactions, latest_nav_map=latest_nav_map, fund_name_map=fund_name_map)
            save_bond_transactions(xlsx_path, self.bond_transactions)
            self.sheet_fund_codes = sorted({normalize_fund_code(txn["fund_code"]) for txn in self.transactions if normalize_fund_code(txn["fund_code"])})
            save_config(self.config_path, self.config)
            self.saved_app_state = copy.deepcopy(self.config.get("app", {}))
            messagebox.showinfo("\u63d0\u793a", f"\u5df2\u4fdd\u5b58\uff1a\n{self.config_path}\n{xlsx_path}")
        except Exception as exc:
            messagebox.showerror("\u4fdd\u5b58\u5931\u8d25", str(exc))

    def _save_settings(self, show_success: bool = False) -> bool:
        try:
            self.collect_form_to_config()
            save_config(self.config_path, self.config)
            self.saved_app_state = copy.deepcopy(self.config.get("app", {}))
            if show_success:
                messagebox.showinfo("\u63d0\u793a", f"\u8bbe\u7f6e\u5df2\u4fdd\u5b58\u5230\uff1a\n{self.config_path}")
            return True
        except Exception as exc:
            messagebox.showerror("\u4fdd\u5b58\u5931\u8d25", str(exc))
            return False

    def _reload_form_from_saved_state(self) -> None:
        app = self.saved_app_state or {}
        self.vars["app.xlsx_path"].set(str(app.get("xlsx_path", "data/jijin.xlsx")))
        self.vars["app.threshold"].set(str(app.get("threshold", 0.40)))
        self.vars["app.ntfy_url"].set(str(app.get("ntfy_url", "")))

    def _has_unsaved_settings(self) -> bool:
        app = self.saved_app_state or {}

        current_xlsx = self.vars["app.xlsx_path"].get().strip() or "data/jijin.xlsx"
        saved_xlsx = str(app.get("xlsx_path", "data/jijin.xlsx")).strip() or "data/jijin.xlsx"
        if current_xlsx != saved_xlsx:
            return True

        current_ntfy = self.vars["app.ntfy_url"].get().strip()
        saved_ntfy = str(app.get("ntfy_url", "")).strip()
        if current_ntfy != saved_ntfy:
            return True

        current_threshold_raw = self.vars["app.threshold"].get().strip() or "0.40"
        current_threshold = to_float(current_threshold_raw, default=None)
        saved_threshold = to_float(app.get("threshold", 0.40), default=0.40)
        if current_threshold is None:
            return True
        if saved_threshold is None:
            saved_threshold = 0.40
        return abs(current_threshold - saved_threshold) > 1e-12

    def _ask_save_settings_action(self, for_close: bool = False) -> str:
        if not self._has_unsaved_settings():
            return "none"

        prompt = "\u901a\u7528\u8bbe\u7f6e\u5df2\u4fee\u6539\uff0c\u662f\u5426\u4fdd\u5b58\u540e\u518d\u9000\u51fa\uff1f" if for_close else "\u901a\u7528\u8bbe\u7f6e\u5df2\u4fee\u6539\uff0c\u662f\u5426\u5148\u4fdd\u5b58\uff1f"
        answer = messagebox.askyesnocancel("\u672a\u4fdd\u5b58\u7684\u8bbe\u7f6e", prompt)
        if answer is None:
            return "cancel"
        if answer:
            return "save" if self._save_settings(show_success=False) else "cancel"
        return "discard"

    def on_tab_changed(self, _event=None) -> None:
        if self._is_switching_tab or self.tabs is None or self.general_tab is None:
            return

        current_tab = self.tabs.select()
        previous_tab = self._last_tab_id
        self._last_tab_id = current_tab

        if previous_tab:
            leaving_general = previous_tab == str(self.general_tab) and current_tab != previous_tab
            if leaving_general:
                action = self._ask_save_settings_action(for_close=False)
                if action == "discard":
                    self._reload_form_from_saved_state()
                elif action == "cancel":
                    self._is_switching_tab = True
                    try:
                        self.tabs.select(previous_tab)
                        self._last_tab_id = previous_tab
                    finally:
                        self._is_switching_tab = False
                    return
            leaving_strategy = self.strategy_tab is not None and previous_tab == str(self.strategy_tab) and current_tab != previous_tab
            if leaving_strategy:
                self._strategy_save_ui_state()

        if self.analysis_tab is not None and current_tab == str(self.analysis_tab):
            self.refresh_analysis_view(fetch_latest=True)
        if self.strategy_tab is not None and current_tab == str(self.strategy_tab):
            self._strategy_fill_from_portfolio()

    def on_window_close(self) -> None:
        action = self._ask_save_settings_action(for_close=True)
        if action == "cancel":
            return
        self._strategy_save_ui_state()
        if self._overlay_job:
            try:
                self.after_cancel(self._overlay_job)
            except Exception:
                pass
            self._overlay_job = None
        self.destroy()

    def get_all_fund_codes(self) -> List[str]:
        return list(self.sheet_fund_codes)

    def _format_fund_display(self, code: str) -> str:
        code = normalize_fund_code(code)
        if not code:
            return ""
        name = str(self.fund_name_cache.get(code, "") or "").strip()
        return f"{name}({code})" if name else code

    def _resolve_selected_fund_code(self) -> str:
        raw = self.selected_fund_var.get().strip()
        code = normalize_fund_code(raw)
        if code:
            return code
        mapped = self.fund_display_to_code.get(raw, "")
        if mapped:
            return mapped
        matched = re.search(r"(\d{6})\s*\)?\s*$", raw)
        if matched:
            return normalize_fund_code(matched.group(1))
        return ""

    def _set_selected_fund_code(self, code: str) -> None:
        code = normalize_fund_code(code)
        self.selected_fund_var.set(self._format_fund_display(code) if code else "")

    def _set_rate_bg_by_value(self, rate: Optional[float]) -> None:
        if not hasattr(self, "summary_rate_label"):
            return
        if rate is None:
            self.summary_rate_label.configure(style="SummaryRateNeutral.TLabel")
        elif rate > 0:
            self.summary_rate_label.configure(style="SummaryRatePos.TLabel")
        else:
            self.summary_rate_label.configure(style="SummaryRateNeg.TLabel")

    @staticmethod
    def _compact_name(name: str, max_len: int = 10) -> str:
        s = str(name or "").strip()
        if len(s) <= max_len:
            return s
        return f"{s[:max_len]}..."

    def refresh_fund_list(self, keep_current: bool = True) -> None:
        prev_code = self._resolve_selected_fund_code()
        funds = self.get_all_fund_codes()
        values = [self._format_fund_display(code) for code in funds]
        self.fund_display_to_code = {display: code for display, code in zip(values, funds)}
        self.fund_combo["values"] = values

        if keep_current and prev_code in funds:
            self._set_selected_fund_code(prev_code)
        elif funds:
            self._set_selected_fund_code(funds[0])
        else:
            self.selected_fund_var.set("")

    def fund_transactions(self, code: str) -> List[Dict]:
        return [x for x in self.transactions if x["fund_code"] == code]

    def get_latest_nav_cached(self, code: str) -> Optional[Tuple[str, float]]:
        return self.latest_nav_cache.get(code)

    def build_display_rows(self, code: str, latest_nav: Optional[float]) -> Tuple[List[Dict], Dict[str, float]]:
        txns = sorted(self.fund_transactions(code), key=lambda x: (x["date"], x["uid"]))
        lots, _, oversold = build_positions(txns)

        total_shares = sum(x["remaining_shares"] for x in lots)
        total_cost = sum(x["remaining_shares"] * x["buy_nav"] for x in lots)
        total_profit = 0.0
        total_rate = 0.0

        rows = []
        for txn in txns:
            rate_str = "-"
            profit_str = "-"
            amount_str = "-"
            rate_value = None
            profit_value = None
            amount_value = None

            shares = float(txn["shares"])
            if txn["action"] == "BUY":
                nav = float(txn["nav"] or 0.0)
                if latest_nav is not None and nav > 0:
                    rate = (latest_nav - nav) / nav
                    profit = shares * (latest_nav - nav)
                    amount = shares * latest_nav
                    rate_value = rate
                    profit_value = profit
                    rate_str = f"{rate * 100:.2f}%"
                    profit_str = f"{profit:.2f}"
                    amount_str = f"{amount:.2f}"
                    amount_value = amount
                elif nav > 0:
                    amount = shares * nav
                    amount_str = f"{amount:.2f}"
                    amount_value = amount
            elif latest_nav is not None:
                amount = shares * latest_nav
                amount_str = f"{amount:.2f}"
                amount_value = amount

            rows.append(
                {
                    "uid": str(txn["uid"]),
                    "date": txn["date"],
                    "shares": f"{shares:.2f}",
                    "nav": "" if txn["nav"] is None else f"{txn['nav']:.4f}",
                    "rate": rate_str,
                    "profit": profit_str,
                    "amount": amount_str,
                    "note": txn.get("note", ""),
                    "date_sort": txn["date"],
                    "rate_sort": rate_value,
                    "profit_sort": profit_value,
                    "amount_sort": amount_value,
                }
            )

        if latest_nav is not None:
            total_profit = total_shares * latest_nav - total_cost
            total_rate = (total_profit / total_cost) if total_cost > 0 else 0.0

        summary = {
            "shares": total_shares,
            "cost": total_cost,
            "profit": total_profit,
            "rate": total_rate,
            "oversold": oversold,
        }
        return rows, summary

    @staticmethod
    def _sort_optional_numeric_rows(rows: List[Dict], key_name: str, descending: bool) -> List[Dict]:
        valid_rows = [x for x in rows if x.get(key_name) is not None]
        empty_rows = [x for x in rows if x.get(key_name) is None]
        valid_rows.sort(key=lambda x: x[key_name], reverse=descending)
        return valid_rows + empty_rows

    def _header_text(self, column: str) -> str:
        base = self.header_labels.get(column, column)
        if column not in self.sortable_columns:
            return base
        if self.sort_column == column:
            icon = "\u25bc" if self.sort_descending else "\u25b2"
            return f"{base} {icon}"
        return f"{base} \u25b3\u25bd"

    def _refresh_tree_headings(self) -> None:
        for column in self.tx_tree["columns"]:
            if column in self.sortable_columns:
                self.tx_tree.heading(
                    column,
                    text=self._header_text(column),
                    command=lambda c=column: self.on_sort_header_click(c),
                )
            else:
                self.tx_tree.heading(column, text=self._header_text(column))

    def on_sort_header_click(self, column: str) -> None:
        if column not in self.sortable_columns:
            return
        if self.sort_column == column:
            self.sort_descending = not self.sort_descending
        else:
            self.sort_column = column
            self.sort_descending = True
        self.refresh_fund_view()

    def apply_row_sort(self, rows: List[Dict]) -> List[Dict]:
        result = list(rows)
        if self.sort_column == "date":
            result.sort(key=lambda x: (x["date_sort"], int(x["uid"])), reverse=self.sort_descending)
        elif self.sort_column == "rate":
            result = self._sort_optional_numeric_rows(result, "rate_sort", descending=self.sort_descending)
        elif self.sort_column == "profit":
            result = self._sort_optional_numeric_rows(result, "profit_sort", descending=self.sort_descending)
        elif self.sort_column == "amount":
            result = self._sort_optional_numeric_rows(result, "amount_sort", descending=self.sort_descending)
        return result

    def _bond_header_text(self, column: str) -> str:
        base = self.bond_header_labels.get(column, column)
        if column not in self.bond_sortable_columns:
            return base
        if self.bond_sort_column == column:
            icon = "\u25bc" if self.bond_sort_descending else "\u25b2"
            return f"{base} {icon}"
        return f"{base} \u25b3\u25bd"

    def _refresh_bond_tree_headings(self) -> None:
        if not hasattr(self, "bond_tree") or self.bond_tree is None:
            return
        for column in self.bond_tree["columns"]:
            if column in self.bond_sortable_columns:
                self.bond_tree.heading(
                    column,
                    text=self._bond_header_text(column),
                    command=lambda c=column: self.on_bond_sort_header_click(c),
                )
            else:
                self.bond_tree.heading(column, text=self._bond_header_text(column))

    def on_bond_sort_header_click(self, column: str) -> None:
        if column not in self.bond_sortable_columns:
            return
        if self.bond_sort_column == column:
            self.bond_sort_descending = not self.bond_sort_descending
        else:
            self.bond_sort_column = column
            self.bond_sort_descending = True
        self.refresh_bond_view(fetch_latest=False)

    def apply_bond_row_sort(self, rows: List[Dict]) -> List[Dict]:
        result = list(rows)
        if self.bond_sort_column == "rate":
            result = self._sort_optional_numeric_rows(result, "rate_sort", descending=self.bond_sort_descending)
        elif self.bond_sort_column == "profit":
            result = self._sort_optional_numeric_rows(result, "profit_sort", descending=self.bond_sort_descending)
        elif self.bond_sort_column == "amount":
            result = self._sort_optional_numeric_rows(result, "amount_sort", descending=self.bond_sort_descending)
        return result

    def _clear_rate_overlays(self) -> None:
        for label in self.rate_overlay_labels.values():
            try:
                label.destroy()
            except Exception:
                pass
        self.rate_overlay_labels = {}

    def _on_rate_overlay_click(self, iid: str) -> None:
        try:
            self.tx_tree.selection_set(iid)
            self.tx_tree.focus(iid)
        except Exception:
            pass

    def _refresh_rate_overlays(self) -> None:
        self._overlay_job = None
        if not hasattr(self, "tx_tree") or not self.tx_tree.winfo_exists():
            return
        columns = list(self.tx_tree["columns"])
        rate_col_token = f"#{columns.index('rate') + 1}" if "rate" in columns else "#4"
        tree_w = self.tx_tree.winfo_width()
        tree_h = self.tx_tree.winfo_height()
        visible_iids = set(str(iid) for iid in self.tx_tree.get_children())
        stale = [iid for iid in list(self.rate_overlay_labels.keys()) if iid not in visible_iids]
        for iid in stale:
            try:
                self.rate_overlay_labels[iid].destroy()
            except Exception:
                pass
            self.rate_overlay_labels.pop(iid, None)

        for iid in self.tx_tree.get_children():
            iid = str(iid)
            rate_value = self.rate_cell_meta.get(str(iid))
            if rate_value is None:
                if iid in self.rate_overlay_labels:
                    try:
                        self.rate_overlay_labels[iid].destroy()
                    except Exception:
                        pass
                    self.rate_overlay_labels.pop(iid, None)
                continue
            bbox = self.tx_tree.bbox(iid, "rate")
            if not bbox:
                if iid in self.rate_overlay_labels:
                    self.rate_overlay_labels[iid].place_forget()
                continue
            x, y, w, h = bbox
            if w <= 2 or h <= 2:
                if iid in self.rate_overlay_labels:
                    self.rate_overlay_labels[iid].place_forget()
                continue
            # Only overlay fully visible rate cells, avoiding stray labels in blank area.
            if x < 0 or y < 0 or x + w > tree_w or y + h > tree_h:
                if iid in self.rate_overlay_labels:
                    self.rate_overlay_labels[iid].place_forget()
                continue
            center_x = x + max(1, w // 2)
            center_y = y + max(1, h // 2)
            if self.tx_tree.identify_row(center_y) != iid or self.tx_tree.identify_column(center_x) != rate_col_token:
                if iid in self.rate_overlay_labels:
                    self.rate_overlay_labels[iid].place_forget()
                continue
            rate_text = self.tx_tree.set(iid, "rate")
            bg = "#ECA9A7" if rate_value > 0 else "#97CFAD"
            label = self.rate_overlay_labels.get(iid)
            if label is None:
                label = tk.Label(
                    self.tx_tree,
                    text=rate_text,
                    bg=bg,
                    fg="#FFFFFF",
                    bd=0,
                    padx=2,
                    pady=0,
                    anchor="center",
                    font=self.fonts["small"],
                )
                label.bind("<Button-1>", lambda _e, rid=iid: self._on_rate_overlay_click(rid))
                self.rate_overlay_labels[iid] = label
            else:
                label.configure(text=rate_text, bg=bg, fg="#FFFFFF")
            label.place(x=x + 1, y=y + 1, width=w - 2, height=h - 2)

    def _queue_rate_overlay_refresh(self, delay_ms: int = 16) -> None:
        if not self.winfo_exists():
            return
        if self._overlay_job:
            try:
                self.after_cancel(self._overlay_job)
            except Exception:
                pass
        self._overlay_job = self.after(delay_ms, self._refresh_rate_overlays)

    def refresh_fund_view(self) -> None:
        code = self._resolve_selected_fund_code()
        self._set_selected_fund_code(code)
        self._refresh_tree_headings()

        self.rate_cell_meta = {}
        self._clear_rate_overlays()
        for item in self.tx_tree.get_children():
            self.tx_tree.delete(item)

        if not code:
            self.summary_fund_var.set("-")
            self.summary_latest_var.set("-")
            self.summary_shares_var.set("-")
            self.summary_cost_var.set("-")
            self.summary_profit_var.set("-")
            self.summary_rate_var.set("-")
            self._set_rate_bg_by_value(None)
            self._queue_rate_overlay_refresh()
            self.refresh_analysis_view()
            return

        if code not in self.fund_name_cache:
            try:
                self.fund_name_cache[code] = fetch_fund_name(code, self.session)
            except Exception:
                self.fund_name_cache[code] = ""
        fund_name = self.fund_name_cache.get(code, "")
        self.summary_fund_var.set(self._compact_name(fund_name or code, max_len=7))

        latest = self.get_latest_nav_cached(code)
        latest_nav = latest[1] if latest else None

        rows, summary = self.build_display_rows(code, latest_nav)
        sorted_rows = self.apply_row_sort(rows)
        for idx, row in enumerate(sorted_rows):
            tags = ("row_even",) if idx % 2 == 0 else ("row_odd",)
            self.tx_tree.insert(
                "",
                tk.END,
                iid=row["uid"],
                tags=tags,
                values=(row["date"], row["shares"], row["nav"], row["rate"], row["profit"], row["amount"]),
            )
            self.rate_cell_meta[str(row["uid"])] = row.get("rate_sort")

        if latest:
            self.summary_latest_var.set(f"{latest[1]:.4f}")
        else:
            self.summary_latest_var.set("-")

        self.summary_shares_var.set(f"{summary['shares']:.2f}")
        self.summary_cost_var.set(f"{summary['cost']:.2f}")
        self.summary_profit_var.set(f"{summary['profit']:.2f}")
        self.summary_rate_var.set(f"{summary['rate'] * 100:.2f}%")
        self._set_rate_bg_by_value(summary["rate"])
        self._queue_rate_overlay_refresh()
        self.after_idle(self._queue_rate_overlay_refresh)
        self.refresh_analysis_view()

    def _build_bond_display_rows(self, fetch_latest: bool = False) -> Tuple[List[Dict], Dict[str, float], List[str]]:
        grouped: Dict[str, Dict[str, float]] = {}
        for txn in self.bond_transactions:
            code = normalize_fund_code(txn.get("fund_code"))
            if not code:
                continue
            shares = to_float(txn.get("shares"), default=None)
            nav = to_float(txn.get("nav"), default=None)
            if shares is None or shares <= 0:
                continue
            if nav is None or nav <= 0:
                continue
            name = str(txn.get("fund_name", "") or "").strip() or str(self.fund_name_cache.get(code, "") or "").strip() or code
            self.fund_name_cache[code] = name
            info = grouped.setdefault(code, {"shares": 0.0, "cost": 0.0, "fund_name": name})
            info["shares"] += float(shares)
            info["cost"] += float(shares) * float(nav)
            if name:
                info["fund_name"] = name

        nav_errors: List[str] = []
        if fetch_latest:
            for code in sorted(grouped.keys()):
                try:
                    nav_date, latest_nav = fetch_latest_nav(code, self.session)
                    self.latest_nav_cache[code] = (nav_date, latest_nav)
                except Exception as exc:
                    nav_errors.append(f"{code}: {exc}")

        rows: List[Dict] = []
        total_cost = 0.0
        total_market = 0.0
        for code in sorted(grouped.keys()):
            shares = float(grouped[code]["shares"])
            cost = float(grouped[code]["cost"])
            avg_nav = (cost / shares) if shares > 0 else 0.0
            latest = self.get_latest_nav_cached(code)
            latest_nav = float(latest[1]) if latest and latest[1] is not None else None
            if latest_nav is not None and cost > 0:
                market = shares * latest_nav
                profit = market - cost
                rate = profit / cost
                rate_text = f"{rate * 100:.2f}%"
                profit_text = f"{profit:.2f}"
                rate_sort = rate
                profit_sort = profit
            else:
                market = cost
                rate_text = "-"
                profit_text = "-"
                rate_sort = None
                profit_sort = None
            amount_text = f"{market:.2f}"

            total_cost += cost
            total_market += market
            rows.append(
                {
                    "iid": code,
                    "fund_name": str(grouped[code].get("fund_name", "") or code),
                    "shares": f"{shares:.2f}",
                    "nav": f"{avg_nav:.4f}",
                    "rate": rate_text,
                    "profit": profit_text,
                    "amount": amount_text,
                    "cost": cost,
                    "rate_sort": rate_sort,
                    "profit_sort": profit_sort,
                    "amount_sort": market,
                }
            )

        total_profit = total_market - total_cost
        summary = {
            "cost": total_cost,
            "market": total_market,
            "profit": total_profit,
            "rate": (total_profit / total_cost) if total_cost > 0 else 0.0,
        }
        return rows, summary, nav_errors

    def _compute_stock_portfolio_details(self) -> List[Dict]:
        by_fund: Dict[str, List[Dict]] = {}
        for txn in self.transactions:
            code = normalize_fund_code(txn.get("fund_code"))
            if not code:
                continue
            by_fund.setdefault(code, []).append(txn)

        rows: List[Dict] = []
        for code, txns in sorted(by_fund.items()):
            lots, _, _ = build_positions(sorted(txns, key=lambda x: (x["date"], x["uid"])))
            shares = sum(float(x["remaining_shares"]) for x in lots)
            cost = sum(float(x["remaining_shares"]) * float(x["buy_nav"]) for x in lots)
            latest = self.get_latest_nav_cached(code)
            if (latest is None or latest[1] is None) and shares > 0:
                try:
                    nav_date, latest_nav = fetch_latest_nav(code, self.session)
                    self.latest_nav_cache[code] = (nav_date, latest_nav)
                    latest = (nav_date, latest_nav)
                except Exception:
                    latest = None
            if latest and latest[1] is not None:
                latest_nav = float(latest[1])
                market = shares * latest_nav
                nav_text = f"{latest_nav:.4f}"
            else:
                latest_nav = None
                market = cost
                nav_text = "-"

            rows.append(
                {
                    "code": code,
                    "shares": shares,
                    "latest_nav": latest_nav,
                    "latest_nav_text": nav_text,
                    "cost": cost,
                    "market": market,
                }
            )
        return rows

    def _compute_stock_portfolio_summary(self) -> Dict[str, float]:
        details = self._compute_stock_portfolio_details()
        total_cost = sum(float(x["cost"]) for x in details)
        total_market = sum(float(x["market"]) for x in details)

        total_profit = total_market - total_cost
        return {
            "cost": total_cost,
            "market": total_market,
            "profit": total_profit,
            "rate": (total_profit / total_cost) if total_cost > 0 else 0.0,
        }

    def show_stock_amount_details(self) -> None:
        details = self._compute_stock_portfolio_details()
        if not details:
            messagebox.showinfo("\u80a1\u7968\u57fa\u91d1\u660e\u7ec6", "\u5f53\u524d\u65e0\u53ef\u5bf9\u8d26\u7684\u80a1\u7968\u57fa\u91d1\u6301\u4ed3")
            return

        dialog = tk.Toplevel(self)
        dialog.title("\u80a1\u7968\u57fa\u91d1\u91d1\u989d\u660e\u7ec6")
        dialog.transient(self)
        dialog.geometry("620x420")
        dialog.minsize(560, 320)

        wrap = ttk.Frame(dialog, style="App.TFrame", padding=8)
        wrap.pack(fill=tk.BOTH, expand=True)
        wrap.columnconfigure(0, weight=1)
        wrap.rowconfigure(1, weight=1)

        total_market = sum(float(x["market"]) for x in details)
        ttk.Label(
            wrap,
            text=f"\u603b\u91d1\u989d\u5408\u8ba1: {total_market:.2f}",
            style="Body.TLabel",
        ).grid(row=0, column=0, sticky="w", pady=(0, 6))

        table_wrap = ttk.Frame(wrap, style="Card.TFrame")
        table_wrap.grid(row=1, column=0, sticky="nsew")
        table_wrap.columnconfigure(0, weight=1)
        table_wrap.rowconfigure(0, weight=1)

        cols = ("code", "shares", "latest_nav", "amount")
        tree = ttk.Treeview(table_wrap, columns=cols, show="headings", style="Data.Treeview")
        tree.heading("code", text="\u57fa\u91d1\u4ee3\u7801")
        tree.heading("shares", text="\u4efd\u989d")
        tree.heading("latest_nav", text="\u6700\u65b0\u51c0\u503c")
        tree.heading("amount", text="\u91d1\u989d")
        tree.column("code", width=120, anchor="center", stretch=True)
        tree.column("shares", width=120, anchor="center", stretch=True)
        tree.column("latest_nav", width=120, anchor="center", stretch=True)
        tree.column("amount", width=140, anchor="center", stretch=True)
        tree.grid(row=0, column=0, sticky="nsew")

        scroll = ttk.Scrollbar(table_wrap, orient=tk.VERTICAL, command=tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=scroll.set)

        for idx, row in enumerate(details):
            tags = ("row_even",) if idx % 2 == 0 else ("row_odd",)
            tree.insert(
                "",
                tk.END,
                tags=tags,
                values=(
                    row["code"],
                    f"{float(row['shares']):.2f}",
                    row["latest_nav_text"],
                    f"{float(row['market']):.2f}",
                ),
            )

    def _compute_bond_portfolio_summary(self) -> Dict[str, float]:
        _, summary, _ = self._build_bond_display_rows(fetch_latest=False)
        return summary

    def _collect_analysis_codes(self) -> List[str]:
        codes = set(self.get_all_fund_codes())
        for txn in self.bond_transactions:
            code = normalize_fund_code(txn.get("fund_code"))
            if code:
                codes.add(code)
        return sorted(codes)

    def _refresh_latest_nav_for_codes(self, codes: List[str]) -> List[str]:
        errors: List[str] = []
        for code in codes:
            try:
                nav_date, latest_nav = fetch_latest_nav(code, self.session)
                self.latest_nav_cache[code] = (nav_date, latest_nav)
            except Exception as exc:
                errors.append(f"{code}: {exc}")
        return errors

    def refresh_bond_view(self, fetch_latest: bool = False) -> None:
        if not hasattr(self, "bond_tree") or self.bond_tree is None:
            return
        rows, _, nav_errors = self._build_bond_display_rows(fetch_latest=fetch_latest)
        self._refresh_bond_tree_headings()
        sorted_rows = self.apply_bond_row_sort(rows)
        for item in self.bond_tree.get_children():
            self.bond_tree.delete(item)
        for idx, row in enumerate(sorted_rows):
            tags = ("row_even",) if idx % 2 == 0 else ("row_odd",)
            self.bond_tree.insert(
                "",
                tk.END,
                iid=row["iid"],
                tags=tags,
                values=(row["fund_name"], row["shares"], row["nav"], row["rate"], row["profit"], row["amount"]),
            )
        if fetch_latest and nav_errors:
            messagebox.showwarning("\u51c0\u503c\u66f4\u65b0\u63d0\u793a", "\u90e8\u5206\u503a\u5377\u57fa\u91d1\u51c0\u503c\u66f4\u65b0\u5931\u8d25\uff1a\n" + "\n".join(nav_errors[:3]))
        self.refresh_analysis_view()

    def refresh_analysis_view(self, fetch_latest: bool = False) -> None:
        if not hasattr(self, "analysis_stock_amount_var") or not hasattr(self, "analysis_bond_amount_var"):
            return
        if fetch_latest:
            self._refresh_latest_nav_for_codes(self._collect_analysis_codes())
        stock = self._compute_stock_portfolio_summary()
        bond = self._compute_bond_portfolio_summary()
        total_market = float(stock.get("market", 0.0)) + float(bond.get("market", 0.0))
        stock_ratio = (float(stock.get("market", 0.0)) / total_market) if total_market > 0 else 0.0
        bond_ratio = (float(bond.get("market", 0.0)) / total_market) if total_market > 0 else 0.0

        self.analysis_stock_amount_var.set(f"{float(stock.get('market', 0.0)):.2f}")
        self.analysis_stock_rate_var.set(f"{float(stock['rate']) * 100:.2f}%")
        self.analysis_stock_profit_var.set(f"{float(stock['profit']):.2f}")
        self.analysis_stock_ratio_var.set(f"{stock_ratio * 100:.2f}%")

        self.analysis_bond_amount_var.set(f"{float(bond.get('market', 0.0)):.2f}")
        self.analysis_bond_rate_var.set(f"{float(bond['rate']) * 100:.2f}%")
        self.analysis_bond_profit_var.set(f"{float(bond['profit']):.2f}")
        self.analysis_bond_ratio_var.set(f"{bond_ratio * 100:.2f}%")

    def refresh_fund_profit(self) -> None:
        self.on_fund_selected()

    def on_fund_selected(self) -> None:
        code = self._resolve_selected_fund_code()
        if not code:
            self.refresh_fund_view()
            return

        # Show cached data immediately, then refresh NAV.
        self.refresh_fund_view()
        try:
            nav_date, latest_nav = fetch_latest_nav(code, self.session)
            self.latest_nav_cache[code] = (nav_date, latest_nav)
            self.refresh_fund_view()
            self._save_trade_doc()
        except Exception:
            messagebox.showerror("\u5237\u65b0\u5931\u8d25", "\u83b7\u53d6\u4e0d\u5230\u6700\u65b0\u57fa\u91d1\u51c0\u503c")

    def _run_startup_sync_and_notify_once(self) -> None:
        if self._startup_notify_done:
            return
        self._startup_notify_done = True

        fund_codes = list(self.get_all_fund_codes())
        if not fund_codes:
            return

        url_text = self.vars["app.ntfy_url"].get().strip() or str(self.config.get("app", {}).get("ntfy_url", "")).strip()
        ntfy_url = normalize_ntfy_url(url_text)

        threshold = to_float(self.vars["app.threshold"].get().strip(), default=None)
        if threshold is None:
            threshold = to_float(self.config.get("app", {}).get("threshold", 0.40), default=0.40)
        threshold = float(threshold if threshold is not None else 0.40)

        t = threading.Thread(
            target=self._startup_sync_worker,
            args=(fund_codes, ntfy_url, threshold),
            daemon=True,
        )
        t.start()

    def _startup_sync_worker(self, fund_codes: List[str], ntfy_url: str, threshold: float) -> None:
        nav_map: Dict[str, Tuple[str, float]] = {}
        nav_errors: List[str] = []
        with requests.Session() as sess:
            for code in fund_codes:
                try:
                    nav_date, latest_nav = fetch_latest_nav(code, sess)
                    nav_map[code] = (nav_date, latest_nav)
                except Exception as exc:
                    nav_errors.append(f"{code}: {exc}")
        self.after(0, lambda: self._startup_sync_apply(nav_map, nav_errors, ntfy_url, threshold))

    def _startup_sync_apply(self, nav_map: Dict[str, Tuple[str, float]], nav_errors: List[str], ntfy_url: str, threshold: float) -> None:
        if nav_map:
            self.latest_nav_cache.update(nav_map)
            self.refresh_fund_view()
            self._save_trade_doc()

        if ntfy_url:
            notify_rows: List[Tuple[str, Dict, float, str]] = []
            for code, (nav_date, latest_nav) in nav_map.items():
                rows, _ = self.build_display_rows(code, latest_nav)
                for row in rows:
                    rate_val = row.get("rate_sort")
                    if rate_val is None or float(rate_val) < threshold:
                        continue
                    notify_rows.append((code, row, latest_nav, nav_date))
            if notify_rows:
                t = threading.Thread(
                    target=self._startup_notify_worker,
                    args=(ntfy_url, threshold, notify_rows, nav_errors),
                    daemon=True,
                )
                t.start()
                return

        if nav_errors:
            messagebox.showwarning("\u51c0\u503c\u66f4\u65b0\u63d0\u793a", "\u90e8\u5206\u57fa\u91d1\u51c0\u503c\u66f4\u65b0\u5931\u8d25\uff1a\n" + "\n".join(nav_errors[:3]))

    def _startup_notify_worker(
        self,
        ntfy_url: str,
        threshold: float,
        notify_rows: List[Tuple[str, Dict, float, str]],
        nav_errors: List[str],
    ) -> None:
        errors: List[str] = []
        sent_keys: List[str] = []
        for code, row, latest_nav, nav_date in notify_rows:
            alert_key = f"startup|{code}|{row['uid']}|{nav_date}|{latest_nav:.6f}|{threshold:.6f}"
            if alert_key in self.ntfy_sent_cache:
                continue
            fund_name = str(self.fund_name_cache.get(code, "") or "").strip() or code
            lines = [
                f"基金名称: {fund_name}",
                f"购买日期: {row['date']}",
                f"份额: {row['shares']}",
                f"成交净值: {row['nav']}",
                f"最新净值: {latest_nav:.4f}",
                f"盈利利率: {row['rate']}",
                f"盈利额: {row['profit']}",
            ]
            try:
                send_ntfy(ntfy_url, title="基金收益提醒", message="\n".join(lines), tags="moneybag,rotating_light")
                sent_keys.append(alert_key)
            except Exception as exc:
                errors.append(str(exc))
        self.after(0, lambda: self._startup_notify_finish(sent_keys, errors, nav_errors))

    def _startup_notify_finish(self, sent_keys: List[str], errors: List[str], nav_errors: List[str]) -> None:
        for key in sent_keys:
            self.ntfy_sent_cache.add(key)
        if errors:
            messagebox.showerror("\u901a\u77e5\u5931\u8d25", f"ntfy \u53d1\u9001\u5931\u8d25\uff1a\n{errors[0]}")
        elif nav_errors:
            messagebox.showwarning("\u51c0\u503c\u66f4\u65b0\u63d0\u793a", "\u90e8\u5206\u57fa\u91d1\u51c0\u503c\u66f4\u65b0\u5931\u8d25\uff1a\n" + "\n".join(nav_errors[:3]))

    def add_fund(self) -> None:
        raw = simpledialog.askstring("\u65b0\u589e\u57fa\u91d1", "\u8f93\u51656\u4f4d\u57fa\u91d1\u4ee3\u7801:")
        if raw is None:
            return
        code = normalize_fund_code(raw)
        if not code:
            messagebox.showwarning("\u63d0\u793a", "\u57fa\u91d1\u4ee3\u7801\u65e0\u6548")
            return

        funds = set(self.config.get("funds", []))
        funds.add(code)
        self.config["funds"] = sorted(funds)
        if code not in self.fund_name_cache:
            try:
                self.fund_name_cache[code] = fetch_fund_name(code, self.session)
            except Exception:
                self.fund_name_cache[code] = ""

        self._persist_fund_changes()
        self.refresh_fund_list(keep_current=False)
        self._set_selected_fund_code(code)
        self.refresh_fund_view()

    @staticmethod
    def _validated_date(raw: str) -> Optional[str]:
        date_str = parse_date(raw)
        if not date_str:
            return None
        try:
            dt.date.fromisoformat(date_str)
            return date_str
        except ValueError:
            return None

    def _save_trade_doc(self) -> bool:
        try:
            current = self.vars["app.xlsx_path"].get().strip() or self.config.get("app", {}).get("xlsx_path", "data/jijin.xlsx")
            self.config.setdefault("app", {})
            self.config["app"]["xlsx_path"] = current
            latest_nav_map = {code: val[1] for code, val in self.latest_nav_cache.items() if val and len(val) >= 2}
            fund_name_map = {normalize_fund_code(code): name for code, name in self.fund_name_cache.items() if normalize_fund_code(code)}
            save_transactions(Path(current), self.transactions, latest_nav_map=latest_nav_map, fund_name_map=fund_name_map)
            save_bond_transactions(Path(current), self.bond_transactions)
            self.sheet_fund_codes = sorted({normalize_fund_code(txn["fund_code"]) for txn in self.transactions if normalize_fund_code(txn["fund_code"])})
            return True
        except Exception as exc:
            messagebox.showerror("\u4fdd\u5b58\u5931\u8d25", f"\u4ea4\u6613\u8bb0\u5f55\u6587\u4ef6\u4fdd\u5b58\u5931\u8d25\uff1a\n{exc}")
            return False

    def _save_runtime_config(self) -> bool:
        try:
            self.config.setdefault("app", {})
            current = self.vars["app.xlsx_path"].get().strip() or self.config.get("app", {}).get("xlsx_path", "data/jijin.xlsx")
            self.config["app"]["xlsx_path"] = current
            self.config["funds"] = sorted(set(self.get_all_fund_codes()))
            save_config(self.config_path, self.config)
            self.saved_app_state = copy.deepcopy(self.config.get("app", {}))
            return True
        except Exception as exc:
            messagebox.showerror("\u4fdd\u5b58\u5931\u8d25", f"\u914d\u7f6e\u4fdd\u5b58\u5931\u8d25\uff1a\n{exc}")
            return False

    def _persist_fund_changes(self) -> bool:
        ok_xlsx = self._save_trade_doc()
        ok_cfg = self._save_runtime_config()
        return ok_xlsx and ok_cfg

    def _center_dialog(self, dialog: tk.Toplevel) -> None:
        self.update_idletasks()
        dialog.update_idletasks()

        parent_x = self.winfo_rootx()
        parent_y = self.winfo_rooty()
        parent_w = self.winfo_width()
        parent_h = self.winfo_height()
        dialog_w = dialog.winfo_reqwidth()
        dialog_h = dialog.winfo_reqheight()

        pos_x = parent_x + max((parent_w - dialog_w) // 2, 0)
        pos_y = parent_y + max((parent_h - dialog_h) // 2, 0)
        dialog.geometry(f"+{pos_x}+{pos_y}")

    def _prompt_buy_dialog(
        self,
        title: str = "\u65b0\u589e\u4e70\u5165",
        initial_date: Optional[str] = None,
        initial_shares: Optional[float] = None,
        initial_nav: Optional[float] = None,
    ) -> Optional[Dict]:
        dialog = tk.Toplevel(self)
        dialog.title(title)
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)

        date_var = tk.StringVar(value=initial_date or dt.date.today().isoformat())
        shares_var = tk.StringVar(value="" if initial_shares is None else f"{initial_shares:.2f}")
        nav_var = tk.StringVar(value="" if initial_nav is None else f"{initial_nav:.4f}")

        body = ttk.Frame(dialog, padding=12)
        body.grid(row=0, column=0, sticky="nsew")

        ttk.Label(body, text="\u4e70\u5165\u65e5\u671f (YYYY-MM-DD)").grid(row=0, column=0, sticky="w")
        ttk.Entry(body, textvariable=date_var, width=24).grid(row=0, column=1, sticky="ew", padx=(8, 0))
        ttk.Label(body, text="\u4efd\u989d").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(body, textvariable=shares_var, width=24).grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=(8, 0))
        ttk.Label(body, text="\u6210\u4ea4\u51c0\u503c").grid(row=2, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(body, textvariable=nav_var, width=24).grid(row=2, column=1, sticky="ew", padx=(8, 0), pady=(8, 0))

        result: Dict[str, float] = {}

        def on_ok() -> None:
            date_str = self._validated_date(date_var.get().strip())
            if not date_str:
                messagebox.showwarning("\u63d0\u793a", "\u65e5\u671f\u683c\u5f0f\u65e0\u6548")
                return

            shares = to_float(shares_var.get(), default=None)
            if shares is None or shares <= 0:
                messagebox.showwarning("\u63d0\u793a", "\u4efd\u989d\u5fc5\u987b\u5927\u4e8e 0")
                return

            nav = to_float(nav_var.get(), default=None)
            if nav is None or nav <= 0:
                messagebox.showwarning("\u63d0\u793a", "\u6210\u4ea4\u51c0\u503c\u5fc5\u987b\u5927\u4e8e 0")
                return

            result["date"] = date_str
            result["shares"] = float(shares)
            result["nav"] = float(nav)
            dialog.destroy()

        def on_cancel() -> None:
            dialog.destroy()

        btns = ttk.Frame(body)
        btns.grid(row=3, column=0, columnspan=2, sticky="e", pady=(12, 0))
        ttk.Button(btns, text="\u786e\u5b9a", command=on_ok).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btns, text="\u53d6\u6d88", command=on_cancel).grid(row=0, column=1)

        dialog.bind("<Return>", lambda _: on_ok())
        dialog.bind("<Escape>", lambda _: on_cancel())
        self._center_dialog(dialog)
        dialog.wait_window()
        return result or None

    def _prompt_bond_buy_dialog(
        self,
        title: str = "\u65b0\u589e\u503a\u5377\u57fa\u91d1",
        initial_code: str = "",
        initial_shares: Optional[float] = None,
        initial_nav: Optional[float] = None,
    ) -> Optional[Dict]:
        dialog = tk.Toplevel(self)
        dialog.title(title)
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)

        code_var = tk.StringVar(value=initial_code)
        shares_var = tk.StringVar(value="" if initial_shares is None else f"{initial_shares:.2f}")
        nav_var = tk.StringVar(value="" if initial_nav is None else f"{initial_nav:.4f}")

        body = ttk.Frame(dialog, padding=12)
        body.grid(row=0, column=0, sticky="nsew")

        ttk.Label(body, text="\u57fa\u91d1\u4ee3\u7801").grid(row=0, column=0, sticky="w")
        ttk.Entry(body, textvariable=code_var, width=24).grid(row=0, column=1, sticky="ew", padx=(8, 0))
        ttk.Label(body, text="\u4efd\u989d").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(body, textvariable=shares_var, width=24).grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=(8, 0))
        ttk.Label(body, text="\u6210\u4ea4\u51c0\u503c").grid(row=2, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(body, textvariable=nav_var, width=24).grid(row=2, column=1, sticky="ew", padx=(8, 0), pady=(8, 0))

        result: Dict[str, float] = {}

        def on_ok() -> None:
            code = normalize_fund_code(code_var.get().strip())
            if not code:
                messagebox.showwarning("\u63d0\u793a", "\u57fa\u91d1\u4ee3\u7801\u65e0\u6548")
                return

            shares = to_float(shares_var.get(), default=None)
            if shares is None or shares <= 0:
                messagebox.showwarning("\u63d0\u793a", "\u4efd\u989d\u5fc5\u987b\u5927\u4e8e 0")
                return

            nav = to_float(nav_var.get(), default=None)
            if nav is None or nav <= 0:
                messagebox.showwarning("\u63d0\u793a", "\u6210\u4ea4\u51c0\u503c\u5fc5\u987b\u5927\u4e8e 0")
                return

            result["fund_code"] = code
            result["shares"] = float(shares)
            result["nav"] = float(nav)
            dialog.destroy()

        def on_cancel() -> None:
            dialog.destroy()

        btns = ttk.Frame(body)
        btns.grid(row=3, column=0, columnspan=2, sticky="e", pady=(12, 0))
        ttk.Button(btns, text="\u786e\u5b9a", command=on_ok).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btns, text="\u53d6\u6d88", command=on_cancel).grid(row=0, column=1)

        dialog.bind("<Return>", lambda _: on_ok())
        dialog.bind("<Escape>", lambda _: on_cancel())
        self._center_dialog(dialog)
        dialog.wait_window()
        return result or None

    def _find_txn_by_uid(self, uid: int) -> Optional[Dict]:
        for txn in self.transactions:
            if int(txn["uid"]) == uid:
                return txn
        return None

    def on_tree_right_click(self, event) -> None:
        row_id = self.tx_tree.identify_row(event.y)
        if not row_id:
            return
        if row_id not in self.tx_tree.selection():
            self.tx_tree.selection_set(row_id)
        self.tx_tree.focus(row_id)
        try:
            self.tx_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.tx_menu.grab_release()

    def on_bond_tree_right_click(self, event) -> None:
        if not hasattr(self, "bond_tree") or self.bond_tree is None:
            return
        row_id = self.bond_tree.identify_row(event.y)
        if not row_id:
            return
        if row_id not in self.bond_tree.selection():
            self.bond_tree.selection_set(row_id)
        self.bond_tree.focus(row_id)
        try:
            self.bond_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.bond_menu.grab_release()

    def add_buy(self) -> None:
        code = self._resolve_selected_fund_code()
        if not code:
            messagebox.showwarning("\u63d0\u793a", "\u8bf7\u5148\u9009\u62e9\u57fa\u91d1")
            return

        form = self._prompt_buy_dialog(title="\u65b0\u589e\u4e70\u5165")
        if form is None:
            return

        txn = {
            "uid": self.next_uid,
            "date": form["date"],
            "fund_code": code,
            "action": "BUY",
            "shares": form["shares"],
            "nav": form["nav"],
            "note": "",
        }
        self.next_uid += 1
        self.transactions.append(txn)
        self.transactions.sort(key=lambda x: (x["date"], x["uid"]))

        funds = set(self.config.get("funds", []))
        funds.add(txn["fund_code"])
        self.config["funds"] = sorted(funds)

        self._persist_fund_changes()
        self.refresh_fund_list(keep_current=True)
        self.refresh_fund_view()

    def edit_selected_record(self) -> None:
        selected = self.tx_tree.selection()
        if len(selected) != 1:
            messagebox.showwarning("\u63d0\u793a", "\u8bf7\u53ea\u9009\u62e9\u4e00\u6761\u8bb0\u5f55\u8fdb\u884c\u4fee\u6539")
            return

        uid = int(selected[0])
        txn = self._find_txn_by_uid(uid)
        if txn is None:
            messagebox.showwarning("\u63d0\u793a", "\u672a\u627e\u5230\u5bf9\u5e94\u8bb0\u5f55")
            return

        form = self._prompt_buy_dialog(
            title="\u4fee\u6539\u8bb0\u5f55",
            initial_date=txn.get("date", dt.date.today().isoformat()),
            initial_shares=float(txn.get("shares", 0.0)),
            initial_nav=to_float(txn.get("nav"), default=None),
        )
        if form is None:
            return

        txn["date"] = form["date"]
        txn["shares"] = form["shares"]
        txn["nav"] = form["nav"]
        self.transactions.sort(key=lambda x: (x["date"], x["uid"]))
        self._save_trade_doc()
        self.refresh_fund_view()

    def _remove_selected_records(self, action_text: str) -> None:
        selected = self.tx_tree.selection()
        if not selected:
            messagebox.showwarning("\u63d0\u793a", f"\u8bf7\u5148\u9009\u62e9\u8981{action_text}\u7684\u8bb0\u5f55")
            return

        if action_text == "\u5356\u51fa":
            confirm_text = f"\u786e\u5b9a\u5356\u51fa\u9009\u4e2d\u7684 {len(selected)} \u6761\u8bb0\u5f55\u5417\uff1f\n\uff08\u5356\u51fa\u7b49\u540c\u4e8e\u5220\u9664\u8be5\u6761\u4e70\u5165\u8bb0\u5f55\uff09"
        else:
            confirm_text = f"\u786e\u5b9a\u5220\u9664\u9009\u4e2d\u7684 {len(selected)} \u6761\u8bb0\u5f55\u5417\uff1f"
        if not messagebox.askyesno("\u786e\u8ba4", confirm_text):
            return

        to_delete = {int(iid) for iid in selected}
        self.transactions = [x for x in self.transactions if x["uid"] not in to_delete]
        self._persist_fund_changes()
        self.refresh_fund_list(keep_current=True)
        self.refresh_fund_view()

    def add_bond_buy(self) -> None:
        form = self._prompt_bond_buy_dialog(title="\u65b0\u589e\u503a\u5377\u57fa\u91d1")
        if form is None:
            return

        code = normalize_fund_code(form["fund_code"])
        if not code:
            return
        fund_name = str(self.fund_name_cache.get(code, "") or "").strip()
        if not fund_name:
            try:
                fund_name = fetch_fund_name(code, self.session)
            except Exception:
                fund_name = code
        self.fund_name_cache[code] = fund_name

        txn = {
            "uid": self.bond_next_uid,
            "fund_code": code,
            "fund_name": fund_name,
            "shares": float(form["shares"]),
            "nav": float(form["nav"]),
            "note": "",
        }
        self.bond_next_uid += 1
        self.bond_transactions.append(txn)
        self.bond_transactions.sort(key=lambda x: int(x["uid"]))

        try:
            nav_date, latest_nav = fetch_latest_nav(code, self.session)
            self.latest_nav_cache[code] = (nav_date, latest_nav)
        except Exception:
            pass

        self._save_trade_doc()
        self.refresh_bond_view(fetch_latest=False)

    def _bond_code_aggregate(self, code: str) -> Optional[Tuple[float, float]]:
        code = normalize_fund_code(code)
        if not code:
            return None
        total_shares = 0.0
        total_cost = 0.0
        for txn in self.bond_transactions:
            if normalize_fund_code(txn.get("fund_code")) != code:
                continue
            shares = to_float(txn.get("shares"), default=None)
            nav = to_float(txn.get("nav"), default=None)
            if shares is None or shares <= 0:
                continue
            if nav is None or nav <= 0:
                continue
            total_shares += float(shares)
            total_cost += float(shares) * float(nav)
        if total_shares <= 0:
            return None
        return total_shares, (total_cost / total_shares)

    def edit_selected_bond_record(self) -> None:
        if not hasattr(self, "bond_tree") or self.bond_tree is None:
            return
        selected = self.bond_tree.selection()
        if len(selected) != 1:
            messagebox.showwarning("\u63d0\u793a", "\u8bf7\u53ea\u9009\u62e9\u4e00\u6761\u8bb0\u5f55\u8fdb\u884c\u4fee\u6539")
            return

        old_code = normalize_fund_code(selected[0])
        if not old_code:
            messagebox.showwarning("\u63d0\u793a", "\u672a\u627e\u5230\u5bf9\u5e94\u503a\u5377\u57fa\u91d1")
            return
        agg = self._bond_code_aggregate(old_code)
        if agg is None:
            messagebox.showwarning("\u63d0\u793a", "\u672a\u627e\u5230\u53ef\u4fee\u6539\u7684\u503a\u5377\u57fa\u91d1\u8bb0\u5f55")
            return
        old_shares, old_nav = agg

        form = self._prompt_bond_buy_dialog(
            title="\u4fee\u6539\u503a\u5377\u57fa\u91d1\u8bb0\u5f55",
            initial_code=old_code,
            initial_shares=old_shares,
            initial_nav=old_nav,
        )
        if form is None:
            return

        new_code = normalize_fund_code(form["fund_code"])
        if not new_code:
            return
        fund_name = str(self.fund_name_cache.get(new_code, "") or "").strip()
        if not fund_name:
            try:
                fund_name = fetch_fund_name(new_code, self.session)
            except Exception:
                fund_name = new_code
        self.fund_name_cache[new_code] = fund_name

        self.bond_transactions = [x for x in self.bond_transactions if normalize_fund_code(x.get("fund_code")) != old_code]
        self.bond_transactions.append(
            {
                "uid": self.bond_next_uid,
                "fund_code": new_code,
                "fund_name": fund_name,
                "shares": float(form["shares"]),
                "nav": float(form["nav"]),
                "note": "",
            }
        )
        self.bond_next_uid += 1
        self.bond_transactions.sort(key=lambda x: int(x["uid"]))

        try:
            nav_date, latest_nav = fetch_latest_nav(new_code, self.session)
            self.latest_nav_cache[new_code] = (nav_date, latest_nav)
        except Exception:
            pass

        self._save_trade_doc()
        self.refresh_bond_view(fetch_latest=False)

    def sell_selected_records(self) -> None:
        self._remove_selected_records("\u5356\u51fa")

    def delete_selected_records(self) -> None:
        self._remove_selected_records("\u5220\u9664")

    def _remove_selected_bond_records(self, action_text: str) -> None:
        if not hasattr(self, "bond_tree") or self.bond_tree is None:
            return
        selected = self.bond_tree.selection()
        if not selected:
            messagebox.showwarning("\u63d0\u793a", f"\u8bf7\u5148\u9009\u62e9\u8981{action_text}\u7684\u8bb0\u5f55")
            return

        if action_text == "\u5356\u51fa":
            confirm_text = f"\u786e\u5b9a\u5356\u51fa\u9009\u4e2d\u7684 {len(selected)} \u4e2a\u57fa\u91d1\u5417\uff1f\n\uff08\u5356\u51fa\u7b49\u540c\u4e8e\u5220\u9664\u8be5\u57fa\u91d1\u7684\u5168\u90e8\u4e70\u5165\u8bb0\u5f55\uff09"
        else:
            confirm_text = f"\u786e\u5b9a\u5220\u9664\u9009\u4e2d\u7684 {len(selected)} \u4e2a\u57fa\u91d1\u8bb0\u5f55\u5417\uff1f"
        if not messagebox.askyesno("\u786e\u8ba4", confirm_text):
            return

        codes = {normalize_fund_code(iid) for iid in selected if normalize_fund_code(iid)}
        self.bond_transactions = [x for x in self.bond_transactions if normalize_fund_code(x.get("fund_code")) not in codes]
        self._save_trade_doc()
        self.refresh_bond_view(fetch_latest=False)

    def sell_selected_bond_records(self) -> None:
        self._remove_selected_bond_records("\u5356\u51fa")

    def delete_selected_bond_records(self) -> None:
        self._remove_selected_bond_records("\u5220\u9664")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="config/funds.yaml", help="config file path")
    args = parser.parse_args()

    app = JijinUI(config_path=Path(args.config))
    app.mainloop()


if __name__ == "__main__":
    main()


